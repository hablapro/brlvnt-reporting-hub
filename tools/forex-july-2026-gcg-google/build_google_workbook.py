"""
Build GCG (US Hispanic) Google Ads July 2026 data workbook.
Source: Google Ads MCP (run_gaql_query), customer 4781995752, pulled 2026-08-19.
Date range 2026-07-01..2026-07-31. Reporting only.
House style matches tools/forex-july-2026-ggmi-meta/build_meta_workbook.py.

Counting rules (see Notes & QA tab):
- Submitted applications = "PO App Form - Step 5 - Submission Completed" ONLY.
- Search campaigns carry Step 5 in metrics.conversions; July also counts 3
  offline GCLID events (2 approved, 1 funded) as primary conversions, so
  metrics.conversions (76) OVERSTATES submitted apps (73). New vs June.
- PMax carries Step 5 in all_conversions only (49); metrics.conversions = 0
  because the campaign's goal configuration excludes it. View-through = 1,
  so the 49 are click-based.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
SUBHEAD_FILL = "D9E1F2"
SECTION_FILL = "2E5496"
ROW_ALT_FILL = "F2F5FB"
GRAY_TEXT = "404040"
BORDER_COLOR = "BFBFBF"

thin = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def title_row(ws, text, row=1, span=9):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)


def subtitle_row(ws, text, row=2, span=9):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
    c.fill = PatternFill("solid", fgColor=SUBHEAD_FILL)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SUBHEAD_FILL)


def section_row(ws, text, row, span=9):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SECTION_FILL)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SECTION_FILL)


def bullet_row(ws, text, row, span=9):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30


def header_row(ws, headers, row):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def data_row(ws, values, row, formats=None, alt=False, total=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if total:
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=SECTION_FILL)
        else:
            c.font = Font(name="Calibri", size=10)
            if alt:
                c.fill = PatternFill("solid", fgColor=ROW_ALT_FILL)
        if formats and formats.get(i):
            c.number_format = formats[i]


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


MONEY = "#,##0.00"
INT = "#,##0"
PCT = "0.00%"

# Campaign data: (name, type, spend, impr, clicks, step5_subs, primary_conv_reported)
SEARCH = [
    ("GCG_US_Esp_NonBrand_TrackA_Trust_Google", "Search", 10944.16, 45604, 4668, 28, 31),
    ("GCG_US_Esp_NonBrand_TrackB_Authority_Google", "Search", 9728.08, 25373, 1848, 20, 20),
    ("GCG_US_Esp_NonBrand_TrackB_Platform_Google", "Search", 6383.33, 38410, 2285, 19, 19),
    ("GCG_US_Esp_Brand_Search_Google", "Search", 2422.57, 16005, 767, 6, 6),
]
PMAX = ("GCG_LeadPMax_q3_FY26", "PMax", 18174.60, 792667, 14009, 49, 0)

s_spend = sum(c[2] for c in SEARCH)
s_impr = sum(c[3] for c in SEARCH)
s_clicks = sum(c[4] for c in SEARCH)
s_subs = sum(c[5] for c in SEARCH)

# ---------------------------------------------------------------------------
# TAB 1: Summary
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Summary")
title_row(ws, "GCG (US Hispanic) — Google Ads — July 2026 Monthly Report")
subtitle_row(ws, "Customer 4781995752 · pulled 2026-08-19 via Google Ads API · MoM vs June · USD · submitted app = PO App Form Step 5")

r = 4
section_row(ws, "HEADLINE", r); r += 1
for t in [
    f"•  Two Google lines this month. Search: ${s_spend:,.2f}, {s_subs} submitted applications, ${s_spend/s_subs:,.2f} CPA. PMax (new, launched week of Jul 13): $18,174.60, 49 submitted applications, $370.91 CPA — in roughly 18 days of delivery.",
    "•  Search spend +30.9% MoM ($22,523.79 → $29,478.13); submitted apps 67 → 73 (+9.0%); CPA $336.18 → $403.81 (+20.1%). Spend still outgrew applications, but less violently than June (+48% spend / -12% apps).",
    "•  The June ad-rank program shows movement where it was applied: Trust impression share 27% → 32.6% and lost-to-rank 64-76% band → 58.7%; Trust went from the worst Search CPA ($433) to 28 submitted apps at $390.86, the volume leader.",
    "•  PMax launched mid-month and matched Search efficiency out of the gate ($370.91 vs $403.81) on the same Step 5 event, click-based (view-through = 1 of 49).",
    "•  Geo: 100.00% of July spend delivered in the United States (location-of-presence). Compliant.",
]:
    bullet_row(ws, t, r); r += 1

r += 1
section_row(ws, "WATCH ITEMS", r); r += 1
for t in [
    "•  PMax goal configuration: Step 5 is NOT a primary conversion goal on the PMax campaign (metrics.conversions = 0; the 49 sit in all_conversions). The campaign is not optimizing toward submitted applications. Ops action → recommendations file; account mechanics stay out of the client deck.",
    "•  Conversion column drift: July Search metrics.conversions (76) includes 3 offline GCLID events (2 approved, 1 funded) newly counted as primary. June's 67 was pure Step 5. Scorecard basis stays Step 5 only (73) — do not quote 76.",
    "•  Authority CPA $486.40 at $5.26 avg CPC (highest in account); 'broker forex usa' broad carries it: $4,233.86 / 15 subs / $282 CPA. The expensive terms around it dilute.",
]:
    bullet_row(ws, t, r); r += 1

set_widths(ws, [110] + [12] * 8)

# ---------------------------------------------------------------------------
# TAB 2: Campaigns
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Campaigns")
title_row(ws, "GCG Google Ads — July Campaigns")
subtitle_row(ws, "Submitted apps = Step 5 only. 'Primary conv (reported)' shows the platform column incl. GCLID offline events — internal reconciliation only.")

headers = ["Campaign", "Type", "Spend", "Impr", "Clicks", "CTR", "Avg CPC", "Submitted apps (Step 5)", "CPA (Step 5)"]
r = 4
header_row(ws, headers, r); r += 1
fmts = {3: MONEY, 4: INT, 5: INT, 6: PCT, 7: MONEY, 8: INT, 9: MONEY}
for i, c in enumerate(SEARCH):
    name, typ, spend, impr, clicks, subs, conv = c
    data_row(ws, [name, typ, spend, impr, clicks, clicks / impr, spend / clicks, subs, spend / subs],
             r, fmts, alt=i % 2 == 1)
    r += 1
name, typ, spend, impr, clicks, subs, conv = PMAX
data_row(ws, [name, typ, spend, impr, clicks, clicks / impr, spend / clicks, subs, spend / subs], r, fmts)
r += 1
data_row(ws, ["SEARCH SUBTOTAL", "", s_spend, s_impr, s_clicks, s_clicks / s_impr, s_spend / s_clicks, s_subs, s_spend / s_subs], r, fmts, total=True)
r += 1
data_row(ws, ["GOOGLE TOTAL (Search + PMax)", "", s_spend + PMAX[2], s_impr + PMAX[3], s_clicks + PMAX[4], "", "", s_subs + PMAX[5], (s_spend + PMAX[2]) / (s_subs + PMAX[5])], r, fmts, total=True)
r += 2
bullet_row(ws, "PMax launched week of 2026-07-13; weekly spend $7,448.86 / $5,596.51 / $5,129.23. Its CPA reflects a partial month.", r)

set_widths(ws, [44, 8, 12, 11, 9, 8, 9, 12, 11])

# ---------------------------------------------------------------------------
# TAB 3: Conversions
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Conversions")
title_row(ws, "GCG Google Ads — July Conversion Actions (all_conversions)")
subtitle_row(ws, "Full-funnel view. Steps 1-4 and Sitewide are secondary actions; Step 5 is the agency scorecard event; GCLID rows are client-journey outcomes (downstream, neutral framing).")

headers = ["Conversion action", "Trust", "Authority", "Platform", "Brand", "PMax", "Total"]
r = 4
header_row(ws, headers, r); r += 1
CONV = [
    ("Step 1 - Application Start", 409, 218, 279, 71, 517, 1494),
    ("Step 2 - Account Set-up", 108, 96, 145, 13, 270, 632),
    ("Step 3 - Platform & Product", 93, 75, 110, 14, 194, 486),
    ("Step 4.1 - eKYC Passed", 38, 21, 21, 6, 48, 134),
    ("Step 5 - Submission Completed", 28, 20, 19, 6, 49, 122),
    ("Step 6 - My Account", 2, 1, 2, 0, 3, 8),
    ("GCLID - Approved", 2, 2, 0, 0, 4, 8),
    ("GCLID - Funded", 1, 2, 0, 0, 4, 7),
    ("GCLID - Traded", 0, 1, 0, 0, 2, 3),
    ("Sitewide (secondary)", 8489, 3981, 5122, 1534, 15731, 34857),
]
ifmt = {i: INT for i in range(2, 8)}
for i, row_vals in enumerate(CONV):
    data_row(ws, list(row_vals), r, ifmt, alt=i % 2 == 1)
    r += 1
r += 1
bullet_row(ws, "Step 1 → Step 5 completion: Search 8.4% (73/977 excl. PMax), PMax 9.5% (49/517). GCLID offline rows exist only where the client's CRM matched a click; treat as directional.", r)

set_widths(ws, [34, 10, 10, 10, 10, 10, 10])

# ---------------------------------------------------------------------------
# TAB 4: Impression Share
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Impression Share")
title_row(ws, "GCG Google Ads — Search Auction Position (July vs June)")
subtitle_row(ws, "June baseline from the June cycle. Lost-to-rank remains the dominant ceiling; June ruling: the lever is ad rank, not budget.")

headers = ["Campaign", "Impr share Jul", "Impr share Jun", "Lost to rank Jul", "Lost to rank Jun", "Lost to budget Jul", "Avg CPC", "CTR"]
r = 4
header_row(ws, headers, r); r += 1
IS = [
    ("TrackA Trust", 0.3261, 0.27, 0.5873, "0.64-0.76 band", 0.0866, 2.34, 0.1024),
    ("TrackB Authority", 0.2860, "0.13-0.24 band", 0.6347, "0.64-0.76 band", 0.0793, 5.26, 0.0728),
    ("TrackB Platform", 0.1811, "0.13-0.24 band", 0.7029, "0.64-0.76 band", 0.1160, 2.79, 0.0595),
    ("Brand", 0.1195, 0.11, 0.7897, "0.64-0.76 band", 0.0909, 3.16, 0.0479),
]
f = {2: PCT, 4: PCT, 6: PCT, 7: MONEY, 8: PCT}
for i, row_vals in enumerate(IS):
    data_row(ws, list(row_vals), r, f, alt=i % 2 == 1)
    r += 1

set_widths(ws, [20, 13, 13, 13, 14, 14, 9, 8])

# ---------------------------------------------------------------------------
# TAB 5: Keywords
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Keywords")
title_row(ws, "GCG Google Ads — Top Keywords by Spend (July)")
subtitle_row(ws, "Top 20 by cost, keyword_view. 'como hacer trading' appears in two ad groups; both rows kept.")

headers = ["Keyword", "Campaign", "Match", "Spend", "Impr", "Clicks", "Subs", "CPA"]
r = 4
header_row(ws, headers, r); r += 1
KW = [
    ("broker forex usa", "TrackB Authority", "Broad", 4233.86, 23047, 1657, 15, 282.26),
    ("broker estados unidos", "TrackB Authority", "Phrase", 2850.14, 1228, 105, 3, 950.05),
    ("plataforma de trading", "TrackB Platform", "Phrase", 2445.17, 13428, 665, 7, 349.31),
    ("invertir en forex", "TrackA Trust", "Broad", 2331.88, 14178, 1497, 2, 1165.94),
    ("forex com español", "Brand", "Broad", 2061.70, 15756, 690, 5, 412.34),
    ("mejores brokers usa", "TrackB Authority", "Phrase", 1756.77, 711, 70, 1, 1756.77),
    ("plataforma trading profesional", "TrackB Platform", "Broad", 1367.44, 11922, 735, 4, 341.86),
    ("forex confiable", "TrackA Trust", "Broad", 1288.42, 6547, 796, 9, 143.16),
    ("tradingview en español", "TrackB Platform", "Phrase", 897.24, 2306, 228, 1, 897.24),
    ("broker forex usa (2nd ad group)", "TrackB Authority", "Phrase", 859.72, 369, 14, 1, 859.72),
    ("trading en espanol", "TrackA Trust", "Phrase", 765.51, 2437, 202, 0, None),
    ("plataforma de trading (alt)", "TrackB Platform", "Broad", 709.28, 5706, 331, 3, 236.43),
    ("inversiones seguras", "TrackA Trust", "Phrase", 701.41, 1432, 217, 1, 701.41),
    ("como hacer trading", "TrackA Trust", "Phrase", 686.91, 1764, 194, 1, 686.91),
    ("como hacer trading (2nd ad group)", "TrackA Trust", "Phrase", 668.66, 1542, 191, 6, 111.44),
    ("plataforma trading español", "TrackA Trust", "Phrase", 640.69, 1839, 155, 0, None),
    ("trading en español", "TrackA Trust", "Broad", 611.01, 4075, 382, 0, None),
    ("aprender trading", "TrackA Trust", "Phrase", 605.52, 1380, 158, 1, 605.52),
    ("mercado forex", "TrackA Trust", "Phrase", 553.29, 1801, 146, 1, 553.29),
    ("herramientas de trading", "TrackB Platform", "Phrase", 304.81, 1004, 83, 0, None),
]
f = {4: MONEY, 5: INT, 6: INT, 7: INT, 8: MONEY}
for i, row_vals in enumerate(KW):
    data_row(ws, list(row_vals), r, f, alt=i % 2 == 1)
    r += 1
r += 1
bullet_row(ws, "Strongest: 'forex confiable' $143 CPA (9 subs), 'como hacer trading' 2nd ad group $111 (6), 'broker forex usa' broad $282 (15). Largest zero/weak lines: 'invertir en forex' $2,331.88 / 2 subs, 'trading en espanol' + 'trading en español' + 'plataforma trading español' combined $2,017.21 / 0 subs.", r)

set_widths(ws, [30, 18, 8, 10, 9, 8, 7, 10])

# ---------------------------------------------------------------------------
# TAB 6: Geo
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Geo")
title_row(ws, "GCG Google Ads — Geo Compliance (July)")
subtitle_row(ws, "GCG is contracted to the United States. geographic_view, location of presence.")

headers = ["Country", "Spend", "Impr", "Clicks", "Share of spend"]
r = 4
header_row(ws, headers, r); r += 1
data_row(ws, ["United States (2840)", 47652.73, 918059, 23577, 1.0], r, {2: MONEY, 3: INT, 4: INT, 5: PCT})
r += 2
bullet_row(ws, "VERDICT: 100.00% of July Google spend (Search + PMax) delivered in the United States. Compliant.", r)

set_widths(ws, [24, 12, 11, 10, 13])

# ---------------------------------------------------------------------------
# TAB 7: MoM
# ---------------------------------------------------------------------------
ws = wb.create_sheet("MoM")
title_row(ws, "GCG Google Ads — Trend (May-Jul 2026)")
subtitle_row(ws, "May-June carried from prior cycles (platform basis). PMax reported separately; it has no prior-month comparator.")

headers = ["Metric", "May", "June", "July", "MoM (Jul v Jun)"]
r = 4
header_row(ws, headers, r); r += 1
MOM = [
    ("Search spend ($)", 15201.00, 22523.79, 29478.13, 0.3087),
    ("Search submitted apps (Step 5)", 76, 67, 73, 0.0896),
    ("Search CPA ($)", 200.01, 336.18, 403.81, 0.2012),
    ("PMax spend ($)", None, None, 18174.60, None),
    ("PMax submitted apps (Step 5, all_conv)", None, None, 49, None),
    ("PMax CPA ($)", None, None, 370.91, None),
]
f = {2: MONEY, 3: MONEY, 4: MONEY, 5: PCT}
for i, row_vals in enumerate(MOM):
    data_row(ws, list(row_vals), r, f, alt=i % 2 == 1)
    r += 1

set_widths(ws, [34, 12, 12, 12, 13])

# ---------------------------------------------------------------------------
# TAB 8: Notes & QA
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Notes & QA")
title_row(ws, "GCG Google Ads — Notes, QA & Counting Rules", span=1)
r = 3
for t in [
    "SOURCE",
    "Google Ads API via MCP run_gaql_query, customer 4781995752, pulled 2026-08-19. Period 2026-07-01..2026-07-31, USD.",
    "",
    "COUNTING RULES (read before using any number)",
    "Submitted applications = 'FOREX.com - US ES - PO App Form - Step 5 - Submission Completed' only. Search July = 73. PMax July = 49.",
    "Search metrics.conversions reads 76: it includes 3 offline GCLID events (2 approved, 1 funded) newly counted as primary in July. June's 67 was pure Step 5. Never quote 76 as submitted apps.",
    "PMax metrics.conversions reads 0: the campaign goal config excludes Step 5 from primary. Its 49 Step-5 conversions sit in all_conversions; view-through = 1, so they are click-based. Goal-config fix goes to the recommendations file (account mechanics, not client deck).",
    "PMax launched week of 2026-07-13: partial-month line. Weekly spend $7,448.86 / $5,596.51 / $5,129.23.",
    "",
    "RECONCILIATION",
    "Search campaign spend sums to $29,478.13 vs client tracker Google $29,478 (delta $0.13, rounding). PMax $18,174.60 vs tracker 'YT (Pmax)' $18,175 (delta $0.40, rounding). Both exact for client purposes; tracker is the client-facing basis.",
    "Geo: one row, US, $47,652.73 = Search $29,478.13 + PMax $18,174.60 exactly.",
    "",
    "CARRY-OVER CHECK (June deck commitment #2: ad-rank program before budget)",
    "Partially delivered by the numbers: Trust IS 27% → 32.6%, lost-to-rank → 58.7% (best in account), CPA $433 → $390.86 with volume leadership (28 subs). Authority and Platform still lose 63-70% to rank; Brand 79%. Budget still rose 31% while the rank ceiling stands — spend outgrew apps again (+31% vs +9%).",
]:
    if t and t == t.upper() and len(t) < 60:
        section_row(ws, t, r, span=1)
    else:
        bullet_row(ws, t, r, span=1)
        ws.row_dimensions[r].height = 28 if t else 6
    r += 1
ws.column_dimensions["A"].width = 130

out = "reports/forex/gcg/2026-07/data/GCG-GoogleAds-July-2026-data.xlsx"
wb.save(out)
print(f"saved {out}")

# self-check: internal sums
assert abs(s_spend - 29478.13) < 0.02, s_spend
assert s_subs == 73
assert abs((s_spend + PMAX[2]) - 47652.73) < 0.01
print("self-check OK: search spend, subs, google total reconcile")
