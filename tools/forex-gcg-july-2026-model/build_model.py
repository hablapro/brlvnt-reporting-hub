#!/usr/bin/env python3
"""GCG July 2026 cross-channel model + figures.json.

Client-facing spend = client budget tracker (basis ruling). Conversions and
CPA never sum across channels; total row carries a dash. GCG only. Mirrors
tools/forex-july-2026-ggmi-model/build_model.py's shape, plus a MoM tab
(DOCTRINE §11 requires a prior-month comparator on every KPI, and GCG has
real June comparators GGMI mostly did not).

Sources: reports/forex/gcg/2026-07/data/GCG-{GoogleAds,Meta,Azerion,
Quantcast,GA4}-July-2026-data.xlsx, data/sources/GCG-client-tracker-
July-2026.xlsx, qa/qa-and-model.md (July + June). Every number here traces
to one of those.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

BASE = Path("reports/forex/gcg/2026-07")
DASH = "—"
num = lambda vals: [v for v in vals if v != DASH]

# ---- July: Channel rows -----------------------------------------------
# (line, spend(tracker), impressions, clicks, apps/results, cost_per_app,
#  conversion source/note, viewability)
CH = [
    ("Google Search", 29478, 125392, 9568, 73, 403.81,
     "Google Ads platform, Step 5 event, click-based (metrics.conversions "
     "reads 76 incl. 3 offline GCLID rows; scorecard basis is Step 5 = 73)",
     DASH),
    ("Google PMax (YT)", 18175, 792667, 14009, 49, 370.91,
     "Google Ads platform, Step 5 in all_conversions only; goal config does "
     "not set it as primary (ops item, see recommendations)", DASH),
    ("Meta", 6940, 493027, 11767, DASH, DASH,
     "Conversion objective (OUTCOME_SALES) delivered as committed; 284 "
     "pixel events are the fb_pixel_custom rollup, directional only, not a "
     "scorecard figure (ruling 2026-08-04)", DASH),
    ("Quantcast (Display)", 29857, 22486572, 2574, DASH, DASH,
     "15 platform results, internal only, base too low for a client figure "
     "(GGMI-parity note)", "49.26%"),
    ("Azerion (Display)", 31477, 4769231, 8892, 80, 393.46,
     "Vendor-reported applications", "64.95% (vendor claims 71.28%, flagged)"),
    ("Native (QC+Azerion)", 20298, 1170255 + 9411481, 816 + 1660, DASH, DASH,
     "Upper-funnel; Quantcast 10 results internal only, Azerion Native has "
     "no conversion tracking", "Azerion 72.68% / Quantcast 57.98%"),
]
TRACKER_TOTAL = 136224  # tracker's own stated total (category subtotals)
tot_impr = sum(num([c[2] for c in CH]))
tot_clicks = sum(num([c[3] for c in CH]))
line_sum = sum(c[1] for c in CH)
# Tracker total (136,224) is $1 off the sum of its own six rounded lines
# (136,225) — a rounding artifact in the tracker's own category subtotals,
# not something to reconcile away. Tolerance, not an exact-match assert.
assert abs(line_sum - TRACKER_TOTAL) <= 1, (line_sum, TRACKER_TOTAL)

wb = Workbook()
ws = wb.active
ws.title = "Cross-Channel Model"
ws.append(["Channel", "Spend (tracker)", "Impressions", "Clicks",
           "Submitted apps / results", "Cost per app", "Conversion source / note",
           "Viewability"])
for c in ws[1]:
    c.font = Font(bold=True)
for row in CH:
    ws.append(list(row))
ws.append(["TOTAL", TRACKER_TOTAL, tot_impr, tot_clicks, DASH, DASH,
           "Conversions and CPA are not summed: each channel reports a "
           "different event from a different system.", DASH])
for cell in ws[ws.max_row]:
    cell.font = Font(bold=True)
ws.append([])
ws.append(["Spend basis",
           "Client budget tracker (last update 08/06/2026, screenshot "
           "supplied by Renzo 2026-08-14), transcribed to data/sources/"
           "GCG-client-tracker-July-2026.xlsx. Platform deltas recalculated "
           "silently, internal only. Total row is the tracker's own stated "
           "total (136,224); the six rounded line items sum to 136,225, a "
           "$1 rounding artifact in the tracker's own category subtotals."])
ws.append(["Impressions/clicks basis",
           "Platform/vendor delivery figures. Google Search+PMax combined "
           "impressions = 918,059 (Search 125,392 + PMax 792,667). Native = "
           "Azerion Native (1,170,255) + Quantcast NativeOnly (9,411,481). "
           "Meta clicks = link clicks."])
ws.append(["Entities", "GCG only. GGMI is a separate cycle."])
for i, w in enumerate([22, 15, 13, 10, 22, 13, 90, 30]):
    ws.column_dimensions[chr(65 + i)].width = w
ws.freeze_panes = "A2"

# ---- MoM vs June --------------------------------------------------------
# (channel, June spend, July spend, June apps/results, July apps/results, note)
MOM = [
    ("Google Search", 22524, 29478, "67 apps @ $336.18 CPA",
     "73 apps @ $403.81 CPA", "Spend +30.9%, apps +9.0%, CPA +20.1%. June "
     "had no PMax line; Search is the like-for-like comparison."),
    ("Google PMax (YT)", DASH, 18175, DASH, "49 apps @ $370.91 CPA",
     "New line, launched week of Jul 13. No June comparator."),
    ("Meta", 30711, 6940, "136 pixel events (CTR campaign)",
     "284 pixel events (conversion campaign, 44% of spend)",
     "Spend -77.4% by design as the Q2 CTR engine wound down and the Q3 "
     "conversion-objective structure took over. June ran one CTR campaign; "
     "July split traffic and conversion objectives, so pixel-event counts "
     "are not like-for-like — compare within objective only."),
    ("Quantcast (Display)", 30559, 29857, "15 results", "15 results",
     "Spend -2.3%. Viewability 46.9% -> 49.26% (+2.36pts), second straight "
     "month of improvement, still below the 70% floor."),
    ("Azerion (Display)", 29586, 31477, "58 apps @ $510.10 CPA",
     "80 apps @ $393.46 CPA", "Spend +6.4%, apps +37.9%, CPA -22.8%. "
     "Viewability 58.8% -> 64.95% (computed), still below the 70% floor."),
    ("Native (QC+Azerion)", 3645, 20298, DASH, DASH,
     "June was an Azerion-only pilot ($3,645). July adds Quantcast Native, "
     "a brand-new campaign (created Jul 2) — not a like-for-like comparison, "
     "stated once."),
]
ws2 = wb.create_sheet("MoM vs June")
ws2.append(["Channel", "June spend", "July spend", "June apps / results",
            "July apps / results", "Note"])
for c in ws2[1]:
    c.font = Font(bold=True)
for row in MOM:
    ws2.append(list(row))
ws2.append(["TOTAL", 117024, TRACKER_TOTAL, DASH, DASH,
            "Spend +16.4%. Conversions and CPA are not summed across "
            "channels; comparators live at the channel level above."])
for cell in ws2[ws2.max_row]:
    cell.font = Font(bold=True)
ws2.append([])
ws2.append(["Source", "June figures are the delivered GCG June 2026 "
            "Performance Report (Summary tab) and June's qa-and-model.md. "
            "July figures are the channel workbooks in data/, tracker-"
            "reconciled (see July qa-and-model.md)."])
for i, w in enumerate([20, 12, 12, 24, 24, 90]):
    ws2.column_dimensions[chr(65 + i)].width = w
ws2.freeze_panes = "A2"

wb.save(BASE / "model" / "GCG-July-2026-cross-channel-model.xlsx")

figures = {
    "entity": "GCG",
    "month": "2026-07",
    "spend_basis": (
        "Client budget tracker, last update 08/06/2026, supplied by Renzo "
        "2026-08-14 and transcribed to data/sources/"
        "GCG-client-tracker-July-2026.xlsx. Platform deltas recalculated "
        "silently: Google Search -$0.13, Google PMax -$0.40, Meta -$0.50, "
        "Azerion display +$715.46 (raw+fee vs tracker), Quantcast display "
        "-$2.25, Native +$231.08 (fee-inclusive raw vs tracker). Total.spend "
        "136,224 is the tracker's own stated total; the six rounded line "
        "items sum to 136,225 (tracker-internal $1 rounding, not a figure "
        "we introduce)."
    ),
    "figures": {
        "google_search.spend": 29478,
        "google_search.impressions": 125392,
        "google_search.clicks": 9568,
        "google_search.submitted_apps": 73,
        "google_search.cost_per_app": 403.81,
        "google_pmax.spend": 18175,
        "google_pmax.impressions": 792667,
        "google_pmax.clicks": 14009,
        "google_pmax.submitted_apps": 49,
        "google_pmax.cost_per_app": 370.91,
        "google.combined_impressions": 918059,
        "meta.spend": 6940,
        "meta.impressions": 493027,
        "meta.clicks": 11767,
        "meta.pixel_events": 284,
        "meta.traffic_spend": 3888,
        "meta.conversion_spend": 3052,
        "quantcast.spend": 29857,
        "quantcast.impressions": 22486572,
        "quantcast.clicks": 2574,
        "quantcast.results": 15,
        "quantcast.viewability": 49.26,
        "azerion.spend": 31477,
        "azerion.impressions": 4769231,
        "azerion.clicks": 8892,
        "azerion.submitted_apps": 80,
        "azerion.cost_per_app": 393.46,
        "azerion.viewability": 64.95,
        "azerion.viewability_vendor_claim": 71.28,
        "native.spend": 20298,
        "native.impressions": tot_impr - 125392 - 792667 - 493027 - 22486572 - 4769231,
        "native.clicks": 816 + 1660,
        "native.azerion_impressions": 1170255,
        "native.azerion_clicks": 816,
        "native.azerion_viewability": 72.68,
        "native.quantcast_impressions": 9411481,
        "native.quantcast_clicks": 1660,
        "native.quantcast_viewability": 57.98,
        "native.quantcast_results": 10,
        "total.spend": TRACKER_TOTAL,
        "total.impressions": tot_impr,
        "total.clicks": tot_clicks,
        "ga4.es_sessions": 66398,
        "ga4.es_users": 29901,
        "ga4.meta_capture_pct": 67.2,
        "mom.google_search_spend_pct": 30.9,
        "mom.google_search_apps_pct": 9.0,
        "mom.google_search_cpa_pct": 20.1,
        "mom.meta_spend_pct": -77.4,
        "mom.quantcast_spend_pct": -2.3,
        "mom.quantcast_viewability_pts": 2.36,
        "mom.azerion_spend_pct": 6.4,
        "mom.azerion_apps_pct": 37.9,
        "mom.azerion_cpa_pct": -22.8,
        "mom.azerion_viewability_pts": 6.15,
        "mom.total_spend_pct": 16.4,
    },
    "allow": [2026, 2025],
    "history": [
        22524, 181470, 12888, 67, 336.18,           # Google (Search-only)
        30711, 3058402, 74572, 136,                 # Meta
        30559, 37233620, 2942, 15,                  # Quantcast (Display)
        29586, 4515400, 19653, 58, 510.10,          # Azerion (Display)
        3645,                                       # Native (Azerion pilot)
        117024,                                     # Total spend
        46.9, 58.8,                                 # June viewability (QC / Azerion)
        80231, 41795,                                # GA4 ES sessions / users
    ],
    "history_source": (
        "June figures taken verbatim from the delivered GCG June 2026 "
        "Performance Report, Summary tab "
        "(reports/forex/gcg/2026-06/output/GCG-June-2026-Performance-"
        "Report.xlsx). June Google Ads ($22,524, 67 apps, $336.18 CPA) is "
        "Search-only; June had no PMax line. Viewability comparators (QC "
        "46.9%, Azerion 58.8%) from June's qa-and-model.md findings. GA4 ES "
        "sessions/users (80,231 / 41,795) from the June column of the July "
        "GA4 workbook's ES Audience tab."
    ),
    "notes": (
        "Two new July lines vs June: Google PMax ($18,175, launched wk of "
        "Jul 13, no June comparator) and Quantcast Native ($10,003, "
        "campaign created Jul 2, first delivery month, no June comparator). "
        "google_search.submitted_apps is 73 (Step 5 only) — never quote the "
        "platform's metrics.conversions=76, which newly includes 3 offline "
        "GCLID rows as of July. google_pmax.submitted_apps=49 sits in "
        "all_conversions only; the PMax goal config does not set Step 5 as "
        "a primary conversion, so the campaign is not optimizing to "
        "submitted applications — an ops item for recommendations/forex/"
        "gcg/, never a client-deck figure. meta.pixel_events (284) is the "
        "fb_pixel_custom rollup on the conversion campaign, directional "
        "only per the 2026-08-04 ruling: no Meta cost-per-app, no "
        "Meta-vs-other-channel efficiency comparison. quantcast.results "
        "(15, display) and native.quantcast_results (10) are platform-"
        "attributed and internal only — GGMI-parity 'awareness framing', "
        "not client scorecard figures. azerion.viewability is the computed "
        "figure (64.95%, viewable/served across the vendor's ad-set rows); "
        "the vendor's own summary tab claims 71.28% — discrepancy flagged "
        "in qa-and-model.md, computed figure is what's used. Both "
        "Quantcast (49.26%) and Azerion (64.95%) display viewability sit "
        "below the 70% IAB floor; Azerion Native (72.68%) is above it. "
        "Client funnel rows (submitted/live/approved/funded/traded) are "
        "NOT included: the July client funnel/BvA dashboard has not been "
        "supplied by Renzo as of this build. Comms-since-June-report is "
        "also still open with Renzo. Neither blocks the model; both block "
        "the narrative phase. mix/traffic-conversion Meta split "
        "(meta.traffic_spend 3,888 + meta.conversion_spend 3,052 = "
        "6,940 = meta.spend) reflects the two campaign objectives, rounded "
        "from $3,887.90 and $3,051.60."
    ),
}
(BASE / "figures.json").write_text(json.dumps(figures, indent=2))
print("model + figures.json written")
print("total impressions", tot_impr, "total clicks", tot_clicks)
print("native impressions (computed)", figures["figures"]["native.impressions"])
