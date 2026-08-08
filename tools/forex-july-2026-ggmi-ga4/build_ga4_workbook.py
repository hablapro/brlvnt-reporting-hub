"""
Build GGMI (LATAM) GA4 July 2026 data workbook.
Source: Google Analytics 4 MCP, property 508849216 (Forex LAT). Date range 2026-07-01..2026-07-31,
MoM vs 2026-06-01..2026-06-30. Reporting only, diagnostic tab excluded from any conversion claim.
House style matches reports/forex/ggmi/2026-07/data/GGMI-Meta-July-2026-data.xlsx.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
SUBHEAD_FILL = "D9E1F2"
SECTION_FILL = "2E5496"
ROW_ALT_FILL = "F2F5FB"
GRAY_TEXT = "404040"
BORDER_COLOR = "BFBFBF"
WARN_FILL = "FCE4E4"

thin = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def title_row(ws, text, row=1, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)


def subtitle_row(ws, text, row=2, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
    c.fill = PatternFill("solid", fgColor=SUBHEAD_FILL)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SUBHEAD_FILL)


def section_row(ws, text, row, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SECTION_FILL)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SECTION_FILL)


def bullet_row(ws, text, row, span=8, warn=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10, color="C00000" if warn else "000000", bold=warn)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if warn:
        for col in range(1, span + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=WARN_FILL)
    ws.row_dimensions[row].height = 30


def header_row(ws, headers, row):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def data_row(ws, values, row, formats=None, alt=False, total=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if total:
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=SECTION_FILL)
        else:
            c.font = Font(name="Calibri", size=10)
            if alt:
                c.fill = PatternFill("solid", fgColor=ROW_ALT_FILL)
        if formats and formats.get(i):
            c.number_format = formats[i]


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


PCT = "0.0%"
INT = "#,##0"
SEC = "#,##0"

# ---------------------------------------------------------------------------
# TAB 1: Summary
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Summary")
title_row(ws, "GGMI (LATAM) — GA4 — July 2026 Monthly Report")
subtitle_row(ws, "Property 508849216 (Forex LAT) · 2026-07-01 to 2026-07-31 · MoM vs June · traffic/engagement/geo source only — NOT a conversion source, see Key Events tab")

r = 4
section_row(ws, "HEADLINE", r); r += 1
for t in [
    "•  37,574 sessions in July, down 24.9% vs June's 50,055. But the drop is almost entirely one bucket: Unassigned sessions collapsed from 19,540 to 3,845 (-80.3%, -15,695 sessions), which is 126% of the entire net decline. Strip Unassigned out and sessions were flat to slightly up (30,514 in June vs 33,729 in July, +10.5%).",
    "•  The Unassigned collapse tracks Bing's own July delivery collapse (dark Jul 1-10, relaunch Jul 11), not a fix to the SA360-GA4 link. Both the tagged bing/cpc row (1,895→527, -72.2%) and the '(unlinked SA360 account)' bucket (17,794→1,307, -92.7%) fell together — if the link had gone live, tagged bing/cpc would have grown as sessions moved OUT of Unassigned, not fallen with it. The link is still not confirmed done.",
    "•  Key-event gap: STILL OPEN for essentially the whole month. Every day July 1-30 shows zero key events on live_start/live_confirmation despite those events firing normally (58-141/day). The designation only went live partway through July 31 and is confirmed persistent into August 1-4 (100% of live_start/live_confirmation now counted as key events). Full evidence on the Key Events tab — diagnostic only, not a performance number.",
    "•  Venezuela anomaly: resolved for the month. 14,552 sessions in June → 1,382 in July (-90.5%), and the remaining Venezuela traffic is now organic/direct-led (591 organic + minor), not paid-driven. Consistent with Bing pausing its Venezuela-heavy legacy campaigns Jul 22 (per the Bing report). The underlying targeting bug is still not fixed — see caveat below.",
]:
    bullet_row(ws, t, r); r += 1

r += 1
section_row(ws, "DATA QUALITY FLAGS — READ BEFORE USING THESE NUMBERS", r); r += 1
for t in [
    "⚠ GA4 records ZERO usable key events for paid channels for essentially all of July. Do not compute a GA4-based CPA, conversion rate, or funnel number for this cycle. Bing/SA360, vendor tracker, and Meta pixel remain the only conversion sources this cycle (see their own workbooks and the cross-channel caveat already on file).",
    "⚠ The Venezuela drop is a volume artifact of Bing's Jul 22 campaign pause, not a fix to the geo-targeting bug the Bing report calls 'dormant, not closed.' If Bing scales legacy-style delivery again, expect Venezuela to reappear.",
    "⚠ UTM casing/medium fragmentation splits real channels into duplicate GA4 rows and, in one case, into the WRONG channel group: 'Meta / social' (127 July sessions) is classified Organic Social, not Paid Social, because its medium string doesn't match GA4's paid-social pattern. See Source/Medium tab.",
    "⚠ Landing-page and country data could not be cross-filtered to Mexico-only at the URL level this cycle — the GA4 MCP's dimensionFilter parameter is broken (documented tool defect, serializes to string server-side). Country-level Mexico cuts ARE reliable (unfiltered pulls, filtered locally); channel mix for Mexico specifically is on the Geo tab. Landing pages tab is property-wide.",
]:
    bullet_row(ws, t, r, warn=True); r += 1

r += 1
section_row(ws, "GEO", r); r += 1
for t in [
    "•  Mexico: 7,310 sessions (19.5% of the property), channel mix Direct 2,730 / Organic 1,215 / Unassigned 1,452 / Display 1,006 / Paid Search 434 / Paid Social 254 — roughly 23% Mexico traffic is paid-attributed at the channel level.",
    "•  United States: 14,192 sessions (37.8%) — the largest country on this property, consistent with prior findings that GCG (US Hispanic) campaigns land here too. Not a GGMI-specific concern; flagged for continuity only.",
]:
    bullet_row(ws, t, r); r += 1

r += 1
section_row(ws, "WHAT THIS DATA CAN AND CANNOT BE USED FOR", r); r += 1
for t in [
    "CAN: traffic volume and MoM trend, engagement quality by channel, landing-page and geo analysis, and diagnostic evidence for the tracking-gap and UTM-audit workstreams.",
    "CANNOT: any conversion, CPA, or funnel number for GGMI. GA4 key events are not a usable performance metric this cycle — see Key Events tab.",
]:
    bullet_row(ws, t, r); r += 1

set_widths(ws, [110, 14, 14, 14, 14, 14, 14, 14])

# ---------------------------------------------------------------------------
# TAB 2: Channels (July + June side by side, sessionDefaultChannelGroup)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Channels")
title_row(ws, "Channels — sessionDefaultChannelGroup", span=10)
subtitle_row(ws, "Sessions, users, engagement by default channel group. July vs June. Sorted by July sessions desc.", span=10)

headers = ["Channel", "Jul Sessions", "Jun Sessions", "MoM %", "Jul Users", "Jul Engagement Rate",
           "Jul Bounce Rate", "Jul Avg Engagement Time (s)", "Jul Engaged Sessions", "Jun Engagement Rate"]
r = 4
header_row(ws, headers, r); r += 1

# (channel, jul_sess, jun_sess, jul_users, jul_eng_rate, jul_bounce, jul_avg_secs, jul_engaged, jun_eng_rate)
rows = [
    ("Direct", 10281, 10772, 5194, 0.6674, 0.3326, 76.5, 6862, 0.6765),
    ("Organic Search", 8387, 9246, 4935, 0.6918, 0.3082, 84.1, 5802, 0.6988),
    ("Display", 5639, 3127, 5400, 0.1695, 0.8305, 30.6, 956, 0.2466),
    ("Referral", 5632, 2163, 5389, 0.0554, 0.9446, 7.6, 312, 0.1683),
    ("Unassigned", 3845, 19540, 3104, 0.4213, 0.5787, 84.2, 1620, 0.5623),
    ("Cross-network", 1911, 519, 1909, 0.0120, 0.9880, 1.4, 23, 0.0193),
    ("Paid Search", 702, 2107, 555, 0.6481, 0.3519, 95.2, 455, 0.5083),
    ("Paid Social", 533, 1745, 498, 0.3996, 0.6004, 45.5, 213, 0.2636),
    ("Organic Social", 201, 232, 128, 0.6517, 0.3483, 122.5, 131, 0.6379),
    ("Mobile Push Notifications", 179, 204, 64, 0.4358, 0.5642, 32.9, 78, 0.4608),
    ("Email", 144, 216, 108, 0.8194, 0.1806, 133.3, 118, 0.8796),
    ("AI Assistant", 105, 172, 53, 0.7619, 0.2381, 71.4, 80, 0.6802),
    ("Organic Video", 12, 6, 9, 0.6667, 0.3333, 278.6, 8, 0.5000),
    ("Paid Other", 2, 3, 1, 0.5000, 0.5000, 0, 1, 0.6667),
    ("Paid Video", 1, 2, 1, 1.0000, 0.0000, 2, 1, 1.0000),
    ("Affiliates", 0, 1, 0, None, None, None, 0, 1.0000),
]

for i, (name, jul, jun, users, eng, bounce, secs, engaged, jun_eng) in enumerate(rows):
    mom = (jul - jun) / jun if jun else None
    vals = [name, jul, jun, mom, users, eng, bounce, secs, engaged, jun_eng]
    data_row(ws, vals, r, formats={2: INT, 3: INT, 4: PCT, 5: INT, 6: PCT, 7: PCT, 8: SEC, 9: INT, 10: PCT}, alt=(i % 2 == 1))
    r += 1

jul_total = sum(x[1] for x in rows)
jun_total = sum(x[2] for x in rows)
data_row(ws, ["TOTAL", jul_total, jun_total, (jul_total - jun_total) / jun_total, "", "", "", "", "", ""], r,
         formats={2: INT, 3: INT, 4: PCT}, total=True)
r += 2
bullet_row(ws, "Read: excluding Unassigned, July sessions were 33,729 vs June's 30,515 (+10.5%). The headline -24.9% total is a channel-mix artifact of Bing's July delivery collapse, not a traffic decline. Display and Referral both grew in absolute terms (Referral driven by s0.2mdn.net — the CM360 creative CDN referrer for untagged ad clicks, see Source/Medium tab); Paid Search and Paid Social both fell in line with Bing's and Meta's own reported spend collapses this cycle.", r, span=10)
set_widths(ws, [22, 12, 12, 10, 10, 14, 12, 16, 14, 14])
ws.freeze_panes = "B5"

# ---------------------------------------------------------------------------
# TAB 3: Source / Medium
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Source-Medium")
title_row(ws, "Source / Medium — July 2026", span=6)
subtitle_row(ws, "Top rows by sessions, July 1-31. UTM casing/scheme duplicates flagged in the Flag column.", span=6)

sm_rows = [
    ("(direct) / (none)", 10281, 5194, 0.6674, 0.3326, ""),
    ("google / organic", 7491, 4573, 0.6731, 0.3269, ""),
    ("s0.2mdn.net / referral", 5250, 5216, 0.0015, 0.9985, "Untagged CM360 creative CDN referrer — ad clicks with no UTMs at all"),
    ("Azerion / display", 2081, 2059, 0.0202, 0.9798, "Casing dup — see rollup below"),
    ("(data not available)", 1890, 1891, 0.0005, 0.9995, "Classified Cross-network, not GGMI paid media"),
    ("Quantcast / display", 1569, 1524, 0.1141, 0.8859, "Casing dup — see rollup below"),
    ("(unlinked SA360 account) / (unlinked SA360 account)", 1307, 1125, 0.5692, 0.4308, "Bing paid search, unlinked SA360 — known issue"),
    ("azerion / display", 1016, 1011, 0.0413, 0.9587, "Casing dup — see rollup below"),
    ("(not set)", 867, 543, 0.0311, 0.9689, "Falls in Unassigned"),
    ("(not set) / inappuser", 809, 700, 0.8591, 0.1409, "Falls in Unassigned"),
    ("tradingview / display", 738, 493, 0.8035, 0.1965, "Partnership tag, not paid media"),
    ("Quantcast / native", 713, 641, 0.1753, 0.8247, "Falls in Unassigned, not Display — new finding, see Unassigned tab"),
    ("bing / cpc", 527, 431, 0.6053, 0.3947, ""),
    ("bing / organic", 513, 268, 0.8090, 0.1910, ""),
    ("meta / paid-social", 462, 436, 0.3636, 0.6364, "Correctly classified Paid Social"),
    ("quantcast / display", 327, 313, 0.3058, 0.6942, "Casing dup — see rollup below"),
    ("ntp.msn.com / referral", 263, 24, 0.9316, 0.0684, ""),
    ("google / cpc", 198, 141, 0.7980, 0.2020, ""),
    ("LATAM_NA_PN / Push", 179, 64, 0.4358, 0.5642, ""),
    ("password.loginandtrade.com / referral", 142, 36, 0.8662, 0.1338, ""),
    ("Meta / social", 127, 95, 0.5984, 0.4016, "MISCLASSIFIED — lands in Organic Social channel group, not Paid Social"),
    ("Azerion / native", 107, 102, 0.1682, 0.8318, "Falls in Unassigned, not Display"),
    ("chatgpt.com / ai-assistant", 88, 45, 0.7955, 0.2045, ""),
    ("et / email", 74, 62, 0.6892, 0.3108, ""),
    ("Meta / paidsocial", 71, 69, 0.6338, 0.3662, "Third Meta medium spelling — correctly Paid Social despite casing"),
]

r = 4
header_row(ws, ["Source / Medium", "Sessions", "Users", "Engagement Rate", "Bounce Rate", "Flag"], r); r += 1
for i, row in enumerate(sm_rows):
    data_row(ws, list(row), r, formats={2: INT, 3: INT, 4: PCT, 5: PCT}, alt=(i % 2 == 1))
    r += 1

r += 1
section_row(ws, "UTM CASING ROLLUP — VENDOR DISPLAY/NATIVE (July)", r, span=6); r += 1
header_row(ws, ["Vendor / Medium", "Capitalized Sessions", "Lowercase Sessions", "Malformed Sessions", "Total", "Cap Share"], r); r += 1
casing_rows = [
    ("Azerion display", 2081, 1016, 0, 3097, 2081/3097),
    ("Quantcast display", 1569, 327, 4, 1900, 1569/1900),
    ("Quantcast native", 713, 0, 0, 713, 1.0),
    ("Azerion native", 107, 0, 0, 107, 1.0),
]
for i, row in enumerate(casing_rows):
    data_row(ws, list(row), r, formats={2: INT, 3: INT, 4: INT, 5: INT, 6: PCT}, alt=(i % 2 == 1))
    r += 1
cap_tot = sum(x[1] for x in casing_rows)
low_tot = sum(x[2] for x in casing_rows)
mal_tot = sum(x[3] for x in casing_rows)
gt = cap_tot + low_tot + mal_tot
data_row(ws, ["TOTAL", cap_tot, low_tot, mal_tot, gt, cap_tot / gt], r, formats={2: INT, 3: INT, 4: INT, 5: INT, 6: PCT}, total=True)
r += 2
bullet_row(ws, "5,817 vendor display/native sessions in July split 76.8% capitalized-scheme / 23.1% lowercase-scheme / 0.07% malformed (comma-joined macro). Consistent with the 14-day 1,116:630 ratio the July 28 platform audit found; the split persists at full-month scale. Each casing variant is a separate GA4 row — reporting Quantcast or Azerion's true session volume requires summing both variants, which this workbook does but a client-facing GA4 pull would not by default.", r, span=6)
r += 1
bullet_row(ws, "Meta medium fragmentation is a distinct, second defect: 3 medium spellings for the same platform (meta/paid-social, Meta/social, Meta/paidsocial) totaling 660 sessions. 'Meta / social' (127 sessions, 19.2% of Meta-platform sessions) reads GA4's channel-grouping rules as organic and is bucketed into Organic Social rather than Paid Social — understating Paid Social and overstating Organic Social channel totals. Not previously quantified; the July 28 UTM audit flagged Meta ad-level tagging as unaudited and this corroborates it from the session side.", r, span=6)
set_widths(ws, [55, 12, 10, 14, 12, 55])
ws.freeze_panes = "B5"

# ---------------------------------------------------------------------------
# TAB 4: Geo
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Geo")
title_row(ws, "Geo — Country Breakdown, July 2026", span=5)
subtitle_row(ws, "GGMI targets Mexico ONLY. Non-Mexico volume is compliance-relevant, not just descriptive.", span=5)

geo_rows = [
    ("United States", 14192, 13397),
    ("Mexico", 7310, 3981),
    ("Colombia", 2098, 921),
    ("Spain", 1514, 840),
    ("Chile", 1483, 1080),
    ("Venezuela", 1382, 918),
    ("Dominican Republic", 1219, 619),
    ("Argentina", 1183, 563),
    ("Ecuador", 827, 412),
    ("Peru", 678, 383),
    ("Costa Rica", 625, 239),
    ("Jamaica", 540, 298),
    ("Brazil", 426, 285),
    ("Bolivia", 411, 237),
    ("Panama", 250, 99),
    ("Guatemala", 248, 158),
]
r = 4
header_row(ws, ["Country", "Jul Sessions", "Jul Users", "% of Jul Total", "Jun Sessions (for VE/MX only)"], r); r += 1
jul_prop_total = 37574
for i, (country, sess, users) in enumerate(geo_rows):
    jun_ref = 9236 if country == "Mexico" else (14552 if country == "Venezuela" else "")
    data_row(ws, [country, sess, users, sess / jul_prop_total, jun_ref], r,
             formats={2: INT, 3: INT, 4: PCT}, alt=(i % 2 == 1))
    r += 1
r += 1

section_row(ws, "VENEZUELA ANOMALY — VERDICT: RESOLVED IN VOLUME, NOT IN CONFIGURATION", r, span=5); r += 1
for t in [
    "•  Venezuela sessions: 14,552 (June) → 1,382 (July), -90.5%. Mexico by contrast fell only 9,236 → 7,310 (-20.9%), a normal-range MoM move, not a compliance breach.",
    "•  July Venezuela traffic composition (unfiltered channel crosstab, partial capture): Organic Search 591, Unassigned 521, remainder split across smaller channels. This is organic/direct-led, NOT paid-media-led — a reversal from June where Venezuela was overwhelmingly Bing-paid-search-driven.",
    "•  This matches the Bing channel finding exactly: the Venezuela-heavy legacy campaigns (PlatformIntercept and siblings) were paused Jul 22, and all 6 new Mexico-only campaigns delivered 100% Mexico. The GA4 drop is the downstream effect of that pause, not a separate fix.",
    "•  DO NOT tell the client the underlying issue is fixed. Per the Bing report, all 9 enabled campaigns still carry positiveGeoTargetType = PRESENCE_OR_INTEREST and Venezuela is still not on the location-negative list (only Canada, Guatemala, US are). The breach is dormant, not closed, and will reappear if Bing scales legacy-style delivery or resumes the paused campaigns without fixing geo targeting first.",
]:
    bullet_row(ws, t, r, span=5); r += 1

r += 1
section_row(ws, "MEXICO CHANNEL MIX (July)", r, span=5); r += 1
header_row(ws, ["Channel", "Mexico Sessions", "% of Mexico Total", "", ""], r); r += 1
mx_rows = [
    ("Direct", 2730), ("Unassigned", 1452), ("Organic Search", 1215),
    ("Display", 1006), ("Paid Search", 434), ("Paid Social", 254),
]
mx_total = 7310
mx_captured = sum(x[1] for x in mx_rows)
for i, (ch, sess) in enumerate(mx_rows):
    data_row(ws, [ch, sess, sess / mx_total, "", ""], r, formats={2: INT, 3: PCT}, alt=(i % 2 == 1))
    r += 1
data_row(ws, ["Other channels (residual, not individually pulled)", mx_total - mx_captured, (mx_total - mx_captured) / mx_total, "", ""], r, formats={2: INT, 3: PCT})
r += 2
bullet_row(ws, "Channel-level paid share for Mexico: Display + Paid Search + Paid Social = 1,694 of 7,310 Mexico sessions (23.2%). This is the closest derivable proxy for 'paid vs unpaid, Mexico' this cycle — a true landing-page-level Mexico cut was not pulled because the GA4 MCP's dimensionFilter is broken (documented tool defect: it serializes to a string server-side instead of filtering) and a full unfiltered landing-page x country cross would be a very large pull for one data point. Flagged as a gap, not fabricated.", r, span=5)
set_widths(ws, [40, 16, 18, 18, 14])
ws.freeze_panes = "B5"

# ---------------------------------------------------------------------------
# TAB 5: Landing Pages
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Landing Pages")
title_row(ws, "Landing Pages — July 2026 (property-wide)", span=5)
subtitle_row(ws, "Top pages by sessions. NOT filtered to Mexico — see Geo tab for the Mexico channel-level proxy and the tool-limitation note.", span=5)

lp_rows = [
    ("(not set)", 10226, 9618, 0.0240, 77351),
    ("/es/", 6027, 3498, 0.6831, 442165),
    ("/es/lp/broker-de-confianza/", 1636, 1625, 0.1296, 26049),
    ("/es/login/", 1253, 567, 0.7933, 90856),
    ("/es/lp/forex-brand-trust-live/", 1009, 987, 0.2993, 27280),
    ("/es/lp/tradingview-forex/", 894, 883, 0.1342, 19314),
    ("/es/lp/plataforma-de-verdad/", 866, 823, 0.4307, 27169),
    ("/es/forex-trading/usd-mxn/", 796, 112, 0.4133, 11362),
    ("/en/", 561, 455, 0.9804, 111618),
    ("/account/login?sso=true", 545, 320, 0.9064, 49275),
    ("/es/trading-platforms/overview/", 531, 469, 0.3672, 12845),
    ("/app/funding/deposit/options", 510, 464, 0.9157, 17205),
    ("/es/trading-platforms/tradingview/", 510, 465, 0.8471, 109233),
    ("/es/trading-accounts/new-trading-account/", 390, 346, 0.5974, 49356),
    ("/app/documents/account-verification", 366, 155, 0.8934, 4951),
    ("/es/trading-platforms/simulated-trading/", 358, 345, 0.7514, 20965),
    ("/es/trading-platforms/trading-tools/", 352, 324, 0.2216, 5316),
    ("/es/about-us/overview/", 331, 329, 0.1571, 5456),
    ("/es/trading-platforms/trading-central/", 323, 308, 0.1950, 4008),
    ("/es/forex-trading/", 319, 280, 0.6959, 21935),
]
r = 4
header_row(ws, ["Landing Page", "Sessions", "Users", "Engagement Rate", "Total Engagement Time (s)"], r); r += 1
for i, row in enumerate(lp_rows):
    data_row(ws, list(row), r, formats={2: INT, 3: INT, 4: PCT, 5: INT}, alt=(i % 2 == 1))
    r += 1
r += 1
bullet_row(ws, "'(not set)' as top landing page (10,226 sessions, 2.4% engagement) is almost entirely app/in-context sessions without a resolvable web landing page — consistent with the '(not set)/inappuser' and 'Quantcast/native' Unassigned rows on the Unassigned tab. /es/ (homepage) and the Bing-driven LP set (broker-de-confianza, forex-brand-trust-live, tradingview-forex, plataforma-de-verdad) are the real top web pages and engage normally (13-68% engagement rate).", r, span=5)
set_widths(ws, [45, 12, 10, 16, 22])
ws.freeze_panes = "B5"

# ---------------------------------------------------------------------------
# TAB 6: Key Events (diagnostic)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Key Events (diagnostic)")
title_row(ws, "Key Events — DIAGNOSTIC ONLY, NOT A CONVERSION METRIC", span=6)
subtitle_row(ws, "Evidence for the GA4 LAT-property key-event tracking gap. Do not use these counts as GGMI performance.", span=6)

r = 4
section_row(ws, "STATUS: GAP CLOSED MID-CYCLE — BUT AFTER 30 OF 31 DAYS", r, span=6); r += 1
for t in [
    "June's diagnosed gap: LAT property's ONLY designated key event was first_open (an app-only event); live_start and live_confirmation fired normally but were not marked as key events, so GA4's keyEvents metric read ~0 for every paid channel.",
    "July verdict: the fix landed, but on July 31 — the last day of the reporting month. Daily eventName x keyEvents pulls show ZERO key events counted for live_start and live_confirmation on every single day from July 1 through July 30, despite those events firing 58-141 times/day (live_start) and 12-36 times/day (live_confirmation) throughout. July 31 is the first day either event counts as a key event at all, and even then only a partial-day fraction counted (58 of 74 live_start events, 14 of 17 live_confirmation events) — consistent with the designation change taking effect mid-day.",
    "Confirmed persistent: August 1-4 pull shows live_start and live_confirmation both at 100% key-event capture (282/282 and 67/67), proving this was a real, sticking configuration change, not a one-day blip.",
    "Net effect for July: total usable key events for the entire month = 86 (58 live_start + 14 live_confirmation + 14 first_open), essentially all of it generated in the final hours of the month. Of the ~82 that could be attributed to a channel, only 7 (8.5%) trace to a Bing-paid source (6 bing/cpc + 1 unlinked-SA360). Quantcast, Azerion, and Meta show 0-1 attributed key events each for the whole month.",
]:
    bullet_row(ws, t, r, span=6); r += 1

r += 1
section_row(ws, "MONTH TOTAL — EVENT NAME x KEY EVENTS (July 2026)", r, span=6); r += 1
header_row(ws, ["Event Name", "Event Count (July)", "Key Events (July)", "% Designated", "Notes", ""], r); r += 1
ke_rows = [
    ("live_start", 2720, 58, 58/2720, "Live application started (web parent event). Designated only from July 31."),
    ("live_confirmation", 682, 14, 14/682, "Live application confirmed (parent event). Designated only from July 31."),
    ("first_open", 14, 14, 1.0, "App-only event; was already the sole designated key event before this cycle."),
    ("live_confirmation_g2", 543, 0, 0, "Platform-suffixed detail event — correctly NOT a key event (would double-count the parent)."),
    ("live_confirmation_mt5", 120, 0, 0, "Same — correctly unmarked."),
    ("live_confirmation_mt4", 19, 0, 0, "Same — correctly unmarked."),
]
for i, row in enumerate(ke_rows):
    data_row(ws, list(row) + [""], r, formats={2: INT, 3: INT, 4: PCT}, alt=(i % 2 == 1))
    r += 1
r += 1

section_row(ws, "DAILY PROOF — live_start and live_confirmation, July 2026", r, span=6); r += 1
header_row(ws, ["Date", "live_start events", "live_start key events", "live_confirmation events", "live_confirmation key events", ""], r); r += 1
daily = [
    ("2026-07-01", 58, 0, 19, 0), ("2026-07-02", 68, 0, 13, 0), ("2026-07-03", 78, 0, 15, 0),
    ("2026-07-04", 76, 0, 21, 0), ("2026-07-05", 79, 0, 20, 0), ("2026-07-06", 96, 0, 26, 0),
    ("2026-07-07", 68, 0, 21, 0), ("2026-07-08", 61, 0, 21, 0), ("2026-07-09", 75, 0, 21, 0),
    ("2026-07-10", 53, 0, 13, 0), ("2026-07-11", 88, 0, 12, 0), ("2026-07-12", 104, 0, 25, 0),
    ("2026-07-13", 92, 0, 24, 0), ("2026-07-14", 75, 0, 27, 0), ("2026-07-15", 75, 0, 19, 0),
    ("2026-07-16", 84, 0, 19, 0), ("2026-07-17", 59, 0, 26, 0), ("2026-07-18", 82, 0, 17, 0),
    ("2026-07-19", 78, 0, 16, 0), ("2026-07-20", 75, 0, 16, 0), ("2026-07-21", 141, 0, 35, 0),
    ("2026-07-22", 93, 0, 23, 0), ("2026-07-23", 118, 0, 28, 0), ("2026-07-24", 96, 0, 21, 0),
    ("2026-07-25", 88, 0, 8, 0), ("2026-07-26", 123, 0, 36, 0), ("2026-07-27", 109, 0, 29, 0),
    ("2026-07-28", 134, 0, 35, 0), ("2026-07-29", 121, 0, 27, 0), ("2026-07-30", 99, 0, 32, 0),
    ("2026-07-31", 74, 58, 17, 14),
]
for i, row in enumerate(daily):
    data_row(ws, list(row) + [""], r, formats={2: INT, 3: INT, 4: INT, 5: INT}, alt=(i % 2 == 1))
    r += 1
r += 1

section_row(ws, "AUGUST 1-4 PERSISTENCE CHECK", r, span=6); r += 1
header_row(ws, ["Event Name", "Event Count (Aug 1-4)", "Key Events (Aug 1-4)", "% Designated", "", ""], r); r += 1
aug_rows = [("live_start", 282, 282, 1.0), ("live_confirmation", 67, 67, 1.0), ("first_open", 1, 1, 1.0)]
for i, row in enumerate(aug_rows):
    data_row(ws, list(row) + ["", ""], r, formats={2: INT, 3: INT, 4: PCT}, alt=(i % 2 == 1))
    r += 1
r += 1

section_row(ws, "CHANNEL ATTRIBUTION OF JULY'S ~82 ATTRIBUTABLE KEY EVENTS (diagnostic, do not use as performance)", r, span=6); r += 1
header_row(ws, ["Source / Medium", "live_start key events", "live_confirmation key events", "first_open key events", "Total", ""], r); r += 1
attr_rows = [
    ("(direct) / (none)", 23, 5, 7, 35),
    ("(not set) / inappuser", 9, 5, 0, 14),
    ("google / organic", 8, 2, 0, 10),
    ("tradingview / display", 9, 1, 0, 10),
    ("bing / cpc", 6, 0, 0, 6),
    ("google-play / organic", 0, 0, 3, 3),
    ("bing / organic", 1, 1, 0, 2),
    ("(unlinked SA360 account)", 1, 0, 0, 1),
    ("azerion / display", 1, 0, 0, 1),
]
for i, row in enumerate(attr_rows):
    data_row(ws, list(row) + [""], r, formats={2: INT, 3: INT, 4: INT, 5: INT}, alt=(i % 2 == 1))
    r += 1
r += 1
bullet_row(ws, "82 of the month's 86 key events reconcile to a channel here (4 unreconciled — long-tail sources cut off by the API's 1,000-row cap on a 1,447-row query; immaterial to the finding). Bing-attributable total (bing/cpc + unlinked SA360) = 7 of 82 (8.5%). Quantcast: 0. Azerion: 1. Meta: 0. This is not a channel-performance ranking — it is proof that virtually no channel had usable GA4 key-event coverage for essentially the whole month.", r, span=6)
set_widths(ws, [45, 20, 20, 16, 12, 4])
ws.freeze_panes = "B5"

# ---------------------------------------------------------------------------
# TAB 7: Unassigned
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Unassigned")
title_row(ws, "Unassigned Channel — Composition, July 2026", span=4)
subtitle_row(ws, "3,845 sessions (10.2% of July total), down from 19,540 in June (-80.3%).", span=4)

r = 4
header_row(ws, ["Source / Medium", "Sessions", "% of Unassigned", ""], r); r += 1
un_rows = [
    ("(unlinked SA360 account) / (unlinked SA360 account)", 1307, 1307/3845, "Bing paid search, unlinked SA360 — known, recommended fix on file"),
    ("(not set)", 867, 867/3845, "No source/medium resolvable"),
    ("(not set) / inappuser", 809, 809/3845, "In-app users, no medium"),
    ("Quantcast / native", 713, 713/3845, "NEW finding: CM360 native tracking ads (campaign 36170375) land here, not in Display"),
    ("Azerion / native", 107, 107/3845, "Same native-ad classification issue"),
    ("Other (long tail, <10 sessions each)", 42, 42/3845, "inappuser variants, Apple/search, copilot.com, FOREX/(not set), FOREX/DOOH"),
]
for i, row in enumerate(un_rows):
    data_row(ws, list(row), r, formats={2: INT, 3: PCT}, alt=(i % 2 == 1))
    r += 1
data_row(ws, ["TOTAL", sum(x[1] for x in un_rows), sum(x[1] for x in un_rows) / 3845, ""], r, formats={2: INT, 3: PCT}, total=True)
r += 2
bullet_row(ws, "Known issue (34.0% of Unassigned): Bing/SA360 traffic, unresolved until SA360 is linked to this GA4 property. Recommended fix already on file with StoneX WebOps.", r, span=4)
r += 1
bullet_row(ws, "New issue this cycle (21.3% of Unassigned): Quantcast and Azerion NATIVE (CM360 tracking-ad) traffic is landing in Unassigned rather than the Display channel that display-format buys from the same vendors get. This is a GA4 channel-grouping rule gap for the native tag type, distinct from the source/medium casing issue on the Source-Medium tab, and was not previously quantified. Worth a line in the next UTM/tracking remediation pass.", r, span=4)
set_widths(ws, [55, 12, 16, 55])
ws.freeze_panes = "B5"

# ---------------------------------------------------------------------------
# TAB 8: MoM
# ---------------------------------------------------------------------------
ws = wb.create_sheet("MoM")
title_row(ws, "Month over Month — July vs June 2026", span=5)
subtitle_row(ws, "Session-level MoM by channel, plus the property-wide and Mexico-only totals.", span=5)

r = 4
header_row(ws, ["Metric", "June", "July", "Δ", "MoM %"], r); r += 1
mom_rows = [
    ("Total sessions (property-wide)", 50055, 37574),
    ("Total sessions excl. Unassigned", 30515, 33729),
    ("Unassigned sessions", 19540, 3845),
    ("Mexico sessions", 9236, 7310),
    ("Venezuela sessions", 14552, 1382),
    ("Direct sessions", 10772, 10281),
    ("Organic Search sessions", 9246, 8387),
    ("Display sessions", 3127, 5639),
    ("Paid Search sessions", 2107, 702),
    ("Paid Social sessions", 1745, 533),
    ("bing / cpc sessions (tagged only)", 1895, 527),
    ("(unlinked SA360 account) sessions", 17794, 1307),
]
for i, (name, jun, jul) in enumerate(mom_rows):
    delta = jul - jun
    mom = delta / jun if jun else None
    data_row(ws, [name, jun, jul, delta, mom], r, formats={2: INT, 3: INT, 4: INT, 5: PCT}, alt=(i % 2 == 1))
    r += 1
r += 1
bullet_row(ws, "Every MoM figure in this tab is sessions/traffic only — none is a conversion or revenue comparison. Pair with the Bing and Meta July workbooks for spend-side MoM; do not merge those with this tab's session counts into a blended CPA, since GA4 cannot currently attribute conversions for this property (Key Events tab).", r, span=5)
set_widths(ws, [38, 12, 12, 12, 12])
ws.freeze_panes = "B5"

# ---------------------------------------------------------------------------
# TAB 9: Notes & QA
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Notes & QA")
title_row(ws, "Notes & QA", span=1)
r = 3
section_row(ws, "SOURCE", r); r += 1
for t in [
    "Google Analytics 4 Data API via the google-analytics MCP (runReport). Property 508849216 (Forex LAT / GGMI). Date ranges: July 2026-07-01 to 2026-07-31; June 2026-06-01 to 2026-06-30 (comparison); August 2026-08-01 to 2026-08-04 (key-event persistence check only). Timezone per property metadata: America/New_York. Currency: USD (not used — no revenue metrics pulled).",
]:
    bullet_row(ws, t, r); r += 1
r += 1
section_row(ws, "RECONCILIATION CHECKS PERFORMED", r); r += 1
for t in [
    "Channel-group totals sum to source/medium crosstab totals for both July (37,574) and June (50,055) within the rows pulled — the July crosstab returned 114-115 distinct source/medium rows against a 128-country, 16-channel dataset; long-tail rows below ~1 session were not individually itemized but are captured in channel-group totals.",
    "Key-event daily figures for live_start (2,720) and live_confirmation (682) sum EXACTLY to their respective month totals from the non-date-dimensioned pull — full daily coverage confirmed for these two events (no truncation). first_open's daily breakdown (7) undercounts its month total (14) because the date x eventName query hit the API's 1,000-row cap (1,095 combinations existed) and cut off some first_open dates; the month-level total (14) is authoritative and used throughout.",
    "Source/medium x eventName key-event attribution (82 of 86 month total) sorted by keyEvents descending, so all nonzero rows should sort ahead of the 1,000-row cutoff (1,447 total combinations existed); the 4-event gap versus the 86 authoritative total is disclosed and treated as immaterial rounding, not investigated further given time-box.",
    "August 1-4 persistence check run specifically to distinguish 'one-day blip' from 'sticking config change' — confirmed the latter (100% designation rate both days' worth of data).",
]:
    bullet_row(ws, t, r); r += 1
r += 1
section_row(ws, "TOOLING GOTCHAS HIT THIS PULL (per UTM-AUDIT-HANDOVER.md, confirmed again)", r); r += 1
for t in [
    "GA4 MCP dimensionFilter is broken — confirmed again this session: a country='Mexico' filter on a country x channel query returned all 128 countries unfiltered, not just Mexico. Worked around by pulling unfiltered and filtering locally in Python (documented gotcha, one retry attempted per bounded-effort policy, then worked around).",
    "Large multi-dimension pulls (date x eventName, sourceMedium x eventName) exceed the tool's output token limit and get written to a scratch file instead of returned inline; both were read and parsed in full via a local Python script rather than the Read tool's line-based paging, since the files are JSON, not line-oriented text.",
    "getConversionEvents (MCP) does not return the key-event flag; eventName x keyEvents was used throughout, per the existing house note.",
]:
    bullet_row(ws, t, r); r += 1
r += 1
section_row(ws, "METRIC DEFINITION NOTES", r); r += 1
for t in [
    "'Key events' in this workbook is GA4's own metric name (post-October-2024 rename of 'conversions') — it counts instances of designated events only, per the current-at-query-time key-event configuration. GA4 does NOT retroactively backfill key-event counts for days before a designation change; this is what makes the July 31 daily proof possible (zero on days before designation, populated the day of and after).",
    "'Unassigned' here is GA4's sessionDefaultChannelGroup value, not a Berelvant-defined bucket. Its composition changes are driven by real changes in GA4's own classification rules and the underlying traffic mix, not by anything in this repo's control.",
    "No revenue, ROAS, or CPA metric appears in this workbook. Per the hard constraint on this task, GA4 is not used as a GGMI conversion source; Bing/SA360, vendor trackers, and the Meta pixel remain the conversion sources of record for this cycle.",
]:
    bullet_row(ws, t, r); r += 1
r += 1
section_row(ws, "ANALYST READ", r); r += 1
for t in [
    "The single most important fact in this workbook is the July 31 timing of the key-event fix: it means the July reporting cycle gets almost none of the benefit of the fix StoneX shipped, even though the fix is real and (per August data) sticking. Flag this explicitly in any client-facing framing so a 'still broken' finding this month doesn't read as contradicting the '(fix in progress)' status already communicated in the June/UTM handover materials — both are true; the fix landed, just at the very end of the window this report covers.",
    "The Venezuela and Unassigned improvements are both real in the data but both trace back to Bing's own July campaign changes (dark period, relaunch, Jul 22 legacy pause) rather than to any GA4 or SA360-side fix. Do not credit GA4/tracking remediation for either.",
]:
    bullet_row(ws, t, r); r += 1
set_widths(ws, [130])

wb.save("/Users/rpro/AI-BRLVNT/Brlvnt-Reporting-Analytics-2026/reports/forex/ggmi/2026-07/data/GGMI-GA4-July-2026-data.xlsx")
print("Saved.")
