#!/usr/bin/env python3
"""GGMI July 2026 cross-channel model + figures.json.

Client-facing spend = client budget tracker (basis ruling). Conversions and
CPA never sum across channels; total row carries a dash. GGMI only.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

BASE = Path("reports/forex/ggmi/2026-07")

# Channel rows: (line, spend(tracker), impressions, clicks, apps, cost_per_app, conv_source)
DASH = "—"
CH = [
    ("Bing (Search)",        10625,  90394,     4090,  20,  531.25, "SA360 offline import, CRM-validated"),
    ("Quantcast (Display)",  39240,  50483289,  19188, DASH, DASH,  "n=15 vendor results too low; awareness framing per standing note"),
    ("Azerion (Display)",    37509,  7577455,   12915, 41,  914.85, "Vendor-reported, not CRM-validated"),
    ("Native (QC+Azerion)",  27630,  18179528,  9953,  DASH, DASH,  "Upper-funnel, no conversion tracking"),
    ("Meta (Social)",        8027,   4250494,   74489, DASH, DASH,  "Ruling 2026-08-04: spend/delivery only; 117 pixel fires unvalidated"),
    ("DOOH (Perion)",        26865,  DASH,      DASH,  DASH, DASH,  "Tracker line only; no delivery data in scope"),
]
TOTAL_SPEND = 149896
num = lambda vals: [v for v in vals if v != DASH]
tot_impr = sum(num([c[2] for c in CH]))
tot_clicks = sum(num([c[3] for c in CH]))
assert sum(c[1] for c in CH) == TOTAL_SPEND, sum(c[1] for c in CH)

wb = Workbook(); ws = wb.active; ws.title = "Cross-Channel Model"
ws.append(["Channel", "Spend (tracker)", "Impressions", "Clicks",
           "Submitted apps", "Cost per app", "Conversion source / note"])
for c in ws[1]: c.font = Font(bold=True)
for row in CH: ws.append(list(row))
ws.append(["TOTAL", TOTAL_SPEND, tot_impr, tot_clicks, DASH, DASH,
           "Conversions and CPA are not summed: each channel reports a different event from a different system."])
for cell in ws[ws.max_row]: cell.font = Font(bold=True)
ws.append([]); ws.append(["Spend basis", "Client budget tracker (last update 08/06/2026), transcribed in data/sources/. Platform deltas recalculated silently, internal only."])
ws.append(["Impressions/clicks basis", "Platform/vendor delivery figures. Native = Azerion native + Quantcast NativeOnly combined (tracker structure). Meta clicks = LINK clicks per the June deck definition (June 407,136); all clicks 91,153 stay in the channel workbook."])
ws.append(["Entities", "GGMI only. GCG is a separate cycle."])
for i, w in enumerate([22, 15, 13, 10, 14, 13, 80]):
    ws.column_dimensions[chr(65 + i)].width = w
ws.freeze_panes = "A2"
wb.save(BASE / "model" / "GGMI-July-2026-cross-channel-model.xlsx")

figures = {
    "entity": "GGMI",
    "month": "2026-07",
    "spend_basis": "Client budget tracker, last update 08/06/2026, supplied by Renzo 2026-08-14 and transcribed to data/sources/GGMI-client-tracker-July-2026.xlsx. Tracker Azerion lines = vendor x 1.10 exactly (internal note only). Platform deltas: Bing -$0.06, Meta +$0.45, Quantcast main -$3.20 vs tracker; recalculated silently per standing ruling.",
    "figures": {
        "bing.spend": 10625,
        "bing.impressions": 90394,
        "bing.clicks": 4090,
        "bing.submitted_apps": 20,
        "bing.cost_per_app": 531,
        "bing.phase_b_spend": 4742,
        "bing.phase_b_cost_per_app": 237,
        "meta.spend": 8027,
        "meta.impressions": 4250494,
        "meta.clicks": 74489,
        "azerion.spend": 37509,
        "azerion.impressions": 7577455,
        "azerion.clicks": 12915,
        "azerion.submitted_apps": 41,
        "azerion.cost_per_app": 915,
        "combined.submitted_apps": 61,
        "combined.cost_per_app": 789,
        "quantcast.spend": 39240,
        "quantcast.impressions": 50483289,
        "quantcast.clicks": 19188,
        "native.spend": 27630,
        "native.impressions": 18179528,
        "native.clicks": 9953,
        "dooh.spend": 26865,
        "total.spend": 149896,
        "total.impressions": tot_impr,
        "total.clicks": tot_clicks,
        "ga4.mexico_sessions": 7310,
        "ga4.unique_visitors": 3981,
    },
    "allow": {},
    "notes": (
        "July is two Bing accounts in one month: dark Jul 1-10, legacy Phase A Jul 11-22 ($5,883, 0 apps), "
        "rebuilt MX_ Phase B Jul 22-31 ($4,742, 20 apps at $237). Conversions re-verified live 2026-08-14, "
        "unchanged. bing.cost_per_app 531 = tracker 10,625 / 20. azerion.cost_per_app 915 = tracker 37,509 / 41 "
        "vendor-reported results (June convention). Meta is spend/delivery only per Renzo ruling 2026-08-04: no "
        "Meta cost-per-app, no Meta-vs-Bing efficiency comparison; the 117 pixel fires stay out of client "
        "artifacts as validated conversions. Quantcast tracker line = main campaign only; NativeOnly sits in the "
        "native line. DOOH is a tracker spend line with no delivery detail in our scope. Geo: Bing delivery 2.7% "
        "non-MX but PRESENCE_OR_INTEREST setting unchanged - breach dormant, not fixed; never say resolved. "
        "Meta and Quantcast geo verified clean; Azerion geo unverifiable from vendor file (says LATAM). "
        "GA4 is diagnostic only this cycle; key-event fix landed Jul 31, August is the first month GA4 can "
        "corroborate submitted apps."
    ),
}
(BASE / "figures.json").write_text(json.dumps(figures, indent=2))
print("model + figures.json written")
print("total impressions", tot_impr, "total clicks", tot_clicks)
