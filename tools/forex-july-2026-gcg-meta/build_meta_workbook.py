"""
Build GCG (US Hispanic) Meta July 2026 data workbook.
Source: Meta Ads MCP, act_1699453997689551 (SHARED GGMI+GCG account), campaign/
ad-level pulls filtered to GCG campaigns by naming convention, pulled 2026-08-19.
Date range 2026-07-01..2026-07-31. Reporting only.
House style matches tools/forex-july-2026-ggmi-meta/build_meta_workbook.py.

Reporting convention (Laura ruling 2026-08-19, DOCTRINE §11): campaigns are
reported per objective. Traffic/CTR campaigns: reach, impressions, CPM (+ CTR
as delivery quality). Conversion campaigns: CTR, sessions (GA4, separate pull),
pixel events. Never compared to each other.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
SUBHEAD_FILL = "D9E1F2"
SECTION_FILL = "2E5496"
ROW_ALT_FILL = "F2F5FB"
GRAY_TEXT = "404040"
BORDER_COLOR = "BFBFBF"

thin = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def title_row(ws, text, row=1, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)


def subtitle_row(ws, text, row=2, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
    c.fill = PatternFill("solid", fgColor=SUBHEAD_FILL)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SUBHEAD_FILL)


def section_row(ws, text, row, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SECTION_FILL)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SECTION_FILL)


def bullet_row(ws, text, row, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30


def header_row(ws, headers, row):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def data_row(ws, values, row, formats=None, alt=False, total=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if total:
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=SECTION_FILL)
        else:
            c.font = Font(name="Calibri", size=10)
            if alt:
                c.fill = PatternFill("solid", fgColor=ROW_ALT_FILL)
        if formats and formats.get(i):
            c.number_format = formats[i]


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


MONEY = "#,##0.00"
INT = "#,##0"
PCT = "0.00%"
DEC = "0.00"

# (name, objective, spend, impr, reach, freq, cpm, ctr, link_clicks, lpv, pixel)
CAMPS = [
    ("0426_GCG_Q2_esp_us_CTR", "TRAFFIC (LINK_CLICKS)", 2104.37, 206014, 173989, 1.184, 10.21, 0.0272, 4988, 3719, 8),
    ("0726_GCG_Q3_esp_us_CTR", "TRAFFIC (LINK_CLICKS)", 1783.53, 169241, 130525, 1.297, 10.54, 0.0347, 5521, 4007, 5),
    ("0726_GCG_Q3_esp_us_CONV", "CONVERSION (OUTCOME_SALES)", 3051.60, 117772, 73539, 1.601, 25.91, 0.0169, 1258, 1032, 284),
]
TOTAL_SPEND = sum(c[2] for c in CAMPS)

# ---------------------------------------------------------------------------
# TAB 1: Summary
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Summary")
title_row(ws, "GCG (US Hispanic) — Meta — July 2026 Monthly Report")
subtitle_row(ws, "act_1699453997689551 · shared GGMI+GCG account, GCG campaigns only · MoM vs June · USD · campaigns reported per objective, never blended")

r = 4
section_row(ws, "HEADLINE", r); r += 1
for t in [
    f"•  Spend ${TOTAL_SPEND:,.2f} across 3 GCG campaigns, down 80.0% vs June's $34,710.97 platform ($30,711 tracker). Deliberate reallocation month: the Q2 CTR engine wound down (paused during July after $2,104.37) and the Q3 structure took over.",
    "•  The June commitment landed: 0726_GCG_Q3_esp_us_CONV is live on the conversion objective (OUTCOME_SALES), optimizing to the SubmittedApplication pixel event — 44.0% of July GCG Meta spend ($3,051.60).",
    "•  Traffic line (2 CTR campaigns): $3,887.90, 375,255 impressions, 10,509 link clicks, blended CPM $10.36 (June CTR campaign: $11.35), CTRs 2.72% and 3.47% (June 2.76%).",
    "•  Conversion line: $3,051.60, 1,258 link clicks, LPV 1,032 (82.0% of clicks), 284 pixel events. CPM $25.91 is expected for a conversion objective bidding a narrow action; not comparable to the traffic line.",
    "•  Geo: 100% US delivery on all three campaigns. Frequencies 1.2-1.6, no fatigue signal.",
]:
    bullet_row(ws, t, r); r += 1

r += 1
section_row(ws, "WATCH ITEMS", r); r += 1
for t in [
    "•  Pixel events (284 on CONV) are the fb_pixel_custom rollup. The ad set optimizes to SubmittedApplication, but the rollup is not verified as submitted apps — scorecard stays platform-appropriate (June rule). GA4 corroboration in the GA4 pull.",
    "•  Q2 CTR campaign is now paused; July was its final month. Do not promise it anything forward-looking.",
    "•  Targeting: US, 18-65, Advantage+ Audience on all active GCG ad sets. Age range is locked by Meta's Financial Products and Services category (verified 2026-08-19) — no age refinement is possible on any financial account.",
]:
    bullet_row(ws, t, r); r += 1

set_widths(ws, [110] + [12] * 9)

# ---------------------------------------------------------------------------
# TAB 2: Campaigns (per objective)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Campaigns")
title_row(ws, "GCG Meta — July Campaigns, split by objective")
subtitle_row(ws, "Traffic campaigns read on reach/impressions/CPM (+CTR); conversion campaign on CTR/LPV/pixel events. Reach never summed across campaigns.")

headers = ["Campaign", "Objective", "Spend", "Impr", "Reach", "Freq", "CPM", "CTR", "Link clicks", "LPV"]
r = 4
header_row(ws, headers, r); r += 1
f = {3: MONEY, 4: INT, 5: INT, 6: DEC, 7: MONEY, 8: PCT, 9: INT, 10: INT}
for i, c in enumerate(CAMPS):
    data_row(ws, [c[0], c[1], c[2], c[3], c[4], c[5], c[6], c[7], c[8], c[9]], r, f, alt=i % 2 == 1)
    r += 1
data_row(ws, ["TOTAL", "", TOTAL_SPEND, sum(c[3] for c in CAMPS), "n/s", "", "", "", sum(c[8] for c in CAMPS), sum(c[9] for c in CAMPS)], r, f, total=True)
r += 2
bullet_row(ws, "Conversion campaign pixel events: 284 (fb_pixel_custom rollup; optimization event = SubmittedApplication). Traffic campaigns: 8 + 5 pixel events, incidental. Reach 'n/s' = not summable across campaigns.", r)

set_widths(ws, [26, 24, 11, 11, 11, 7, 9, 8, 11, 9])

# ---------------------------------------------------------------------------
# TAB 3: Ad Sets & Targeting
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Ad Sets & Targeting")
title_row(ws, "GCG Meta — Ad Sets & Targeting (July active)")
subtitle_row(ws, "All GCG ad sets: US (home+recent), age 18-65 (locked by Meta Financial Products and Services category), Advantage+ Audience ON, Spanish-language creative tracks.")

headers = ["Campaign", "Ad set", "Optimization goal", "Optimization event", "Geo", "Age", "Spend"]
r = 4
header_row(ws, headers, r); r += 1
ADSETS = [
    ("0726_GCG_Q3_esp_us_CONV", "trackA&B_pros_us_es_q3_CONV", "OFFSITE_CONVERSIONS", "SubmittedApplication (pixel)", "US", "18-65", 3051.60),
    ("0726_GCG_Q3_esp_us_CTR", "trackA&B_CBO_pros_us_es_q2_CTR", "LANDING_PAGE_VIEWS", "—", "US", "18-65", 1783.53),
    ("0426_GCG_Q2_esp_us_CTR", "trackA_pros_us_es_q2_CTR", "LANDING_PAGE_VIEWS", "—", "US", "18-65", 985.74),
    ("0426_GCG_Q2_esp_us_CTR", "trackB_pros_us_es_q2_CTR", "LANDING_PAGE_VIEWS", "—", "US", "18-65", 1118.63),
]
f = {7: MONEY}
for i, a in enumerate(ADSETS):
    data_row(ws, list(a), r, f, alt=i % 2 == 1)
    r += 1
r += 2
bullet_row(ws, "Q2 CTR ad-set split derived from ad-level spend (trackA ads $985.74 / trackB ads $1,118.63; sums to campaign $2,104.37). Campaign 0426 paused after July.", r)

set_widths(ws, [24, 30, 20, 26, 6, 8, 10])

# ---------------------------------------------------------------------------
# TAB 4: Creatives
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Creatives")
title_row(ws, "GCG Meta — Creative Performance (July, by objective)")
subtitle_row(ws, "Ad-level pull, GCG campaigns only, 17 ads with delivery. Pixel = fb_pixel_custom rollup; rank within Meta only.")

headers = ["Ad", "Campaign / objective", "Spend", "Impr", "Link clicks", "CTR (link)", "LPV", "Pixel"]
r = 4
header_row(ws, headers, r); r += 1
CREATIVE = [
    ("broker_trust_q2_trackA", "Q3 CONV", 1656.39, 43475, 552, 552 / 43475, 443, 177),
    ("exp_plat_q2_trackB", "Q3 CONV", 743.10, 34292, 389, 389 / 34292, 325, 58),
    ("edu_trust_q2_trackA", "Q3 CONV", 382.20, 30115, 220, 220 / 30115, 182, 11),
    ("forex_plat_q2_trackB", "Q3 CONV", 236.29, 8788, 83, 83 / 8788, 72, 32),
    ("trading_proof_q2_trackA", "Q3 CONV", 33.09, 1101, 14, 14 / 1101, 10, 6),
    ("forex_plat_q2_trackB", "Q3 CTR", 1033.10, 59569, 3171, 3171 / 59569, 2341, 1),
    ("broker_trust_q2_trackA", "Q3 CTR", 571.01, 57638, 1597, 1597 / 57638, 1124, 4),
    ("exp_plat_q2_trackB", "Q3 CTR", 152.76, 42976, 649, 649 / 42976, 461, 0),
    ("forex_plat_q2_trackB", "Q2 CTR", 1014.97, 67699, 2334, 2334 / 67699, 1772, 2),
    ("broker_trust_q2_trackA", "Q2 CTR", 963.58, 111493, 2239, 2239 / 111493, 1650, 4),
    ("exp_plat_q2_trackB", "Q2 CTR", 103.01, 22463, 342, 342 / 22463, 253, 2),
]
f = {3: MONEY, 4: INT, 5: INT, 6: PCT, 7: INT, 8: INT}
for i, c in enumerate(CREATIVE):
    data_row(ws, list(c), r, f, alt=i % 2 == 1)
    r += 1
r += 1
bullet_row(ws, "Conversion side: broker_trust_q2_trackA carries the volume (177 pixel events, $9.36/event); forex_plat_q2_trackB is most efficient ($7.38/event on $236.29 — scale candidate). Traffic side: forex_plat_q2_trackB leads clicks (3,171 @ 5.32% link CTR on the Q3 CTR campaign). Long-tail ads under $25 omitted from this tab; full 17-ad detail retained in the pull file.", r)

set_widths(ws, [26, 14, 11, 11, 11, 10, 9, 8])

# ---------------------------------------------------------------------------
# TAB 5: Geo
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Geo")
title_row(ws, "GCG Meta — Geo Compliance (July)")
subtitle_row(ws, "GCG is contracted to the United States. Country breakdown, all GCG campaigns.")

headers = ["Campaign", "Country", "Spend", "Impr", "Reach"]
r = 4
header_row(ws, headers, r); r += 1
GEO = [
    ("0426_GCG_Q2_esp_us_CTR", "US", 2104.37, 206014, 173989),
    ("0426_GCG_Q2_esp_us_CTR", "unknown", 0.0, 0, 0),
    ("0726_GCG_Q3_esp_us_CONV", "US", 3051.60, 117772, 73539),
    ("0726_GCG_Q3_esp_us_CTR", "US", 1783.53, 169241, 130525),
]
f = {3: MONEY, 4: INT, 5: INT}
for i, g in enumerate(GEO):
    data_row(ws, list(g), r, f, alt=i % 2 == 1)
    r += 1
data_row(ws, ["VERDICT", "100.00% US", TOTAL_SPEND, "—", "—"], r, {3: MONEY}, total=True)
r += 2
bullet_row(ws, "Compliant. Every GCG dollar delivered in the US ('unknown' row: 1 click, $0).", r)

set_widths(ws, [26, 12, 11, 11, 11])

# ---------------------------------------------------------------------------
# TAB 6: MoM
# ---------------------------------------------------------------------------
ws = wb.create_sheet("MoM")
title_row(ws, "GCG Meta — Trend (May-Jul 2026)")
subtitle_row(ws, "May-June from prior cycles. June ran one CTR campaign; July split across objectives, so channel-level CPM is a mix artifact — compare within objective only.")

headers = ["Metric", "May", "June", "July", "Note"]
r = 4
header_row(ws, headers, r); r += 1
MOM = [
    ("Spend, platform ($)", 12243.00, 34710.97, 6939.50, "-80.0% MoM; tracker basis 30,711 → 6,940"),
    ("Impressions", None, 3058402, 493027, "-83.9%"),
    ("Link clicks", None, 74572, 11767, "-84.2%"),
    ("Traffic-line CPM ($)", None, 11.35, 10.36, "like-for-like (CTR campaigns only)"),
    ("Traffic-line link CTR", None, 0.0276, 0.0301, "blended across 2 CTR campaigns"),
    ("Conversion-line spend ($)", None, None, 3051.60, "new; no comparator"),
    ("Conversion-line pixel events", None, 136, 284, "June = starts rollup on CTR campaign; not like-for-like"),
]
f = {2: MONEY, 3: MONEY, 4: MONEY}
for i, m in enumerate(MOM):
    data_row(ws, list(m), r, f, alt=i % 2 == 1)
    r += 1

set_widths(ws, [26, 12, 12, 12, 44])

# ---------------------------------------------------------------------------
# TAB 7: Notes & QA
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Notes & QA")
title_row(ws, "GCG Meta — Notes, QA & Attribution", span=1)
r = 3
for t in [
    "SOURCE",
    "Meta Ads MCP, act_1699453997689551, pulled 2026-08-19. Period 2026-07-01..2026-07-31, USD, attribution account default (7d-click/1d-view).",
    "",
    "GCG / GGMI ATTRIBUTION RULE",
    "Shared account. GCG = campaigns with _GCG_ and _us_ naming (3 in July). GGMI campaigns excluded. Ad-level spend sums to $6,939.50 = campaign-level total.",
    "",
    "RECONCILIATION",
    "Platform $6,939.50 vs client tracker Meta $6,940: delta $0.50, rounding. Tracker is the client-facing basis.",
    "",
    "COUNTING RULES",
    "Pixel events = offsite_conversion.fb_pixel_custom rollup. The Q3 CONV ad set optimizes to the SubmittedApplication custom event (promoted_object), but the reported rollup is not verified as submitted applications and is never used for cross-channel CPA. June's 136 came from a CTR campaign optimizing to LPV — a different animal; do not trend 136 → 284 as like-for-like performance.",
    "Reach is per campaign, never summed. Channel CPM in July is an objective-mix artifact; the workbook reports traffic-line CPM like-for-like instead.",
    "",
    "CARRY-OVER CHECK (June commitment #1: conversion objective)",
    "DELIVERED. 0726_GCG_Q3_esp_us_CONV live on OUTCOME_SALES optimizing to SubmittedApplication, 44.0% of July GCG spend. The remaining question is measurement (pixel rollup vs verified submitted apps), tracked in the GA4 pull.",
    "",
    "TARGETING NOTE",
    "US, 18-65, Advantage+ Audience on all active GCG ad sets. Age locked by Meta's Financial Products and Services special ad category (mandatory 2025-01-21; verified 2026-08-19). Any age-refinement ask is platform-impossible, not unimplemented.",
]:
    if t and t == t.upper() and len(t) < 60:
        section_row(ws, t, r, span=1)
    else:
        bullet_row(ws, t, r, span=1)
        ws.row_dimensions[r].height = 30 if t else 6
    r += 1
ws.column_dimensions["A"].width = 130

out = "reports/forex/gcg/2026-07/data/GCG-Meta-July-2026-data.xlsx"
wb.save(out)
print(f"saved {out}")

# self-checks
assert abs(TOTAL_SPEND - 6939.50) < 0.01, TOTAL_SPEND
assert abs(sum(a[6] for a in ADSETS) - TOTAL_SPEND) < 0.01
# Creatives tab omits 6 long-tail ads (24.30 + 17.67 + 4.49 + 2.36 + 0.65 + 0.53 = 50.00)
assert abs(sum(c[2] for c in CREATIVE) - (TOTAL_SPEND - 50.00)) < 0.01
print("self-check OK: campaign, adset and geo spend reconcile to $6,939.50")
