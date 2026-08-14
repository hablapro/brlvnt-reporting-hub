#!/usr/bin/env python3
"""GGMI Azerion July 2026 channel workbook.

Sources: two vendor xlsx in data/sources/ (display + native), client tracker
transcription, June figures.json for MoM. Vendor spend is vendor-basis; the
client tracker line is vendor x 1.10 exactly. Client-facing spend = tracker.
"""
import json, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

BASE = Path("reports/forex/ggmi/2026-07")
SRC = BASE / "data/sources"
DISPLAY = SRC / "Forex GGMI (LATAM) — July 2026 Azerion Report.xlsx"
NATIVE = SRC / "Forex GGMI (LATAM) Native — July 2026 Azerion Report.xlsx"
NATIVE_JUNE = SRC / "Forex GGMI (LATAM) Native — June 2026 Azerion Report.xlsx"
OUT = BASE / "data/GGMI-Azerion-July-2026-data.xlsx"

def rows(path, sheet):
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    return [list(r) for r in ws.iter_rows(values_only=True)]

def kpis(path):
    d = {}
    for r in rows(path, "Summary"):
        if r[0] and r[1] is not None and r[0] != "KPI":
            d[str(r[0]).strip()] = r[1]
    return d

disp_k, nat_k, natj_k = kpis(DISPLAY), kpis(NATIVE), kpis(NATIVE_JUNE)
disp_perf = rows(DISPLAY, "Performance")
nat_perf = rows(NATIVE, "Performance")
june = json.load(open("reports/forex/ggmi/2026-06/figures.json"))["figures"]

# ---- QA reconciliations (fail loudly) ----
qa = []
def check(name, a, b, tol=0.02):
    ok = abs(a - b) <= tol
    qa.append((name, round(a, 2), round(b, 2), "PASS" if ok else "FAIL"))
    assert ok, f"{name}: {a} vs {b}"

adset = [r for r in disp_perf[1:14] if r[0] == "LATAM"]
total = disp_perf[13]
check("Display ad-set spend = total", sum(r[2] for r in adset), total[2])
check("Display ad-set impressions = total", sum(r[3] for r in adset), total[3])
check("Display ad-set clicks = total", sum(r[4] for r in adset), total[4])
check("Display ad-set results = total", sum(r[10] for r in adset), total[10])
check("Display total = Summary KPI spend", total[2], disp_k["Spend"])

nat_adset = [r for r in nat_perf[1:8] if r[0] == "LATAM"]
nat_total = nat_perf[7]
check("Native ad-set spend = total", sum(r[2] for r in nat_adset), nat_total[2])
check("Native ad-set clicks = total", sum(r[4] for r in nat_adset), nat_total[4])
check("Native total = Summary KPI spend", nat_total[2], nat_k["Spend"])

dev = rows(DISPLAY, "Diagnostics")
check("Display device spend = total", sum(r[1] for r in dev[2:5]), total[2])
wk = [r for r in disp_perf if r[0] and str(r[0]).startswith("July ")]
check("Display weekly spend = total", sum(r[1] for r in wk), total[2])
check("Display weekly results = total", sum(r[8] for r in wk), total[10])

TRACKER_AZ, TRACKER_NATIVE, QC_NATIVE = 37509, 27630, 10003.47
check("Tracker Azerion = vendor x1.10", disp_k["Spend"] * 1.10, TRACKER_AZ, tol=1.0)
check("Tracker Native = Az native x1.10 + QC native",
      nat_k["Spend"] * 1.10 + QC_NATIVE, TRACKER_NATIVE, tol=3.0)

results = int(disp_k["Results (Conversions)"])
cpa_client = TRACKER_AZ / results

# ---- build ----
wb = Workbook(); wb.remove(wb.active)
def sheet(title, header, data, widths=None):
    ws = wb.create_sheet(title)
    ws.append(header)
    for c in ws[1]: c.font = Font(bold=True)
    for r in data: ws.append(r)
    for i, w in enumerate(widths or []):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = "A2"
    return ws

sheet("Summary", ["Metric", "July 2026", "Basis"], [
    ("Spend (client-facing, tracker)", TRACKER_AZ, "Client budget tracker line 'Azerion'"),
    ("Spend (vendor display)", round(disp_k["Spend"], 2), "Vendor EOM xlsx; tracker = vendor x1.10"),
    ("Impressions (display)", disp_k["Impressions"], "Vendor"),
    ("Clicks (display, Advanced IVT)", disp_k["Clicks"], "Vendor"),
    ("CTR (display)", round(disp_k["CTR"], 6), "Vendor"),
    ("CPM (display)", disp_k["CPM"], "Vendor, flat rate"),
    ("Device reach (display)", round(disp_k["Device Reach"]), "Vendor"),
    ("Results (vendor-reported submitted apps)", results, "Vendor pixel/step tracking, NOT CRM-validated"),
    ("Cost per submitted app (client-facing)", round(cpa_client, 2), "Tracker spend / vendor results (June convention)"),
    ("CPA (vendor basis)", round(disp_k["CPA (Cost Per Result)"], 2), "Vendor"),
    ("Viewability (display)", disp_k["Viewability"], "Vendor; IAB floor 0.70 -> PASS"),
    ("", "", ""),
    ("Native spend (vendor)", round(nat_k["Spend"], 2), "Vendor Native EOM xlsx"),
    ("Native spend in tracker 'Native' line", round(nat_k["Spend"] * 1.10, 2), "x1.10; + Quantcast NativeOnly 10,003.47 = tracker 27,630"),
    ("Native impressions", nat_k["Impressions"], "Vendor"),
    ("Native clicks", nat_k["Clicks"], "Vendor"),
    ("Native CTR", round(nat_k["CTR"], 6), "Vendor"),
    ("Native viewability", nat_k["Viewability"], "Vendor; PASS vs 0.70 floor"),
    ("Native results", "n/a", "Upper-funnel, no conversion column in vendor file"),
], [42, 16, 60])

sheet("Display Performance", [str(c) for c in disp_perf[0]],
      [r for r in disp_perf[1:14]], [10, 26] + [13] * 12)
sheet("Native Performance", [str(c) for c in nat_perf[0][:12]],
      [r[:12] for r in nat_perf[1:8]], [10, 26] + [13] * 10)
sheet("Weekly Trend Display", [str(c) for c in disp_perf[16]],
      [r for r in disp_perf[17:22]], [12] + [13] * 11)
nat_wk_hdr = next(i for i, r in enumerate(nat_perf) if r[0] == "Week")
sheet("Weekly Trend Native", [str(c) for c in nat_perf[nat_wk_hdr][:10]],
      [r[:10] for r in nat_perf[nat_wk_hdr + 1:nat_wk_hdr + 6]], [12] + [13] * 9)
sheet("Devices", ["Channel", "Device", "Spend", "Impressions", "Clicks"],
      [["Display"] + r[:4] for r in dev[2:5]] +
      [["Native"] + r[:4] for r in rows(NATIVE, "Diagnostics")[2:5]], [10, 14, 13, 14, 12])

sheet("Tracker Reconciliation", ["Line", "Vendor", "x1.10", "Tracker", "Delta"], [
    ("Azerion (display)", round(disp_k["Spend"], 2), round(disp_k["Spend"] * 1.10, 2), TRACKER_AZ,
     round(TRACKER_AZ - disp_k["Spend"] * 1.10, 2)),
    ("Native (Azerion part)", round(nat_k["Spend"], 2), round(nat_k["Spend"] * 1.10, 2), "", ""),
    ("Native (+ Quantcast NativeOnly)", QC_NATIVE, "", TRACKER_NATIVE,
     round(TRACKER_NATIVE - (nat_k["Spend"] * 1.10 + QC_NATIVE), 2)),
], [32, 13, 13, 13, 10])

sheet("MoM", ["Metric", "June 2026", "July 2026", "Change %"], [
    ("Spend (client-facing)", june["azerion.spend"], TRACKER_AZ,
     round((TRACKER_AZ / june["azerion.spend"] - 1) * 100, 1)),
    ("Impressions (display)", june["azerion.impressions"], disp_k["Impressions"],
     round((disp_k["Impressions"] / june["azerion.impressions"] - 1) * 100, 1)),
    ("Clicks (display)", june["azerion.clicks"], disp_k["Clicks"],
     round((disp_k["Clicks"] / june["azerion.clicks"] - 1) * 100, 1)),
    ("Submitted apps (vendor-reported)", june["azerion.submitted_apps"], results,
     round((results / june["azerion.submitted_apps"] - 1) * 100, 1)),
    ("Cost per app (client-facing)", june["azerion.cost_per_app"], round(cpa_client, 2),
     round((cpa_client / june["azerion.cost_per_app"] - 1) * 100, 1)),
    ("Native spend (vendor)", round(natj_k.get("Spend", 0), 2), round(nat_k["Spend"], 2), "new/partial June"),
], [34, 14, 14, 14])

sheet("Notes & QA", ["Check / note", "A", "B", "Status"],
      [list(q) for q in qa] + [
    ("", "", "", ""),
    ("Source: vendor EOM email 2026-08-04, Gmail msg 19fce71106bf7329; files untouched in data/sources/", "", "", ""),
    ("Client-facing spend = tracker ($37,509 display; Native line $27,630 shared with Quantcast NativeOnly)", "", "", ""),
    ("Tracker line = vendor x 1.10 exactly, both lines. Internal only; never in client artifacts.", "", "", ""),
    ("Berelvant 7.5% tech fee: internal/billing only, not client-facing, not in vendor billing.", "", "", ""),
    ("Results are VENDOR-reported submitted applications, not CRM-validated (unlike Bing/SA360).", "", "", ""),
    ("GEO: vendor file says 'LATAM', no country breakdown. Mexico-only delivery UNVERIFIABLE from vendor data. Same gap as June. Disclose.", "", "", ""),
    ("Viewability: display 82.47%, native 77.34% - both PASS the 70% IAB floor.", "", "", ""),
    ("Native launched Jun 29; July is its first full month. June native file kept for reference.", "", "", ""),
], [70, 12, 12, 8])

wb.save(OUT)
print("saved", OUT)
for q in qa: print(q)
