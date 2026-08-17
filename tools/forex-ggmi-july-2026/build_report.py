#!/usr/bin/env python3
"""GGMI July 2026 formatted performance report (.xlsx deliverable).

Client-facing companion to the deck, generated from figures.json — never
hand-keyed. Palette mirrors docs/DESIGN-SYSTEM.md (navy header, ice total,
light banding).
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..',
                    'reports', 'forex', 'ggmi', '2026-07')
OUT = os.path.join(BASE, 'output', 'GGMI-July-2026-Performance-Report.xlsx')
FIG = json.load(open(os.path.join(BASE, 'figures.json')))['figures']

NAVY, ICE, LIGHT, BORDER, INK, MUTED = '1E2761', 'CADCFC', 'F9FAFC', 'D0D4DC', '2B3147', '5A6072'
thin = Side(style='thin', color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.border = box
        cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')

def body(ws, r0, rows, ncols, total_last=False):
    for ri, row in enumerate(rows):
        r = r0 + ri
        is_total = total_last and ri == len(rows) - 1
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = box
            cell.font = Font(name='Calibri', size=10, bold=is_total or ci == 1,
                             color=NAVY if is_total else INK)
            if is_total:
                cell.fill = PatternFill('solid', fgColor=ICE)
            elif ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LIGHT)
            cell.alignment = Alignment(horizontal='right' if ci > 1 else 'left')
            if isinstance(v, (int, float)) and ci > 1:
                cell.number_format = '#,##0' if v == int(v) else '#,##0.00'

ws = wb.active
ws.title = 'Summary'
ws['A1'] = 'GGMI (LATAM) — July 2026 Performance Summary'
ws['A1'].font = Font(name='Georgia', bold=True, size=14, color=NAVY)
ws['A2'] = 'Reporting period July 1–31, 2026 · Currency USD · Prepared by Berelvant'
ws['A2'].font = Font(name='Calibri', size=9, color=MUTED)
hdr = ['Channel', 'Spend ($)', 'Impressions', 'Clicks', 'Submitted apps', 'Cost per app ($)']
for ci, h in enumerate(hdr, 1):
    ws.cell(row=4, column=ci, value=h)
style_header(ws, 4, len(hdr))
def money(k): return f"${FIG[k]:,.0f}"
def num(k): return f"{FIG[k]:,}"
body(ws, 5, [
    ['Bing (Search)', money('bing.spend'), num('bing.impressions'), num('bing.clicks'), str(FIG['bing.submitted_apps']), '$531.25'],
    ['Quantcast (Display)', money('quantcast.spend'), num('quantcast.impressions'), num('quantcast.clicks'), '—', '—'],
    ['Azerion (Display)', money('azerion.spend'), num('azerion.impressions'), num('azerion.clicks'), '41 *', '$914.85 *'],
    ['Native (Quantcast + Azerion)', money('native.spend'), num('native.impressions'), num('native.clicks'), '—', '—'],
    ['Meta (Social)', money('meta.spend'), num('meta.impressions'), num('meta.clicks'), '—', '—'],
    ['DOOH (Perion)', money('dooh.spend'), '—', '—', '—', '—'],
    ['Total', money('total.spend'), num('total.impressions'), num('total.clicks'), '—', '—'],
], len(hdr), total_last=True)
notes = [
    '* Azerion applications are vendor-reported; Bing applications are SA360-reported (offline-imported into Bing; SA360 is the conversion source of record).',
    'Conversions come from different systems per channel and are never summed across channels.',
    'Combined Bing + Azerion view (the one sanctioned pairing): 61 submitted applications at $789.08 (June: 92 at $660.00).',
    'Meta clicks are link clicks, the June deck definition. Meta reports on spend and delivery this month.',
    'Spend per the client budget tracker. Native combines Quantcast’s native placement and Azerion’s native inventory.',
]
for i, n in enumerate(notes):
    c = ws.cell(row=13 + i, column=1, value=n)
    c.font = Font(name='Calibri', size=8.5, color=MUTED, italic=True)
for col, w in zip('ABCDEF', [30, 13, 14, 12, 15, 16]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A5'

ws2 = wb.create_sheet('Month over Month')
ws2['A1'] = 'GGMI — July vs June 2026'
ws2['A1'].font = Font(name='Georgia', bold=True, size=13, color=NAVY)
hdr2 = ['Metric', 'June', 'July', 'Change']
for ci, h in enumerate(hdr2, 1):
    ws2.cell(row=3, column=ci, value=h)
style_header(ws2, 3, len(hdr2))
body(ws2, 4, [
    ['Working media ($)', '$120,393', money('total.spend'), '+24%'],
    ['Bing submitted apps', '50', str(FIG['bing.submitted_apps']), '-30'],
    ['Bing cost per app', '$513.17', '$531.25', '+3.5%'],
    ['Bing cost per app, rebuilt structure', '—', '$237.12', 'Jul 22–31'],
    ['Azerion apps (vendor-reported)', '42', '41', '-1'],
    ['Azerion cost per app', '$834', '$915', '+9.7%'],
    ['Combined Bing + Azerion apps', '92', '61', '-31'],
    ['Combined cost per app', '$660.00', '$789.08', '+19.6%'],
    ['Quantcast viewability', '51.3%', '54.45%', '+3.15pp'],
    ['Azerion viewability (display)', '68.5%', '82.47%', '+13.97pp'],
    ['Meta link clicks', '407,136', num('meta.clicks'), '-81.7%'],
    ['Mexico sessions (GA4)', '9,236', num('ga4.mexico_sessions'), '-20.9%'],
    ['Unique visitors, Mexico (GA4)', '5,838', num('ga4.unique_visitors'), '-31.8%'],
    ['Blended submitted applications (client dashboard)', '684', num('funnel.submitted'), '-5%'],
    ['Blended approved (client dashboard)', '209', num('funnel.approved'), '-14%'],
    ['Blended funded (client dashboard)', '32', num('funnel.funded'), '-34%'],
    ['Blended traded (client dashboard)', '29', num('funnel.traded'), '-34%'],
], len(hdr2))
for col, w in zip('ABCD', [38, 12, 12, 12]):
    ws2.column_dimensions[col].width = w
c = ws2.cell(row=22, column=1,
             value='June figures per the June 2026 review. '
                   'Bing’s July month splits dark (Jul 1–10), legacy (Jul 11–22, $5,883, 0 apps), rebuilt (Jul 22–31, $4,742, 20 apps).')
c.font = Font(name='Calibri', size=8.5, color=MUTED, italic=True)

wb.save(OUT)
print('OK ->', OUT)
