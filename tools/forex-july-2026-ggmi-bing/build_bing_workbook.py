#!/usr/bin/env python3
"""Build GGMI Bing July 2026 data workbook. Sources: bing-ads MCP acct 31003116 + sa360 MCP 5372690580/9697709980."""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "reports/forex/ggmi/2026-07/data/GGMI-Bing-July-2026-data.xlsx"
KW = "/Users/rpro/.claude/projects/-Users-rpro-AI-BRLVNT-Brlvnt-Reporting-Analytics-2026/d5ddf545-a0a0-4950-9b09-f9b7b0815326/tool-results/mcp-bing-ads-bing_ads_keyword_performance-1785868586321.txt"

# ---------- house style ----------
NAVY = "1F3864"; BLUE = "2E5E8C"; LGREY = "F2F2F2"; RED = "FFC7CE"; GREEN = "C6EFCE"; AMBER = "FFEB9C"
TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUB   = Font(name="Calibri", size=9, italic=True, color="404040")
HDR   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY  = Font(name="Calibri", size=10)
BOLD  = Font(name="Calibri", size=10, bold=True)
NOTE  = Font(name="Calibri", size=10)
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '"$"#,##0.00'; MONEY0 = '"$"#,##0'; PCT = '0.0%'; INT = '#,##0'

def sheet(wb, name, title, subtitle, widths):
    ws = wb.create_sheet(name)
    ncol = len(widths)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(1, 1, title); c.font = TITLE; c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c = ws.cell(2, 2 - 1, subtitle); c.font = SUB; c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[2].height = 28
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def header(ws, row, labels, freeze=True):
    for i, l in enumerate(labels, 1):
        c = ws.cell(row, i, l); c.font = HDR; c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BOX
    ws.row_dimensions[row].height = 30
    if freeze:
        ws.freeze_panes = ws.cell(row + 1, 1)
    return row + 1

def row(ws, r, vals, fmts, bold=False, fill=None):
    for i, (v, f) in enumerate(zip(vals, fmts), 1):
        c = ws.cell(r, i, v); c.font = BOLD if bold else BODY; c.border = BOX
        if f: c.number_format = f
        if fill: c.fill = PatternFill("solid", fgColor=fill)
        if i == 1: c.alignment = Alignment(horizontal="left", indent=1)
        else: c.alignment = Alignment(horizontal="right")
    return r + 1

def notes(ws, r, lines, col_span=1):
    for t in lines:
        if col_span > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_span)
        c = ws.cell(r, 1, t); c.font = NOTE
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[r].height = max(15, 15 * (1 + len(t) // 110))
        r += 1
    return r

wb = openpyxl.Workbook(); wb.remove(wb.active)

# ============================================================ 1. CAMPAIGNS
CAMPAIGNS = [
    # name, status, phase, spend, impr, clicks, conv(SA360 primary)
    ("FX_LATAM_spanish_AO_GEN_policytest_v2_brlvnt", "Paused Jul 22", "A legacy", 4039.90, 24964, 1796, 0),
    ("FX_LATAM_Spanish_MX_GEN_Tradingview_brlvnt",   "Active",        "B new",    2027.48, 12011,  411, 18),
    ("FX_LATAM_Spanish_MX_GEN_Forex_brlvnt",         "Active",        "B new",    1671.85, 10636,  294, 0),
    ("FX_LATAM_spanish_MX_GEN_BrandGeneric_brlvnt",  "Paused Jul 22", "A legacy", 1473.81, 20445,  940, 0),
    ("FX_LATAM_Spanish_MX_GEN_Upper_funnel_brlvnt",  "Active",        "B new",     786.81, 11764,  382, 2),
    ("FX_LATAM_spanish_MX_GEN_PlatformIntercept_brlvnt","Paused Jul 22","A legacy", 254.18,  7224,  175, 0),
    ("FX_LATAM_Spanish_MX_GEN_MT5_brlvnt",           "Active",        "B new",     168.24,  1242,   31, 0),
    ("FX_LATAM_spanish_AO_GEN_TradingView_brlvnt",   "Paused Jul 22", "A legacy",   92.46,  1002,   28, 0),
    ("FX_LATAM_Spanish_MX_GEN_Competitor_brlvnt",    "Active",        "B new",      52.06,   686,   22, 0),
    ("FX_LATAM_Spanish_MX_BRD_Brand_brlvnt",         "Active",        "B new",      32.29,    18,    5, 0),
    ("FX_LATAM_spanish_AO_Brand_brlvnt",             "Paused Jul 22", "A legacy",   22.22,    99,    4, 0),
    ("FX_LATAM_Spanish_MX_GEN_Silver_brlvnt",        "Active",        "B new",       3.64,    49,    2, 0),
    ("FX_LATAM_Spanish_MX_GEN_MT4_brlvnt",           "Active",        "B new",       0.00,   251,    0, 0),
    ("FX_LATAM_Spanish_MX_GEN_Gold_brlvnt",          "Active",        "B new",       0.00,     3,    0, 0),
]
ws = sheet(wb, "Campaigns", "GGMI Bing (LATAM) — July 2026 Campaign Performance",
    "Bing acct 31003116 (FOREX.com LATAM) + SA360 customer 5372690580 / login 9697709980. USD. "
    "Spend/impr/clicks from Bing-direct; Conversions from SA360 metrics.conversions (Primary = submitted applications) "
    "because GGMI conversions are offline-imported and read 0 in the Bing API.",
    [46, 14, 11, 13, 11, 10, 9, 11, 8, 13])
r = header(ws, 4, ["Campaign", "Status", "Phase", "Spend", "Impressions", "Clicks", "CTR", "Avg CPC", "Conv", "CPA"])
for n, st, ph, sp, im, cl, cv in CAMPAIGNS:
    ctr = cl / im if im else 0
    cpc = sp / cl if cl else 0
    cpa = sp / cv if cv else None
    fill = GREEN if cv else None
    r = row(ws, r, [n.replace("FX_LATAM_spanish_", "").replace("FX_LATAM_Spanish_", ""), st, ph, sp, im, cl, ctr, cpc, cv, cpa if cpa else "n/a"],
            [None, None, None, MONEY, INT, INT, PCT, MONEY, INT, MONEY], fill=fill)
TS = sum(c[3] for c in CAMPAIGNS); TI = sum(c[4] for c in CAMPAIGNS)
TC = sum(c[5] for c in CAMPAIGNS); TV = sum(c[6] for c in CAMPAIGNS)
r = row(ws, r, ["TOTAL", "", "", TS, TI, TC, TC/TI, TS/TC, TV, TS/TV],
        [None, None, None, MONEY, INT, INT, PCT, MONEY, INT, MONEY], bold=True, fill=LGREY)
r += 1
r = notes(ws, r, [
    "READ: July is not one month of like-for-like delivery. The account was dark Jul 1-10, ran the legacy set Jul 11-22, "
    "then the rebuilt Mexico-only MX_ structure took over Jul 22-31. Grade the two phases separately (see 'Phases' tab), not the month.",
    "Only two campaigns produced submitted applications, both in the new structure: MX_GEN_Tradingview (18) and MX_GEN_Upper_funnel (2).",
    "The five legacy campaigns spent $5,882.57 across 11 days and returned 0 submitted applications. See 'Funnel' tab — they generated 207 application "
    "starts but nothing past Step 2, so this is a funnel collapse, not purely an import-lag artifact.",
], 10)

# ============================================================ 2. PHASES
ws = sheet(wb, "Phases", "GGMI Bing — July Phase Split vs June Baseline",
    "The like-for-like view. Phase A = the legacy campaigns that carried June, run Jul 11-22. Phase B = the rebuilt Mexico-only "
    "MX_ structure, live Jul 22-31. June column is the full-month SA360 baseline (re-verified 2026-08-04, unchanged at 50 conv).",
    [34, 15, 15, 15, 15])
r = header(ws, 4, ["Metric", "June 2026 (full)", "July Phase A (legacy, Jul 11-22)", "July Phase B (new MX_, Jul 22-31)", "July total"])
PH = {
    "A": [c for c in CAMPAIGNS if c[2] == "A legacy"],
    "B": [c for c in CAMPAIGNS if c[2] == "B new"],
}
aS = sum(c[3] for c in PH["A"]); aC = sum(c[5] for c in PH["A"]); aI = sum(c[4] for c in PH["A"]); aV = sum(c[6] for c in PH["A"])
bS = sum(c[3] for c in PH["B"]); bC = sum(c[5] for c in PH["B"]); bI = sum(c[4] for c in PH["B"]); bV = sum(c[6] for c in PH["B"])
JUN = dict(spend=25658.61, impr=466582, clicks=21480, conv=50, days=30)
r = row(ws, r, ["Days delivering", JUN["days"], 12, 10, 21], [None, INT, INT, INT, INT])
r = row(ws, r, ["Spend", JUN["spend"], aS, bS, TS], [None, MONEY, MONEY, MONEY, MONEY])
r = row(ws, r, ["Spend / day", JUN["spend"]/30, aS/12, bS/10, TS/21], [None, MONEY, MONEY, MONEY, MONEY])
r = row(ws, r, ["Impressions", JUN["impr"], aI, bI, TI], [None, INT, INT, INT, INT])
r = row(ws, r, ["Clicks", JUN["clicks"], aC, bC, TC], [None, INT, INT, INT, INT])
r = row(ws, r, ["CTR", JUN["clicks"]/JUN["impr"], aC/aI, bC/bI, TC/TI], [None, PCT, PCT, PCT, PCT])
r = row(ws, r, ["Avg CPC", JUN["spend"]/JUN["clicks"], aS/aC, bS/bC, TS/TC], [None, MONEY, MONEY, MONEY, MONEY])
r = row(ws, r, ["Submitted applications", JUN["conv"], aV, bV, TV], [None, INT, INT, INT, INT], bold=True)
r = row(ws, r, ["CPA", JUN["spend"]/JUN["conv"], "n/a — 0 conv", bS/bV, TS/TV], [None, MONEY, None, MONEY, MONEY], bold=True)
r = row(ws, r, ["CPA, converting campaigns only", JUN["spend"]/JUN["conv"], "n/a",
                (2027.48+786.81)/bV, (2027.48+786.81)/TV], [None, MONEY, None, MONEY, MONEY])
r = row(ws, r, ["Mexico share of spend", 0.5075, (4039.90+1473.81+9.51+58.23+16.11)/aS, 1.0, 10339.93/TS],
        [None, PCT, PCT, PCT, PCT], fill=GREEN)
r += 1
r = notes(ws, r, [
    "THE JULY RESULT: the rebuilt Mexico-only structure delivered 20 submitted applications on $4,742 in 10 days — $237.12 cost per "
    "submitted app account-wide, $140.71 across the two campaigns that actually converted, against a June baseline of $513.17. That is the number to carry "
    "into the July report, and it is earned on 100% Mexico delivery.",
    "PHASE A IS THE COST: $5,882.57 spent Jul 11-22 on the legacy set for zero submitted applications. Roughly 55% of the month's spend bought no "
    "measurable primary outcome. At June's $513 cost per submitted app that spend would have predicted ~11 submissions.",
    "Caveat on Phase B CPA: 10 days is a short read and the newest clicks are ~4 days old at pull time, so Phase B conversions may still rise. "
    "Direction is safe; treat $237 as a ceiling estimate that should improve, not a settled monthly cost per submitted app.",
    "June baseline re-verified today at 27/9/14 = 50 conversions, byte-identical to what the June report booked. June is mature; no restatement needed.",
], 5)

# ============================================================ 3. DAILY
DAILY = [
    ("2026-07-01", 0, 0, 0, 0, "dark"), ("2026-07-02", 0, 0, 0, 0, "dark"), ("2026-07-03", 0, 0, 0, 0, "dark"),
    ("2026-07-04", 0, 0, 0, 0, "dark"), ("2026-07-05", 0, 0, 0, 0, "dark"), ("2026-07-06", 0, 0, 0, 0, "dark"),
    ("2026-07-07", 0, 0, 0, 0, "dark"), ("2026-07-08", 0, 0, 0, 0, "dark"), ("2026-07-09", 0, 0, 0, 0, "dark"),
    ("2026-07-10", 0, 0, 0, 0, "dark"),
    ("2026-07-11", 436.55, 5106, 416, 0, "A legacy"), ("2026-07-12", 364.63, 5142, 173, 0, "A legacy"),
    ("2026-07-13", 441.88, 4213, 162, 0, "A legacy"), ("2026-07-14", 562.23, 5449, 222, 0, "A legacy"),
    ("2026-07-15", 559.90, 4676, 252, 0, "A legacy"), ("2026-07-16", 699.13, 6234, 342, 0, "A legacy"),
    ("2026-07-17", 614.23, 4505, 263, 0, "A legacy"), ("2026-07-18", 670.31, 4353, 464, 0, "A legacy"),
    ("2026-07-19", 406.84, 3071, 199, 0, "A legacy"), ("2026-07-20", 518.09, 4364, 203, 0, "A legacy"),
    ("2026-07-21", 448.17, 4226, 177, 0, "A legacy"),
    ("2026-07-22", 577.22, 6519, 210, 0, "crossover"),
    ("2026-07-23", 639.51, 5874, 210, 1, "B new"), ("2026-07-24", 322.88, 3548, 102, 1, "B new"),
    ("2026-07-25", 230.02, 1934, 44, 1, "B new"), ("2026-07-26", 218.30, 1902, 44, 1, "B new"),
    ("2026-07-27", 582.18, 4159, 119, 2, "B new"), ("2026-07-28", 559.63, 4247, 122, 0, "B new"),
    ("2026-07-29", 723.75, 4550, 152, 6, "B new"), ("2026-07-30", 607.16, 4569, 128, 6, "B new"),
    ("2026-07-31", 442.33, 3684, 86, 2, "B new"),
]
ws = sheet(wb, "Daily Trend", "GGMI Bing — Daily Delivery, July 2026",
    "SA360 campaign x segments.date, Jul 1-31. Dates with no delivery are shown explicitly as zeros so the 10-day dark period is visible. "
    "Phase tags: dark = no delivery, A legacy = pre-rebuild campaigns, crossover = both sets live, B new = Mexico-only MX_ structure.",
    [13, 13, 13, 10, 9, 14])
r = header(ws, 4, ["Date", "Spend", "Impressions", "Clicks", "Conv", "Phase"])
for d, sp, im, cl, cv, ph in DAILY:
    fill = LGREY if ph == "dark" else (GREEN if cv else (AMBER if ph == "crossover" else None))
    r = row(ws, r, [d, sp, im, cl, cv, ph], [None, MONEY, INT, INT, INT, None], fill=fill)
r = row(ws, r, ["TOTAL", sum(x[1] for x in DAILY), sum(x[2] for x in DAILY), sum(x[3] for x in DAILY), sum(x[4] for x in DAILY), ""],
        [None, MONEY, INT, INT, INT, None], bold=True, fill=LGREY)
r += 1
r = notes(ws, r, [
    "Jul 1-10: zero delivery. The account was still dark from the June pause; campaigns were relaunched Sat Jul 11.",
    "Jul 22: the only day both structures delivered. Legacy campaigns end Jul 22; MX_GEN_Tradingview/Forex/Upper_funnel/MT5/MT4/Silver start Jul 22, "
    "Competitor and Gold Jul 24, MX_BRD_Brand Jul 24.",
    "All 20 submitted applications land Jul 23-31, none before. They concentrate late (Jul 29-31 = 14 of 20), consistent with the new structure "
    "ramping and with import recency.",
    "Effective delivery days in July: 21 of 31 (68%). Any 'July vs June' spend comparison that ignores this reads as a budget cut when it was a 10-day outage.",
], 6)

# ============================================================ 4. GEO
GEO = [  # country, clicks, spend, status
    ("Mexico", 3902, 10339.93, "TARGET"),
    ("Venezuela", 137, 169.76, "forbidden"), ("Nicaragua", 15, 41.25, "forbidden"),
    ("Australia", 7, 17.51, "forbidden"), ("Colombia", 7, 14.05, "forbidden"),
    ("Dominican Republic", 5, 13.98, "forbidden"), ("United States", 4, 12.19, "forbidden"),
    ("Panama", 3, 6.14, "forbidden"), ("Argentina", 1, 3.81, "forbidden"),
    ("Brazil", 4, 3.42, "forbidden"), ("Peru", 2, 1.42, "forbidden"),
    ("Mauritius", 1, 0.85, "forbidden"), ("Chile", 2, 0.63, "forbidden"),
]
ws = sheet(wb, "Geo Compliance", "GGMI Bing — Delivery by User Location, July 2026 (Mexico-only mandate)",
    "SA360 user_location_view, targeting_location = false (where the user physically was), Jul 1-31, 69 country rows returned. "
    "GGMI is contracted to Mexico only; any non-Mexico delivery is a compliance breach, not just inefficiency.",
    [26, 11, 13, 13, 14])
r = header(ws, 4, ["Country", "Clicks", "Spend", "% of spend", "Status"])
for c, cl, sp, st in GEO:
    fill = GREEN if st == "TARGET" else RED
    r = row(ws, r, [c, cl, sp, sp / TS, st], [None, INT, MONEY, PCT, None], fill=fill)
r = row(ws, r, ["TOTAL", sum(g[1] for g in GEO), sum(g[2] for g in GEO), 1.0, ""],
        [None, INT, MONEY, PCT, None], bold=True, fill=LGREY)
r += 1
r = header(ws, r, ["Campaign", "Mexico $", "non-MX $", "Mexico %", "Verdict"], freeze=False)
GEOC = [
    ("AO_GEN_policytest_v2 (legacy)", 4039.90, 0.00), ("MX_GEN_Tradingview (new)", 2027.48, 0.00),
    ("MX_GEN_Forex (new)", 1671.85, 0.00), ("MX_GEN_BrandGeneric (legacy)", 1473.81, 0.00),
    ("MX_GEN_Upper_funnel (new)", 786.81, 0.00), ("MX_GEN_MT5 (new)", 168.24, 0.00),
    ("MX_GEN_Competitor (new)", 52.06, 0.00), ("MX_BRD_Brand (new)", 32.29, 0.00),
    ("MX_GEN_Silver (new)", 3.64, 0.00), ("MX_GEN_MT4 / Gold (new)", 0.00, 0.00),
    ("AO_Brand (legacy)", 16.11, 6.11), ("AO_GEN_TradingView (legacy)", 58.23, 34.23),
    ("MX_GEN_PlatformIntercept (legacy)", 9.51, 244.67),
]
for n, mx, nmx in GEOC:
    tot = mx + nmx
    pct = mx / tot if tot else 1.0
    v = "clean" if nmx == 0 else ("LEAK" if pct < 0.9 else "minor leak")
    r = row(ws, r, [n, mx, nmx, pct, v], [None, MONEY, MONEY, PCT, None],
            fill=GREEN if nmx == 0 else RED)
r += 1
r = notes(ws, r, [
    "THE BREACH IS CLOSED. Non-Mexico delivery is $285.01 of $10,624.94 = 2.7% of July spend, against 49% in June and 51.6% in the "
    "trailing-30-day read at the Jul 13 audit. Every one of the 6 new MX_ campaigns delivered 100% Mexico.",
    "All residual leakage sits in three legacy campaigns that were paused Jul 22: PlatformIntercept ($244.67, 96% of the leak, almost all Venezuela), "
    "AO_GEN_TradingView ($34.23) and AO_Brand ($6.11). With those paused, the structural cause — positiveGeoTargetType = PRESENCE_OR_INTEREST — "
    "no longer has a live campaign to act on.",
    "VERIFIED 2026-08-04 — THE SETTING WAS NEVER FIXED. All 9 currently enabled campaigns still carry positiveGeoTargetType = PRESENCE_OR_INTEREST "
    "(negativeGeoTargetType = PRESENCE, which is correct). July's 100%-Mexico result is therefore a consequence of low volume and of the leaking "
    "legacy campaigns being paused — NOT of a corrected setting. The breach is DORMANT, NOT FIXED, and will return as volume scales. "
    "Do not tell the client the geo issue is resolved. See the 'Config Audit' tab.",
    "56 further countries appear with impressions but $0 spend (Spain, UK, Germany, India, Japan and others), all on the same three legacy campaigns. "
    "Impression-only spill costs nothing but confirms the targeting was open.",
    "Conversions do not carry geo in an offline import, so no CPA-by-country view is possible. This is a delivery view only.",
], 5)

# ============================================================ 5. FUNNEL
FUNNEL = [
    # campaign, phase, sitewide, g2_s1, g2_s2, g2_s3, g2_s4, g2_live, mt5_live, demo
    ("MX_GEN_Tradingview", "B new", 776, 134, 74, 16, 16, 16, 2, 12),
    ("MX_GEN_Upper_funnel", "B new", 358, 29, 9, 2, 2, 2, 0, 2),
    ("MX_GEN_Forex", "B new", 405, 15, 4, 4, 2, 0, 0, 4),
    ("MX_GEN_MT5", "B new", 112, 4, 0, 0, 0, 0, 0, 0),
    ("MX_GEN_Competitor", "B new", 18, 0, 0, 0, 0, 0, 0, 0),
    ("MX_BRD_Brand", "B new", 13, 0, 0, 0, 0, 0, 0, 0),
    ("AO_GEN_policytest_v2", "A legacy", 1525, 104, 28, 0, 0, 0, 0, 27),
    ("MX_GEN_PlatformIntercept", "A legacy", 838, 81, 2, 0, 0, 0, 0, 0),
    ("MX_GEN_BrandGeneric", "A legacy", 540, 22, 6, 0, 0, 0, 0, 0),
    ("AO_GEN_TradingView", "A legacy", 8, 2, 0, 0, 0, 0, 0, 0),
    ("AO_Brand", "A legacy", 62, 0, 0, 0, 0, 0, 0, 0),
]
ws = sheet(wb, "Funnel", "GGMI Bing — Application Funnel by Campaign, July 2026",
    "SA360 campaign x segments.conversion_action_name, all_conversions, Jul 1-31. G2 = G2 Raw Spread account path. "
    "Duplicate conversion actions carrying a timestamp suffix were excluded to avoid double-counting (see Notes & QA). "
    "'Submitted' is the confirmation/ThankYou page that fires on application submission — the primary agency KPI and the only action counted in "
    "metrics.conversions. Approved / Funded / Traded are separate downstream GCLID goals and are NOT in this table.",
    [30, 11, 11, 10, 10, 10, 10, 11, 11, 10, 13])
r = header(ws, 4, ["Campaign", "Phase", "Sitewide", "G2 Step 1", "G2 Step 2", "G2 Step 3", "G2 Step 4",
                   "G2 submitted", "MT5 submitted", "Demo", "Step1 to submitted"])
for n, ph, sw, s1, s2, s3, s4, lv, ml, dm in FUNNEL:
    conv = (lv + ml) / s1 if s1 else None
    fill = GREEN if (lv + ml) else (RED if s1 else None)
    r = row(ws, r, [n, ph, sw, s1, s2, s3, s4, lv, ml, dm, conv if conv is not None else "n/a"],
            [None, None, INT, INT, INT, INT, INT, INT, INT, INT, PCT], fill=fill)
tot = [sum(f[i] for f in FUNNEL) for i in range(2, 10)]
r = row(ws, r, ["TOTAL", ""] + tot + [(tot[5] + tot[6]) / tot[1]],
        [None, None, INT, INT, INT, INT, INT, INT, INT, INT, PCT], bold=True, fill=LGREY)
r += 1
r = notes(ws, r, [
    "THIS IS THE MOST IMPORTANT TAB IN THE PULL. The legacy campaigns did generate demand — 207 application starts (Step 1) across "
    "policytest_v2, PlatformIntercept and BrandGeneric — and converted none of them past Step 2. Zero Step 3, zero Step 4, zero submissions.",
    "That is not import lag. Steps 1 and 2 imported normally for those same campaigns through the same pipeline, and the new campaigns' Step 3, "
    "Step 4 and submissions imported fine over a shorter window. A pipeline that delivers Step 2 but never Step 3 for three campaigns "
    "over 11 days is a funnel failure, not a reporting delay.",
    "By contrast MX_GEN_Tradingview carries the whole path: 134 starts, 74 Step 2, 16 Step 3, 16 Step 4, 16 submitted G2 + 2 submitted MT5. "
    "A 13.4% start-to-submitted rate is the benchmark to hold the rest of the account to.",
    "MX_GEN_Forex is the clearest open opportunity: $1,671.85 spent, 15 starts, reaches Step 4 twice and then stops at zero submissions. "
    "Second-largest spend in the new structure with no primary outcome yet.",
    "This corroborates the GCG Q2 post-mortem finding that the start-to-submit funnel, not top-of-funnel volume, is where GGMI loses accounts. "
    "Same pattern, different account, one month later.",
], 11)

# ============================================================ 6. KEYWORDS
with open(KW) as f:
    kws = json.load(f)
kws = [k for k in kws if k["Spend"] > 0]
kws.sort(key=lambda k: -k["Spend"])
ws = sheet(wb, "Keywords", "GGMI Bing — July 2026 Keyword Detail",
    "bing-ads MCP keyword performance, acct 31003116, Jul 1-31. All keywords with spend > $0, sorted by spend desc. "
    "104 keywords returned, 67 with spend. Keyword spend reconciles to $10,624.94 against the campaign total (PASS). "
    "Conversions read 0 at keyword level because GGMI conversions are offline-imported — use the Funnel tab for outcomes.",
    [30, 26, 13, 11, 11, 12, 10, 10, 8, 34])
r = header(ws, 4, ["Keyword", "Ad group", "Match", "Spend", "Impressions", "Clicks", "CTR", "Avg CPC", "QS", "Campaign"])
for k in kws:
    try:
        qs = int(k.get("QualityScore"))
    except (TypeError, ValueError):
        qs = None
    fill = RED if (qs is not None and qs <= 4) else None
    r = row(ws, r, [k["Keyword"], k["AdGroupName"], k["BidMatchType"], k["Spend"], k["Impressions"], k["Clicks"],
                    k["Clicks"] / k["Impressions"] if k["Impressions"] else 0,
                    k["AverageCpc"], qs, k["CampaignName"].replace("FX_LATAM_spanish_", "").replace("FX_LATAM_Spanish_", "")],
            [None, None, None, MONEY, INT, INT, PCT, MONEY, INT, None], fill=fill)
r = row(ws, r, ["TOTAL (spend > $0)", "", "", sum(k["Spend"] for k in kws), sum(k["Impressions"] for k in kws),
                sum(k["Clicks"] for k in kws), "", "", "", ""],
        [None, None, None, MONEY, INT, INT, None, None, None, None], bold=True, fill=LGREY)
r += 1
r = notes(ws, r, [
    "Spend is highly concentrated: 'trading online' (phrase, legacy policytest_v2) took $2,338.37 = 22% of the month, and returned no confirmation. "
    "'tradingview' exact in the new MX_GEN_Tradingview campaign took $1,566.66 and sits in the ad group that produced 17 of the month's 20 confirmations.",
    "Four keywords carry Quality Score 3 with meaningful impressions, all on the MetaTrader theme: 'mt4' and 'metatrader 4' and 'mt5' in "
    "PlatformIntercept, 'mt5' in policytest_v2. The MT theme has been a persistent QS problem since June; it needs ad and landing relevance work, not bid cuts.",
    "'metatrader' phrase carries QS 10 and drove 146 clicks for $196.39 in PlatformIntercept — the MT intent is healthy, the specific model-number "
    "keywords are what score badly.",
], 10)

# ============================================================ 7. GOALS
GOALS = [
    ("40059107", "GGMI - Sitewide", "Active", "RecordingConversions", True),
    ("40059122", "GGMI - G2 Raw Spread - App Form Step 1", "Active", "RecordingConversions", True),
    ("40059144", "GGMI - G2 Raw Spread - App Form Step 2", "Active", "RecordingConversions", True),
    ("40059152", "GGMI - MT5 Raw Spread - App Form Step 1", "Active", "RecordingConversions", True),
    ("40059170", "GGMI - G2 - Demo Confirmation", "Active", "RecordingConversions", True),
    ("40059184", "GGMI - G2 Raw Spread - Live Confirmation = SUBMITTED APP (PRIMARY KPI)", "Active", "RecordingConversions", True),
    ("40059257", "GGMI - MT5 Raw Spread - Live Confirmation = SUBMITTED APP (PRIMARY KPI)", "Active", "RecordingConversions", True),
]
ws = sheet(wb, "Conversion Goals", "GGMI Bing — Conversion Goal & Bidding-Signal Status (2026-08-04)",
    "bing-ads MCP conversion goals, acct 31003116. The 7 GGMI offline-import goals only; ~40 legacy CIMA/US/UK goals are omitted "
    "(all Paused or TagInactive). Compare against the Jul 13 audit, which found the two primary goals stale.",
    [12, 48, 12, 24, 20, 22])
r = header(ws, 4, ["Goal ID", "Goal name", "Status", "Tracking status", "ExcludeFromBidding", "Change vs Jul 13"])
for gid, n, st, ts, ex in GOALS:
    chg = "FIXED — was NoRecentConversions" if "SUBMITTED APP" in n else "unchanged"
    r = row(ws, r, [gid, n, st, ts, "TRUE — blocks bidding", chg], [None]*6,
            fill=GREEN if "FIXED" in chg else None)
r += 1
r = notes(ws, r, [
    "RESOLVED since the Jul 13 audit: both primary submitted-application goals (40059184, 40059257) now report RecordingConversions. "
    "On Jul 13 they read NoRecentConversions and the audit flagged the primary-KPI import as possibly broken. The offline upload is current. "
    "All 7 GGMI goals are Active and recording.",
    "STILL OPEN — the bidding signal is still switched off. All 7 goals carry ExcludeFromBidding = TRUE, so no conversion data reaches any bid "
    "strategy. Every campaign runs MANUAL_CPC. The account is now measuring correctly and still bidding blind. This is the third consecutive month "
    "this item has been carried (June report, Jul 6 remediation row 3, Jul 13 audit item 1).",
    "Consequence for July's result: the $140.71 cost per submitted app on MX_GEN_Tradingview was achieved with manual bidding and no conversion feedback. "
    "That is upside, not a ceiling — but it also means nothing is automatically protecting it.",
], 6)


# ============================================================ 7b. CONFIG AUDIT
CONFIG = [
    # campaign, monthly budget, Jul spend, Aug1-4 spend, geo type, ad group language, exact bid, phrase bid
    ("MX_GEN_Tradingview",  30000, 2027.48, 790.43, "PRESENCE_OR_INTEREST", "English",     8,    7),
    ("MX_GEN_Forex",        30000, 1671.85, 393.38, "PRESENCE_OR_INTEREST", "(not set)",   20,   15),
    ("MX_GEN_Upper_funnel", 30000,  786.81, 159.21, "PRESENCE_OR_INTEREST", "not checked", None, None),
    ("MX_GEN_MT5",          30000,  168.24,  84.60, "PRESENCE_OR_INTEREST", "not checked", None, None),
    ("MX_GEN_Competitor",    2800,   52.06,  41.10, "PRESENCE_OR_INTEREST", "not checked", None, None),
    ("MX_BRD_Brand",        12000,   32.29,  13.13, "PRESENCE_OR_INTEREST", "not checked", None, None),
    ("MX_GEN_Silver",        6000,    3.64,   0.00, "PRESENCE_OR_INTEREST", "not checked", None, None),
    ("MX_GEN_MT4",           None,    0.00,  18.27, "PRESENCE_OR_INTEREST", "not checked", None, None),
    ("MX_GEN_Gold",          None,    0.00,   0.00, "PRESENCE_OR_INTEREST", "not checked", None, None),
]
ws = sheet(wb, "Config Audit", "GGMI Bing — Live Configuration Audit of the 9 Enabled Campaigns (2026-08-04)",
    "sa360 MCP campaign.geo_target_type_setting + campaign_criterion (LOCATION), bing-ads MCP ad groups and budget utilization. "
    "This tab is the evidence for what is still misconfigured after the July rebuild. Budget column is the Bing MonthlyBudget field.",
    [24, 15, 13, 13, 24, 15, 11, 11, 15])
r = header(ws, 4, ["Campaign", "Monthly budget", "Jul spend", "Aug 1-4 spend", "positiveGeoTargetType",
                   "Ad group language", "Exact bid", "Phrase bid", "Budget used (Jul)"])
for n, bud, jul, aug, geo, lang, eb, pb in CONFIG:
    used = jul / bud if bud else None
    r = row(ws, r, [n, bud if bud else "not returned", jul, aug, geo, lang,
                    eb if eb else "not checked", pb if pb else "not checked", used if used is not None else "n/a"],
            [None, MONEY0, MONEY, MONEY, None, None, MONEY, MONEY, PCT], fill=RED)
r += 1
r = header(ws, r, ["Setting", "Value found", "Correct?", "Comment"], freeze=False)
AUDIT = [
    ("positiveGeoTargetType", "PRESENCE_OR_INTEREST on all 9", "NO",
     "Must be PRESENCE for a Mexico-only mandate. Serves people merely 'interested in' Mexico. Root cause of the June/July breach, still unfixed."),
    ("negativeGeoTargetType", "PRESENCE on all 9", "yes", "Correct."),
    ("Location targets (positive)", "Mexico (2484) + Mexico City (20703)", "redundant",
     "Mexico City is nested inside Mexico. Harmless but sloppy; remove the city row. MX_GEN_Competitor instead carries 2484 + 9450400, "
     "an ID that did not resolve in the geo_target_constant lookup — verify what it is."),
    ("Location targets (negative)", "Canada, Guatemala, United States", "incomplete",
     "Venezuela is NOT excluded, and Venezuela was the single largest source of forbidden delivery in both June and July. Spain, Colombia, "
     "Argentina, Peru, Nicaragua also not excluded. Note: with positiveGeoTargetType = PRESENCE these negatives become unnecessary — "
     "fix the setting rather than chasing countries."),
    ("Ad group language", "English on MX_GEN_Tradingview; not set on MX_GEN_Forex", "NO",
     "These are Spanish-language campaigns targeting Mexico. Neither ad group is set to Spanish and the two differ from each other. "
     "Only 2 of 9 campaigns were checked — audit the remaining 7."),
    ("Network", "OwnedAndOperatedOnly", "yes", "Bing search only, no syndicated partners. Reasonable quality choice; keep."),
    ("Ad rotation", "OptimizeForClicks", "NO",
     "Optimising for clicks, not conversions. Consistent with the account having no usable conversion signal."),
    ("Bid strategy", "InheritFromParent -> EnhancedCpc", "conflicted",
     "Bing reports Enhanced CPC while SA360 reports MANUAL_CPC — the June/July reconcile discrepancy persists. Either way eCPC cannot work: "
     "all 7 conversion goals are ExcludeFromBidding = TRUE, so there is no signal to enhance against."),
    ("CPC bids", "Forex $20 exact / $15 phrase vs Tradingview $8 / $7", "review",
     "Forex bids are ~2.5x Tradingview's on the campaign with the weaker conversion record ($1,671.85 and 0 submitted apps in July). "
     "Tradingview delivered 18 at lower bids."),
    ("RSA creative depth", "COULD NOT PULL", "unknown",
     "bing_ads_list_ads is broken (400 NullRequest, server-side). June's finding of 4 headlines / 2 descriptions against the 8+/3+ standard "
     "remains unverified. Recorded in KNOWN-BUGS.md."),
]
for a, b, c, d in AUDIT:
    fill = GREEN if c == "yes" else (RED if c == "NO" else AMBER)
    r = row(ws, r, [a, b, c, d], [None, None, None, None], fill=fill)
r += 1
r = notes(ws, r, [
    "Budgets are enormous relative to delivery: roughly $140,800/month is configured across the enabled campaigns against $4,742 of actual "
    "Phase B spend. The account is delivery-constrained, not budget-constrained — raising budgets will not raise volume. Impression share and "
    "bid/quality work will.",
], 9)

# ============================================================ 7c. SEARCH QUERIES
import json as _json
SQ = "/Users/rpro/.claude/projects/-Users-rpro-AI-BRLVNT-Brlvnt-Reporting-Analytics-2026/d5ddf545-a0a0-4950-9b09-f9b7b0815326/tool-results/mcp-bing-ads-bing_ads_search_term_report-1785870263824.txt"
import re as _re
qs = [q for q in _json.load(open(SQ)) if q.get("SearchQuery")]
LOW = _re.compile(r"(?i)gratis|descargar|download|curso|manual|pdf|paper trading|prueba gratis|iniciar sesi|p[aá]gina oficial|tradingview\.com|que es|c[oó]mo operar|como operar|simulador")
COMP = _re.compile(r"(?i)oanda|etoro|libertex|capital\.com|fxpro|hfm|naga|exness|xtb|plus500|avatrade|ig\.com|webull|xm broker|forex factory")
def cls(q):
    if COMP.search(q): return "competitor / off-brand"
    if LOW.search(q): return "low intent (free / download / info / navigational)"
    return "commercial"
for q in qs: q["_c"] = cls(q["SearchQuery"])
qs.sort(key=lambda q: -q["Spend"])
ws = sheet(wb, "Search Queries", "GGMI Bing — Search Query Detail, Jul 25 - Aug 4 2026",
    "bing-ads MCP search term report, full account, Jul 25 - Aug 4 (the window covering the new MX_ structure). "
    "Covered spend $2,812.87 across 129 queries — a partial view, not the full month, because Bing only reports queries above a volume "
    "threshold. Intent classification is ours, applied by pattern match, and is directional not definitive.",
    [42, 26, 13, 10, 11, 10, 26, 44])
r = header(ws, 4, ["Search query", "Matched keyword", "Spend", "Clicks", "Impr", "CTR", "Campaign", "Intent classification"])
for q in qs[:70]:
    fill = RED if q["_c"] != "commercial" else None
    r = row(ws, r, [q["SearchQuery"], q["Keyword"], q["Spend"], q["Clicks"], q["Impressions"],
                    q["Clicks"]/q["Impressions"] if q["Impressions"] else 0,
                    q["CampaignName"].replace("FX_LATAM_spanish_","").replace("FX_LATAM_Spanish_",""), q["_c"]],
            [None, None, MONEY, INT, INT, PCT, None, None], fill=fill)
r += 1
tot = sum(q["Spend"] for q in qs)
for c in ["commercial", "low intent (free / download / info / navigational)", "competitor / off-brand"]:
    sub = sum(q["Spend"] for q in qs if q["_c"] == c)
    r = row(ws, r, [c, "", sub, "", "", sub/tot, "", ""], [None, None, MONEY, None, None, PCT, None, None],
            bold=True, fill=GREEN if c == "commercial" else RED)
r += 1
r = notes(ws, r, [
    "ONE QUERY IS THE ACCOUNT. 'tradingview' alone took $1,468.54 on 277 clicks — 52% of all covered spend. The TradingView theme is also "
    "the only thing converting (17 of July's 20 submitted applications came from the MX_GEN_Tradingview 'Exact' ad group). Do NOT cut this theme.",
    "But a meaningful slice of that TradingView traffic has no commercial intent for a broker: 'tradingview.com' ($164.71 across two keywords), "
    "'tradingview iniciar sesion' + 'iniciar sesion gratis' ($93.72 — people trying to log in to TradingView), 'tradingview pagina oficial' ($16.01). "
    "That is roughly $275 of navigational traffic looking for TradingView the product, not a trading account.",
    "Add-negatives list, highest value first: gratis, descargar, download, 'iniciar sesion', 'pagina oficial', 'tradingview.com', 'paper trading', "
    "curso, manual, pdf, 'que es'. Keep the core 'tradingview' and 'trading view' terms. This is the 'fix relevance, do not cut' path already "
    "agreed for this theme in June.",
    "MX_GEN_MT5 is the worst offender proportionally: ~20% of its covered spend went to download/manual intent ('metatrader 5 descargar pc', "
    "'manual metatrader 5 espanol pdf', 'download mt5'), plus 'mp5 x7 descargar' (an audio-player query, wholly irrelevant) and "
    "'xm broker descargar para pc' (a competitor). This is also where the QS 3 keywords sit.",
    "MX_GEN_Forex is paying for informational queries: 'que es forex', 'que es el forex trading', 'forex que es y como operar'. Those belong in "
    "the upper-funnel campaign at upper-funnel bids, not on a $20 exact bid.",
    "Coverage caveat: $2,812.87 of covered spend against $10,624.94 of July spend. Treat the percentages as a sample. A full search-term "
    "export from the Bing UI would be needed to size the waste precisely.",
], 8)

# ============================================================ 8. NOTES & QA
ws = sheet(wb, "Notes & QA", "GGMI Bing July 2026 — Sources, Reconciliation, QA and Open Items",
    "Pulled 2026-08-04 for the July reporting cycle. Reporting only; no account changes were made.", [120])
r = 4
def block(r, title, lines):
    c = ws.cell(r, 1, title); c.font = Font(name="Calibri", size=11, bold=True, color=NAVY); r += 1
    return notes(ws, r, lines) + 1

r = block(r, "SOURCES", [
    "bing-ads MCP, account 31003116 (FOREX.com LATAM, USD): campaign performance Jul 1-31, keyword performance Jul 1-31, conversion goals (live config).",
    "sa360 MCP, customer 5372690580 / login 9697709980 (Bing engine account 31003116): campaign performance Jul 1-31, ad_group performance, "
    "campaign x segments.date, campaign x segments.conversion_action_name, user_location_view (targeting_location = false).",
    "June baseline: reports/forex/ggmi/2026-06/data/GGMI-Bing-SA360-June-2026-data.xlsx, re-verified live on 2026-08-04.",
    "Prior context: reports/forex/ggmi/2026-07/qa/BING-quick-audit-2026-07-13.md; recommendations/forex/ggmi/GGMI-Bing-SA360-remediation-June-2026.md.",
    "Config audit (2026-08-04): sa360 campaign.geo_target_type_setting, campaign_criterion LOCATION, geo_target_constant lookup; bing-ads ad groups "
    "(campaigns 627650449, 627650452 only), budget utilization, search term report Jul 25 - Aug 4. bing_ads_list_ads is broken so RSA depth is unverified.",
    "August 1-4 2026 pulled as a forward pacing check: Bing-direct $1,500.12 vs SA360 $1,458.99 (delta $41.13, Aug 4 intraday sync lag), 6 submitted applications.",
])
r = block(r, "RECONCILIATION — ALL PASS", [
    "Bing-direct campaign spend total $10,624.95 vs SA360 $10,624.94. Delta $0.01, rounding on BrandGeneric. PASS.",
    "Keyword spend total $10,624.94 = campaign spend total $10,624.94. PASS.",
    "Geo spend total (Mexico $10,339.93 + non-Mexico $285.01) = $10,624.94. PASS.",
    "Daily spend total across 31 days = $10,624.94; daily clicks 4,090 = SA360 clicks 4,090. PASS.",
    "Ad-group conversions (17 + 1 + 2) = campaign conversions (18 + 2) = 20. PASS.",
    "Conversion-action detail: G2 live conf 16 + 2 + MT5 live conf 2 = 20 = metrics.conversions. PASS.",
])
r = block(r, "METRIC DEFINITIONS AND DEPARTURES FROM SOURCE", [
    "Conversions = SA360 metrics.conversions, which for this account counts ONLY the two 'Live Confirmation' actions (40059184 G2, 40059257 MT5). "
    "Bing-direct reports 0 conversions on every campaign because these are offline-imported; that 0 is a reporting artifact, not a real zero. "
    "Do not use Bing-native conversion figures for GGMI in any deliverable.",
    "PRIMARY KPI = SUBMITTED APPLICATIONS. The goals named 'Live Confirmation' fire on the post-submission confirmation / ThankYou page, so a "
    "Live Confirmation IS a submitted application. Confirmed from the account's own URL-based goals: 'FX ES Step 1-4' map to /en/step/1 through "
    "/en/step/4 (the form pages), and the confirmation goals map to /en/step/ThankYou (submission complete). This matches the June workbook's "
    "'submitted applications' wording — carry it forward unchanged, and keep the agency scorecard on cost per submitted application.",
    "DOWNSTREAM STEPS ARE SEPARATE GOALS and are deliberately excluded from the primary KPI: 'GCLID - Approved', 'GCLID - Funded', 'GCLID - Traded'. "
    "Those belong to the client's journey, not the agency scorecard. Note one live data point in July — MX_GEN_MT5 recorded 1 GCLID-Approved with 0 "
    "submitted applications in the same month, which is expected when the application was submitted in a prior period. Do not read Approved as a "
    "subset of July submissions.",
    "all_conversions in the Funnel tab excludes duplicate conversion actions that carry a timestamp suffix (for example 'GGMI - ... - Sitewide' and "
    "'GGMI - ... - Sitewide-1777462778694539000' both fire with different counts). Counting both would inflate every funnel step. metrics.conversions "
    "is unaffected — it only counts the non-suffixed live-confirmation actions.",
    "CTR, CPC, CPA calculated per the shared KPI set. No platform metric was silently remapped.",
])
r = block(r, "QA FLAGS AND OPEN ITEMS", [
    "1. GEO SETTING VERIFIED AND STILL WRONG. All 9 enabled campaigns carry positiveGeoTargetType = PRESENCE_OR_INTEREST as of 2026-08-04. "
    "July's 100% Mexico delivery is a low-volume artifact plus the pausing of the leaking legacy campaigns, not a fix. The breach is dormant and "
    "returns as volume scales. Venezuela — the largest leak in both June and July — is still not excluded. Highest-priority item on the account. "
    "Do NOT state to the client that the geo problem is solved.",
    "2. CONVERSION MATURITY. The newest July clicks are 4 days old at pull time and the live-confirmation window is 90 days. Phase B conversions "
    "will likely rise. Re-pull before the deck is finalised and restate if the count moves. June's figure was re-verified as stable, so the "
    "June-to-July comparison is sound in direction.",
    "3. LEGACY ZERO-CONVERSION ANOMALY. policytest_v2 spent $4,039.90 in July at 100% Mexico with a 7.19% CTR and produced 104 application starts "
    "and zero submitted applications, after converting 27 times in June. The funnel data argues this is a genuine drop-off rather than lag, but the cause is "
    "not established. Worth one focused look before it is characterised in client-facing material.",
    "4. CLIENT TRACKER BASIS NOT YET RECONCILED. Client-facing spend must match the client's own tracker. The July GGMI tracker has not been "
    "received or compared. $10,624.94 is the platform figure. Reconcile before any client-facing number is published.",
    "5. BIDDING SIGNAL STILL OFF. All 7 GGMI goals remain ExcludeFromBidding = TRUE for the third consecutive month. Execution sits with the "
    "paid-media/PPC agent under explicit approval; this repository does not change the account.",
    "6. 155 paused legacy campaigns and ~40 dead conversion goals remain in the account. Housekeeping, no performance impact.",
])
r = block(r, "ANALYST READ", [
    "July is a rebuild month and should be reported as one. The account delivered on 21 of 31 days and changed structure mid-month, so any "
    "single-month like-for-like comparison against June is misleading. Read it as two phases.",
    "The good news is real and defensible: the Mexico-only mandate is being met at 97.3% of spend, both primary conversion goals are importing "
    "again, and the rebuilt structure produced 20 submitted applications in 10 days at a cost per submitted app well below June's $513 — $237 account-wide, "
    "$141 on the two campaigns that converted.",
    "The cost is equally real: $5,882.57 went to the legacy set in the first half of the month for zero primary outcomes, and the account still "
    "bids without a conversion signal.",
    "The single highest-value finding for the client conversation is the funnel: 207 application starts on the legacy campaigns died between "
    "Step 2 and Step 3, while the new TradingView campaign carries 13.4% of its starts through to a submitted application. That is a fixable, "
    "attributable gap and it matches the GCG Q2 post-mortem conclusion.",
])
ws.column_dimensions["A"].width = 118

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("wrote", OUT)
for s in wb.sheetnames: print("  -", s)
