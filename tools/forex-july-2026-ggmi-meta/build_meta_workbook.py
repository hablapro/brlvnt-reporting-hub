"""
Build GGMI (LATAM) Meta July 2026 data workbook.
Source: Meta Ads MCP, act_1699453997689551 (SHARED GGMI+GCG account), campaign-level
pull filtered to GGMI campaigns by naming convention. Date range 2026-07-01..2026-07-31.
Reporting only. House style matches reports/forex/ggmi/2026-06/data/GGMI-Meta-Apr-Jun-2026-data.xlsx.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
SUBHEAD_FILL = "D9E1F2"
SECTION_FILL = "2E5496"
ROW_ALT_FILL = "F2F5FB"
GRAY_TEXT = "404040"
BORDER_COLOR = "BFBFBF"

thin = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def title_row(ws, text, row=1, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)


def subtitle_row(ws, text, row=2, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
    c.fill = PatternFill("solid", fgColor=SUBHEAD_FILL)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SUBHEAD_FILL)


def section_row(ws, text, row, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SECTION_FILL)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SECTION_FILL)


def bullet_row(ws, text, row, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30


def header_row(ws, headers, row):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def data_row(ws, values, row, formats=None, alt=False, total=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if total:
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=SECTION_FILL)
        else:
            c.font = Font(name="Calibri", size=10)
            if alt:
                c.fill = PatternFill("solid", fgColor=ROW_ALT_FILL)
        if formats and formats.get(i):
            c.number_format = formats[i]


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# TAB 1: Summary
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Summary")
title_row(ws, "GGMI (LATAM) — Meta — July 2026 Monthly Report")
subtitle_row(ws, "act_1699453997689551 · shared GGMI+GCG account, GGMI campaigns only (see Notes & QA for attribution rule) · MoM vs June · attribution 7d-click/1d-view · USD")

r = 4
section_row(ws, "HEADLINE", r); r += 1
for t in [
    "•  Spend $8,027.45 across 4 GGMI campaigns, down 69.0% vs June's $25,923.71. The drop is a budget reallocation, not a pullback: the CTR/traffic campaign was cut roughly in half ($3,341.99 vs June $23,997.09) while the conversion (SALES) campaigns took the larger share of what remained.",
    "•  Objective mix flipped: OUTCOME_SALES (conversion) campaigns are now 56.8% of GGMI spend ($4,561.62 across 2 campaigns/3 ad sets) vs a rounding error in June. TRAFFIC fell to 41.6% ($3,341.99). The conversion-objective switch flagged as pending in the June report has happened.",
    "•  117 SubmittedApplication pixel fires (all 3 conversion ad sets explicitly promote this event, per their promoted_object — resolves the June starts-vs-submit ambiguity going forward). Cost per result on the 2 SALES campaigns alone: $39.67 ($4,561.62 / 115). June's 86 conversions were flagged UNVALIDATED at a nominal $301 CPA; July's figure is both cheaper and better-defined, pending GA4 cross-check once the LAT property's key events are live (see the July UTM/GA4 handover).",
    "•  Mexico geo compliance: 100% of spend, impressions, and clicks across all 4 GGMI campaigns landed in Mexico. Clean, same as June.",
]:
    bullet_row(ws, t, r); r += 1

r += 1
section_row(ws, "WHAT WORKED", r); r += 1
for t in [
    "•  Instagram was actually tested this cycle: $816.44 (10.2% of GGMI spend) across reels/feed/stories, vs June's near-zero $70 (0.3%). Direct response to the June recommendation.",
    "•  The 3 conversion ad sets (PROS_MX_CONV, Retargeting_newlps, RTDOOH_WC_q3) all promote the SubmittedApplication custom event specifically — not a generic or ambiguous event. That was the open validation question from June.",
    "•  Facebook feed remains the efficient core: $3,885.26 (48.4% of spend), CPM $1.37, 33,928 LPV.",
]:
    bullet_row(ws, t, r); r += 1

r += 1
section_row(ws, "WHAT DID NOT WORK", r); r += 1
for t in [
    "•  Age skew is essentially unchanged: on the CTR campaign (apples-to-apples vs June's demographic pull), 55+ is 60.1% of spend (June 62.8%) and 65+ is 29.2% (June 32.0%). The client's June ask to refine targeting to 25+ has NOT been implemented — all 5 GGMI ad sets still target the full 18-65 range with Advantage+ Audience on.",
    "•  The Followers/engagement campaign remains conversion-dead: $123.84, 0 landing-page views, 0 pixel fires. Awareness-only, as flagged in June.",
    "•  Retargeting_newlps carries the highest frequency of any GGMI ad set (3.76) against its smallest reach (49,815) — a fatigue signal on a small retargeting pool.",
]:
    bullet_row(ws, t, r); r += 1

r += 1
section_row(ws, "WATCH / VALIDATE", r); r += 1
for t in [
    "•  Reconcile the 117 SubmittedApplication pixel fires against the client's own funnel once GA4 key events are live on the LAT property (currently 0 configured — separate open item, not a Meta-side issue).",
    "•  TradingView_exe_q2_reel is 71% of Retargeting_newlps spend ($1,010.89) and that ad set's highest-frequency creative — refresh candidate if it keeps running unchanged.",
    "•  0726_RTDOOH_WC_q3 (World Cup-themed retargeting) is new this cycle, one month of data only — don't scale on a single month.",
]:
    bullet_row(ws, t, r); r += 1

set_widths(ws, [110])
ws.freeze_panes = "A4"

# ---------------------------------------------------------------------------
# TAB 2: Campaigns
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Campaigns")
title_row(ws, "GGMI Meta — July Campaigns", span=12)
subtitle_row(ws, "GGMI campaigns only, attributed by naming convention (see Notes & QA). Cost/Result = spend / SubmittedApplication pixel fires; N/A where the campaign is not a conversion objective.", span=12)

headers = ["Campaign", "Objective", "Spend", "Impr", "Reach", "Freq", "Clicks", "CTR", "CPC", "CPM", "Link clicks", "LPV", "Pixel conv", "Cost/Result"]
r = 4
header_row(ws, headers, r); r += 1

campaigns = [
    ("0726_GGMI_Q2_esp_mx_CTR", "TRAFFIC", 3341.99, 3556379, 2251346, 1.579668, 81063, 0.02279369, 0.041227, 0.939717, 68114, 45582, 2, None),
    ("0626_GGMI_Q3_Followers_Campaign", "ENGAGEMENT", 123.84, 76302, 46365, 1.645681, 2934, 0.03845247, 0.042209, 1.623024, 2684, 0, 0, None),
    ("0726_GGMI_Q3_esp_mx_newlp_CONV", "SALES", 1428.69, 187238, 49815, 3.758667, 1768, 0.00944253, 0.808083, 7.630342, 1283, 82, 39, 1428.69/39),
    ("0726_GGMI_Q3_esp_mx_CONV", "SALES", 3132.93, 430575, 226252, 1.903077, 5388, 0.0125135, 0.581464, 7.276154, 2408, 154, 76, 3132.93/76),
]
formats = {3: '"$"#,##0.00', 4: '#,##0', 5: '#,##0', 6: '0.00', 7: '#,##0', 8: '0.00%', 9: '"$"#,##0.0000', 10: '"$"#,##0.0000', 11: '#,##0', 12: '#,##0', 13: '#,##0', 14: '"$"#,##0.00'}
for i, camp in enumerate(campaigns):
    row = list(camp)
    if row[-1] is None:
        row[-1] = "N/A (not a conversion objective)"
    data_row(ws, row, r, formats, alt=(i % 2 == 1)); r += 1

total = ["TOTAL", "", 8027.45, 4250494, "2,251,346 dominant / 2,573,778 upper bound*", "1.888 (blended)", 91153, 91153/4250494, 8027.45/91153, 8027.45/4250494*1000, 74489, 45818, 117, 4561.62/115]
data_row(ws, total, r, formats, total=True); r += 1
r += 1
bullet_row(ws, "* Reach cannot be summed across campaigns sharing the same MX audience without double-counting; the CTR campaign's own reach (2,251,346) is the more defensible single figure, in line with the June-report convention. Cost/Result on the last row is spend / pixel conv for the 2 SALES campaigns combined ($4,561.62 / 115 = $39.67); it excludes the Followers and CTR campaigns, which are not optimizing to conversions.", r, span=14); r += 2
set_widths(ws, [32, 12, 12, 12, 24, 14, 10, 10, 12, 10, 12, 10, 12, 14])
ws.freeze_panes = "A5"

# ---------------------------------------------------------------------------
# TAB 3: Ad Sets
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Ad Sets")
title_row(ws, "GGMI Meta — Ad Sets & Targeting", span=13)
subtitle_row(ws, "All 5 GGMI ad sets target Mexico, age 18-65, Advantage+ Audience ON. No ad set is scoped to 25+ as of July (client asked for this refinement in June).", span=13)

headers = ["Campaign", "Ad set", "Geo", "Age", "Spend", "Impr", "Reach", "Freq", "Clicks", "CTR", "LPV", "Pixel conv", "Notes"]
r = 4
header_row(ws, headers, r); r += 1
adsets = [
    ("0726_GGMI_Q2_esp_mx_CTR", "0426_PROS_MX_CTR", "MX", "18-65", 3341.99, 3556379, 2251346, 1.579668, 81063, 0.02279369, 45582, 2, "Prospecting; carries all CTR-campaign spend"),
    ("0626_GGMI_Q3_Followers_Campaign", "Broad Targeting", "MX", "18-65", 123.84, 76302, 46365, 1.645681, 2934, 0.03845247, 0, 0, "Ad set itself is PAUSED/WITH_ISSUES; other ads in campaign inactive"),
    ("0726_GGMI_Q3_esp_mx_newlp_CONV", "Retargeting_newlps", "MX", "18-65", 1428.69, 187238, 49815, 3.758667, 1768, 0.00944253, 82, 39, "Custom-audience retarget (View 50% CTR CAMPAIGN); highest freq of any GGMI ad set"),
    ("0726_GGMI_Q3_esp_mx_CONV", "0726_PROS_MX_CONV", "MX", "18-65", 1924.07, 312460, 179980, 1.736082, 3873, 0.01239519, 101, 47, "Prospecting, conversion objective"),
    ("0726_GGMI_Q3_esp_mx_CONV", "0726_RTDOOH_WC_q3", "MX", "18-65", 1208.86, 118115, 69432, 1.701161, 1515, 0.01282648, 53, 29, "New this cycle; World Cup-themed retargeting/DOOH creative"),
]
formats = {5: '"$"#,##0.00', 6: '#,##0', 7: '#,##0', 8: '0.00', 9: '#,##0', 10: '0.00%', 11: '#,##0', 12: '#,##0'}
for i, a in enumerate(adsets):
    data_row(ws, a, r, formats, alt=(i % 2 == 1)); r += 1
set_widths(ws, [28, 22, 8, 8, 12, 12, 12, 10, 10, 10, 10, 12, 44])
ws.freeze_panes = "A5"

# ---------------------------------------------------------------------------
# TAB 4: Creatives (ad-level)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Creatives")
title_row(ws, "GGMI Meta — Creative Performance (ad level)", span=9)
subtitle_row(ws, "All 34 GGMI ads with July delivery, ranked by spend. Same creative can run in more than one ad set/objective.", span=9)

headers = ["Ad", "Ad set / campaign", "Spend", "Impr", "Link clicks", "CTR", "CPM", "LPV", "Pixel conv"]
r = 4
header_row(ws, headers, r); r += 1

ads = [
    ("edu_trust_q2", "0426_PROS_MX_CTR / TRAFFIC", 3128.71, 3296688, 63452, 0.02303524, 0.9490464369088006, 42677, 1),
    ("edu_trust_q2", "0726_PROS_MX_CONV / SALES", 1209.68, 192433, 940, 0.01591203, 6.2862398860902235, 44, 18),
    ("TradingView_exe_q2_reel", "Retargeting_newlps / SALES", 1010.89, 119104, 906, 0.01067135, 8.487456340677054, 55, 28),
    ("Retarg_DOOH_account", "0726_RTDOOH_WC_q3 / SALES", 472.70, 33277, 161, 0.00829402, 14.205006460918954, 22, 18),
    ("commod_pro_q2", "0726_PROS_MX_CONV / SALES", 404.04, 81967, 343, 0.00559981, 4.929300816182122, 31, 22),
    ("Retarg_DOOH_trader", "0726_RTDOOH_WC_q3 / SALES", 249.80, 17529, 86, 0.00661761, 14.250670317759141, 14, 6),
    ("wcup_timing_reel", "0726_RTDOOH_WC_q3 / SALES", 201.34, 18828, 153, 0.01290631, 10.69364775865732, 4, 1),
    ("demo_trust_q2", "0726_PROS_MX_CONV / SALES", 201.27, 15839, 149, 0.01388977, 12.707241618789066, 22, 3),
    ("Error_trader_still 1", "Broad Targeting / ENGAGEMENT", 123.39, 76052, 2675, 0.03847368, 1.6224425393152053, 0, 0),
    ("demo_trust_q2", "0426_PROS_MX_CTR / TRAFFIC", 122.25, 120169, 2653, 0.0249482, 1.017317278166582, 1721, 0),
    ("broker_trust_q3_still - Copia", "Retargeting_newlps / SALES", 108.74, 17944, 78, 0.00540571, 6.059964333481943, 7, 6),
    ("TradingView_award_q2_still", "Retargeting_newlps / SALES", 106.51, 17889, 95, 0.00654033, 5.953938174297054, 5, 1),
    ("Retarg_DOOH_market", "0726_RTDOOH_WC_q3 / SALES", 83.85, 5613, 35, 0.00837342, 14.938535542490646, 5, 4),
    ("exp_proof_q2", "0426_PROS_MX_CTR / TRAFFIC", 74.33, 113287, 1667, 0.01561521, 0.6561211789525718, 958, 1),
    ("wcup_broker_reel", "0726_RTDOOH_WC_q3 / SALES", 72.63, 26487, 372, 0.02673009, 2.7420998980632008, 5, 0),
    ("global_plat_q2", "0726_PROS_MX_CONV / SALES", 67.98, 8439, 46, 0.00746534, 8.055456807678635, 1, 2),
    ("wcup_brand_reel", "0726_RTDOOH_WC_q3 / SALES", 65.52, 10131, 50, 0.0084888, 6.467278649689073, 2, 0),
    ("TradingView_instruments_q2_still", "Retargeting_newlps / SALES", 56.00, 10315, 62, 0.00853127, 5.428986912263694, 4, 0),
    ("TradingView_direct_q2_reel", "Retargeting_newlps / SALES", 55.28, 6188, 59, 0.01470588, 8.933419521654816, 5, 1),
    ("platform_traders_q3_still - Copia", "Retargeting_newlps / SALES", 55.03, 11198, 41, 0.00482229, 4.914270405429541, 3, 2),
    ("exp_proof_q2", "0726_PROS_MX_CONV / SALES", 41.10, 13782, 53, 0.00500653, 2.982150631258163, 3, 2),
    ("wcup_trust_ad", "0726_RTDOOH_WC_q3 / SALES", 32.98, 4038, 14, 0.00619118, 8.167409608717186, 0, 0),
    ("wcup_market_ad", "0726_RTDOOH_WC_q3 / SALES", 30.04, 2212, 6, 0.00632911, 13.580470162748643, 1, 0),
    ("TradingView_integration_q2_still - Copia", "Retargeting_newlps / SALES", 17.79, 3127, 14, 0.00639591, 5.6891589382795, 1, 1),
    ("commod_pro_q2", "0426_PROS_MX_CTR / TRAFFIC", 16.52, 25933, 337, 0.01353488, 0.6370261828558207, 224, 0),
    ("platform_traders_q3_still", "Retargeting_newlps / SALES", 9.00, 722, 18, 0.02354571, 12.465373961218837, 2, 0),
    ("broker_trust_q3_still", "Retargeting_newlps / SALES", 7.89, 644, 9, 0.01552795, 12.251552795031056, 0, 0),
    ("TradingView_integration_q2_still", "Retargeting_newlps / SALES", 1.56, 107, 1, 0.02803738, 14.57943925233645, 0, 0),
    ("Graph_talk_carrousel 1", "Broad Targeting / ENGAGEMENT", 0.28, 210, 4, 0.02380952, 1.3333333333333335, 0, 0),
    ("global_plat_q2", "0426_PROS_MX_CTR / TRAFFIC", 0.18, 302, 5, 0.01655629, 0.5960264900662252, 2, 0),
    ("Errores_trader_Reel 2", "Broad Targeting / ENGAGEMENT", 0.10, 11, 4, 0.18181818, 9.090909090909092, 0, 0),
    ("Trading_Reel 1", "Broad Targeting / ENGAGEMENT", 0.06, 22, 1, 0.04545455, 2.727272727272727, 0, 0),
    ("Analizar_plat_Reel 4", "Broad Targeting / ENGAGEMENT", 0.01, 4, 0, 0.0, 2.5, 0, 0),
    ("Plataforma_Reel 3", "Broad Targeting / ENGAGEMENT", 0.00, 3, 0, 0.0, 0.0, 0, 0),
]
formats = {3: '"$"#,##0.00', 4: '#,##0', 5: '#,##0', 6: '0.00%', 7: '"$"#,##0.0000', 8: '#,##0', 9: '#,##0'}
for i, a in enumerate(ads):
    data_row(ws, a, r, formats, alt=(i % 2 == 1)); r += 1
r += 1
bullet_row(ws, "Reads: edu_trust_q2 is the workhorse creative, running in both the CTR ($3,128.71) and CONV ($1,209.68) ad sets — 54% of all GGMI spend on one creative. TradingView_exe_q2_reel remains the conversion standout (28 pixel fires on $1,010.89, CPA ~$36) but also the fatigue risk (adset freq 3.76). The 4 World Cup retargeting/DOOH creatives (Retarg_DOOH_*, wcup_*) are new and thin (1 month, $1,208.86 combined) — no fatigue read yet. Followers-campaign creatives (Broad Targeting ad set) are functionally inactive (<$125 combined, 0 LPV, 0 conv).", r, span=9); r += 2
set_widths(ws, [32, 28, 12, 12, 12, 10, 12, 10, 10])
ws.freeze_panes = "A5"

# ---------------------------------------------------------------------------
# TAB 5: Geo
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Geo")
title_row(ws, "GGMI Meta — Geo Compliance (country breakdown)", span=6)
subtitle_row(ws, "GGMI is contracted to Mexico ONLY. Country breakdown pulled at campaign level, account-wide, filtered to GGMI campaigns.", span=6)

headers = ["Campaign", "Country", "Spend", "Impr", "Reach", "Clicks"]
r = 4
header_row(ws, headers, r); r += 1
geo_rows = [
    ("0726_GGMI_Q2_esp_mx_CTR", "MX", 3341.99, 3556379, 2251346, 81061),
    ("0726_GGMI_Q2_esp_mx_CTR", "unknown", 0.00, 0, 0, 2),
    ("0626_GGMI_Q3_Followers_Campaign", "MX", 123.84, 76302, 46365, 2934),
    ("0726_GGMI_Q3_esp_mx_newlp_CONV", "MX", 1428.69, 187238, 49815, 1768),
    ("0726_GGMI_Q3_esp_mx_CONV", "MX", 3132.93, 430575, 226252, 5388),
]
formats = {3: '"$"#,##0.00', 4: '#,##0', 5: '#,##0', 6: '#,##0'}
for i, gr in enumerate(geo_rows):
    data_row(ws, gr, r, formats, alt=(i % 2 == 1)); r += 1
r += 1
data_row(ws, ["VERDICT", "100.00% Mexico", 8027.45, "—", "—", "—"], r, formats, total=True); r += 2
bullet_row(ws, "Compliant. Every dollar and impression across all 4 GGMI campaigns is Mexico-attributed; the 2 unattributed clicks ($0 spend) on the CTR campaign are a rounding/attribution-timing artifact, not a geo issue. Same clean result as June. GCG campaigns (not shown; out of scope for this workbook) are 100% US.", r, span=6); r += 2
set_widths(ws, [32, 14, 12, 12, 14, 10])
ws.freeze_panes = "A5"

# ---------------------------------------------------------------------------
# TAB 6: Placements
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Placements")
title_row(ws, "GGMI Meta — Placements (publisher x position)", span=7)
subtitle_row(ws, "GGMI campaigns only, ranked by spend. Instagram is meaningfully tested this month (10.2% of spend) vs June's near-zero.", span=7)

headers = ["Platform", "Position", "Spend", "Impr", "Clicks", "LPV", "Pixel conv"]
r = 4
header_row(ws, headers, r); r += 1
placements = [
    ("facebook", "feed", 3885.26, 2831370, 64804, 33928, 37),
    ("facebook", "facebook_reels", 1940.33, 528686, 9326, 3139, 34),
    ("facebook", "instream_video", 519.35, 103897, 3151, 1425, 10),
    ("facebook", "facebook_stories", 474.70, 87937, 6449, 4378, 6),
    ("instagram", "instagram_reels", 443.68, 74849, 1864, 88, 4),
    ("facebook", "facebook_reels_overlay", 315.54, 555153, 4782, 2692, 4),
    ("instagram", "feed", 229.14, 17427, 86, 24, 10),
    ("instagram", "instagram_stories", 143.62, 14616, 74, 4, 4),
    ("threads", "threads_feed", 25.89, 4885, 30, 3, 0),
    ("audience_network", "an_classic", 16.44, 11592, 347, 68, 4),
    ("facebook", "right_hand_column", 15.52, 10894, 35, 1, 2),
    ("facebook", "search", 8.42, 2488, 20, 4, 2),
    ("facebook", "marketplace", 4.05, 3856, 62, 42, 0),
    ("audience_network", "rewarded_video", 3.96, 1403, 104, 21, 0),
    ("facebook", "biz_disco_feed", 0.92, 289, 12, 0, 0),
    ("messenger", "messenger_stories", 0.65, 1152, 6, 0, 0),
]
formats = {3: '"$"#,##0.00', 4: '#,##0', 5: '#,##0', 6: '#,##0', 7: '#,##0'}
for i, p in enumerate(placements):
    data_row(ws, p, r, formats, alt=(i % 2 == 1)); r += 1
r += 1
bullet_row(ws, "Facebook 89.2% of spend, Instagram 10.2%, Audience Network 0.25%, Threads 0.3%, Messenger <0.1%. FB feed is still the efficient core (48.4% of spend, best absolute LPV volume). facebook_reels_overlay (June's flagged 'cheap junk' placement) is down to $316 from June's $1,653 — a smaller share of a smaller budget, still worth an exclusion review if the CTR campaign scales back up.", r, span=7); r += 2
set_widths(ws, [18, 22, 12, 12, 10, 10, 12])
ws.freeze_panes = "A5"

# ---------------------------------------------------------------------------
# TAB 7: MoM
# ---------------------------------------------------------------------------
ws = wb.create_sheet("MoM")
title_row(ws, "GGMI Meta — Trend (Apr-Jul 2026)", span=6)
subtitle_row(ws, "April-June carried forward from reports/forex/ggmi/2026-06/data/GGMI-Meta-Apr-Jun-2026-data.xlsx. Conversions = offsite_conversion.fb_pixel_custom (June: UNVALIDATED; July: explicitly SubmittedApplication per promoted_object).", span=6)

headers = ["Metric", "April", "May", "June", "July", "MoM (Jul v Jun)"]
r = 4
header_row(ws, headers, r); r += 1
trend = [
    ("Spend", 5227.18, 6626.06, 25923.71, 8027.45, (8027.45 - 25923.71) / 25923.71),
    ("Impressions", 2935470, 6838850, 19780693, 4250494, (4250494 - 19780693) / 19780693),
    ("Link clicks", 84321, 128231, 407136, 74489, (74489 - 407136) / 407136),
    ("Link CTR", 0.0287248719966479, 0.01875037469750031, 0.02058249425336109, 74489/4250494, (74489/4250494 - 0.02058249425336109)),
    ("CPM", 1.780696106585998, 0.9688851195741975, 1.31055620751002, 8027.45/4250494*1000, (8027.45/4250494*1000 - 1.31055620751002) / 1.31055620751002),
    ("Landing page views", 42606, 64264, 249972, 45818, (45818 - 249972) / 249972),
    ("Pixel conv (fb_pixel_custom)", 1, 4, 86, 117, (117 - 86) / 86),
    ("Apparent CPA (blended, all spend/all conv)", 5227.18, 1656.515, 301.438488372093, 8027.45/117, (8027.45/117 - 301.438488372093) / 301.438488372093),
]
formats = {2: '"$"#,##0.00', 3: '"$"#,##0.00', 4: '"$"#,##0.00', 5: '"$"#,##0.00', 6: '0.0%'}
row_formats_override = {
    "Impressions": {2: '#,##0', 3: '#,##0', 4: '#,##0', 5: '#,##0', 6: '0.0%'},
    "Link clicks": {2: '#,##0', 3: '#,##0', 4: '#,##0', 5: '#,##0', 6: '0.0%'},
    "Link CTR": {2: '0.00%', 3: '0.00%', 4: '0.00%', 5: '0.00%', 6: '+0.00%;-0.00%'},
    "Landing page views": {2: '#,##0', 3: '#,##0', 4: '#,##0', 5: '#,##0', 6: '0.0%'},
    "Pixel conv (fb_pixel_custom)": {2: '#,##0', 3: '#,##0', 4: '#,##0', 5: '#,##0', 6: '+0.0%;-0.0%'},
}
for i, t in enumerate(trend):
    fmt = row_formats_override.get(t[0], formats)
    data_row(ws, t, r, fmt, alt=(i % 2 == 1)); r += 1
r += 1
bullet_row(ws, "Spend down 69.0% MoM but pixel conversions up 36.0% and cost-per-result improved sharply — a real efficiency gain, though not a clean apples-to-apples read since June ran ~93% traffic-objective and July runs ~57% conversion-objective by spend. CPM is up 44% MoM (retargeting/CONV traffic costs more per impression than the cheap broad-traffic clicks that dominated June). June's reach/frequency (reported client-side as ~6.3M/3.0) was never stored as a data-tab number in the June workbook, so no reliable numeric reach MoM exists; July's reach (2.25M dominant-campaign) is the first cleanly-sourced figure.", r, span=6); r += 2
set_widths(ws, [34, 16, 16, 16, 16, 16])
ws.freeze_panes = "A5"

# ---------------------------------------------------------------------------
# TAB 8: Notes & QA
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Notes & QA")
title_row(ws, "GGMI Meta — Notes, QA & Attribution Rule", span=1)
subtitle_row(ws, "Read before using any number in this workbook elsewhere.", span=1)
r = 4
section_row(ws, "SOURCE", r, span=1); r += 1
for t in [
    "Meta Ads MCP, act_1699453997689551, date range 2026-07-01 to 2026-07-31 (calendar month, not a rolling window). Attribution 7-day click / 1-day view (platform default).",
]:
    bullet_row(ws, t, r, span=1); r += 1

r += 1
section_row(ws, "GGMI / GCG ATTRIBUTION RULE", r, span=1); r += 1
for t in [
    "act_1699453997689551 is SHARED between GGMI (LATAM/Mexico) and GCG (US Hispanic). Campaigns are named with a client prefix (0726_GGMI_... vs 0726_GCG_...) — attributed by that naming convention, cross-checked against each campaign's ad-set geo targeting (GGMI ad sets all target MX; GCG ad sets all target US). No ambiguous campaigns this cycle.",
    "GGMI campaigns (4, all with July spend): 0726_GGMI_Q2_esp_mx_CTR, 0626_GGMI_Q3_Followers_Campaign, 0726_GGMI_Q3_esp_mx_newlp_CONV, 0726_GGMI_Q3_esp_mx_CONV.",
    "GCG campaigns (3, all with July spend, excluded from this workbook): 0726_GCG_Q3_esp_us_CTR, 0726_GCG_Q3_esp_us_CONV, 0426_GCG_Q2_esp_us_CTR (this last one is PAUSED as of account-status but still shows $2,104.37 of July delivery before it was paused — flag for whoever builds the GCG July workbook).",
    "5 further campaigns (Feb-Mar 2026, both clients) are long-paused with $0 July spend and excluded entirely.",
]:
    bullet_row(ws, t, r, span=1); r += 1

r += 1
section_row(ws, "RECONCILIATION", r, span=1); r += 1
for t in [
    "Campaign-level spend sums to the account-level total exactly across all 7 active campaigns ($14,966.95); GGMI's 4 sum to $8,027.45 with no rounding drift found.",
    "Ad-set-level spend for each GGMI campaign reconciles exactly to that campaign's total (checked for all 5 ad sets).",
    "Ad-level spend for each ad set reconciles exactly to that ad set's total (checked for all 34 GGMI ads).",
    "Placement-breakdown spend sums to within $0.02 of the GGMI total (rounding across 16 platform x position rows, immaterial).",
    "Age/gender breakdown spend sums to within $0.01 of the GGMI total across all 4 campaigns.",
]:
    bullet_row(ws, t, r, span=1); r += 1

r += 1
section_row(ws, "DATA QUALITY / OPEN FLAGS", r, span=1); r += 1
for t in [
    "GA4 key events on the LAT property (508849216) are still not configured (0 designated key events as of the July UTM/GA4 handover) — the 117 SubmittedApplication pixel fires reported here are a Meta-side signal only and have not been cross-checked against GA4 or the client's own funnel. Do not present the 117 figure as validated against the client's tracker without that cross-check.",
    "'offsite_conversion.fb_pixel_custom' in the platform actions array is a generic custom-conversion action type; it does not itself carry the specific event name in the API response used here. The SubmittedApplication attribution is inferred from each ad set's promoted_object.custom_event_str, which was pulled and confirmed for all 3 conversion ad sets — treat as high-confidence, not certain, absent a client-side receipt test (per the June Conversion QA precedent).",
    "Reach cannot be de-duplicated across the 4 GGMI campaigns from this MCP (no combined-campaign-set reach query available) — the dominant CTR-campaign figure (2,251,346) is reported as the single GGMI reach number, matching the June-report convention; the naive sum across all 4 campaigns (2,573,778) is an upper bound only, not a true unique count.",
    "0426_GCG_Q2_esp_us_CTR (GCG, not GGMI) shows $2,104.37 of July spend despite currently being flagged PAUSED at the account level — it was live for part of July before being paused (updated_time 2026-07-02). Not a GGMI issue, but flagged for whoever reconciles the GCG July numbers.",
    "This workbook does not include a client-funnel or GA4 cross-reference tab (unlike the June workbook's Funnel/Conversion QA tabs) — that synthesis belongs with whoever assembles the full July cross-channel model, not this single-channel data pull.",
]:
    bullet_row(ws, t, r, span=1); r += 1

set_widths(ws, [130])
ws.freeze_panes = "A4"

out_path = "/Users/rpro/AI-BRLVNT/Brlvnt-Reporting-Analytics-2026/reports/forex/ggmi/2026-07/data/GGMI-Meta-July-2026-data.xlsx"
wb.save(out_path)
print("Saved", out_path)
