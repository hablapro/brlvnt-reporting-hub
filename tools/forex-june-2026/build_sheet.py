#!/usr/bin/env python3
"""Build the GGMI June 2026 Performance Report workbook (.xlsx).

Four tabs (Summary / Performance / Diagnostics / Data Notes) formatted to
the Google Sheets Styling Standard in dashboards/dashboard-spec.md, for
upload to Drive with conversion to Google Sheets. Mirrors the May report
(1XsQTZ...) structure. Numbers: model/GGMI-Jun-2026-cross-channel-model.xlsx
and the data/ workbooks. Narrative: approved v2 draft.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = '1A2A3A'
BLUE = '3B5998'
GRAY = 'F3F3F3'
BODY = '32373E'

F_TITLE = Font(name='Arial', size=16, bold=True, color='FFFFFF')
F_SECTION = Font(name='Arial', size=12, bold=True, color=NAVY)
F_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_LABEL = Font(name='Arial', size=10, bold=True, color=BODY)
F_BODY = Font(name='Arial', size=10, color=BODY)
FILL_NAVY = PatternFill('solid', fgColor=NAVY)
FILL_BLUE = PatternFill('solid', fgColor=BLUE)
FILL_GRAY = PatternFill('solid', fgColor=GRAY)
TOP_BORDER = Border(top=Side(style='thin', color='555555'))
WRAP = Alignment(wrap_text=True, vertical='top')
RIGHT = Alignment(horizontal='right')
CENTER_V = Alignment(vertical='center')

CUR = '$#,##0.00'
CUR0 = '$#,##0'
INT = '#,##0'
PCT3 = '0.000%'
PCT2 = '0.00%'

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, '..', '..', 'reports', 'forex', 'ggmi', '2026-06',
                   'output', 'GGMI-June-2026-Performance-Report.xlsx')

wb = Workbook()


def title_bar(ws, ncols, text):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = F_TITLE
    c.fill = FILL_NAVY
    c.alignment = CENTER_V
    ws.row_dimensions[1].height = 38


def section(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = F_SECTION
    return row + 1


def table(ws, row, headers, rows, fmts=None, total_last=False):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.font = F_HDR
        c.fill = FILL_BLUE
    row += 1
    for ri, r in enumerate(rows):
        is_total = total_last and ri == len(rows) - 1
        for ci, v in enumerate(r, 1):
            c = ws.cell(row, ci, v)
            c.font = F_LABEL if (ci == 1 or is_total) else F_BODY
            if ri % 2 == 0 or is_total:
                c.fill = FILL_GRAY
            if is_total:
                c.border = TOP_BORDER
            if ci > 1:
                c.alignment = RIGHT
                if fmts and fmts[ci - 1] and isinstance(v, (int, float)):
                    c.number_format = fmts[ci - 1]
        row += 1
    return row


def para(ws, row, ncols, text, bold=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = F_LABEL if bold else F_BODY
    c.alignment = WRAP
    lines = max(1, len(text) // 110 + 1)
    ws.row_dimensions[row].height = 13 * lines + 4
    return row + 1


def meta_rows(ws, row, pairs):
    for k, v in pairs:
        a = ws.cell(row, 1, k)
        a.font = F_LABEL
        a.fill = FILL_GRAY
        b = ws.cell(row, 2, v)
        b.font = F_BODY
        b.fill = FILL_GRAY
        for ci in range(3, 9):
            ws.cell(row, ci).fill = FILL_GRAY
        row += 1
    return row


# ============================ SUMMARY =======================================
ws = wb.active
ws.title = 'Summary'
NC = 8
title_bar(ws, NC, 'FOREX.com GGMI (LATAM) — June 2026 Performance Review')
r = 2
r = meta_rows(ws, r, [
    ('Reporting Period', 'June 1 – 30, 2026'),
    ('Comparison Basis', 'vs May 2026 (month over month)'),
    ('Currency', 'USD'),
    ('Timezone', 'America/New_York'),
    ('Channels', 'Bing Ads, Meta, Azerion (Programmatic), Quantcast (Programmatic)'),
    ('Attribution', 'Per channel; conversions not summed across channels (see Data Notes)'),
    ('Revenue / ROAS', 'Not tracked; excluded'),
    ('Data Freshness', 'Pulled July 7, 2026'),
    ('Prepared by', 'Berelvant'),
])
r += 1
r = section(ws, r, NC, 'Cross-Channel Performance Summary')
r = table(ws, r,
          ['Channel', 'Spend', 'Impressions', 'Clicks', 'CTR', 'CPM',
           'Conversions', 'CPA'],
          [
              ['Bing Ads', 25658.61, 466582, 21480, 0.046037, '—', 50, 513.17],
              ['Meta', 25923.71, 19780693, 407136, 0.020582, 1.31,
               'Held (validation)', '—'],
              ['Azerion', 34555.83, 7679074, 9910, 0.001291, 4.50, 42, 822.76],
              ['Quantcast', 33784.20, 41964872, 11284, 0.000269, 0.81,
               '11 (view-through)', '—'],
              ['Total', 119922.35, 69891221, 449810, '—', '—', '—', '—'],
          ],
          fmts=[None, CUR, INT, INT, PCT3, CUR, INT, CUR],
          total_last=True)
r = para(ws, r, NC,
         'Conversions are measured on different events per channel and are '
         'not summed. Bing recorded 50 submitted applications at a $513 CPA '
         '(SA360 Primary, the live-account confirmation goal), up from 33 in '
         'May. Azerion counts submitted applications (42) alongside 440 '
         'Step 1 starts. Meta platform conversions are held from reporting '
         'pending a placement audit and pixel-event review. Quantcast '
         'recorded 11 view-through conversions, a directional signal only.')
r += 1
r = section(ws, r, NC, 'Executive Summary')
r = para(ws, r, NC,
         'June was the biggest GGMI media month of 2026: $119,922 across '
         'four channels, up 56% from May, delivering 69.9M impressions and '
         '449.8K clicks. Application volume followed the spend. Search '
         'produced 50 submitted applications at $513 each (up from 33 in '
         'May) and Azerion added 42 at $823, holding efficiency through a '
         '27% budget increase. Site traffic recovered with the spend: 9,236 '
         'Mexico sessions (+61% vs May) and 5,838 unique visitors (+125%). '
         'The recovery is real but bought. Media drove essentially all of '
         'the June gain while organic search sat at its low, down 75% since '
         'January per Search Console. We are approximately 31% into the '
         'annual GGMI budget.')
r += 1
r = section(ws, r, NC, 'Campaign Highlights')
for b in [
    '•  Bing held efficiency through a 61% scale-up: 50 submitted '
    'applications at $513 (May: 33 at $484), now across 3 campaigns; '
    'Platform Intercept leads at a $285 CPA.',
    '•  Azerion applications rose 14% (37 to 42) on a 27% budget increase; '
    'application starts nearly doubled to 440; Global Market converted at '
    '$407.',
    '•  Meta scaled 291% at a $1.31 CPM, 100% Mexico, in its first month '
    'off a pure traffic objective; click quality is the June issue and its '
    'conversions stay held.',
    '•  Quantcast delivered a deliberate reach month: 42M impressions '
    '(+155%) at a $0.81 CPM (-50%), reaching 18.2M devices.',
]:
    r = para(ws, r, NC, b)
r += 1
r = section(ws, r, NC, 'Suggested Optimizations')
for b in [
    '•  Complete the Mexico presence-only restriction on search: 49% of '
    'June spend ($12,637) served out-of-market users; the June 3 exclusion '
    'is already working.',
    '•  Run the Meta placement-level audit and exclusions before '
    'reinstating its conversion reporting; rebalance the 55+ age skew (63% '
    'of June spend).',
    '•  Apply the 49-site Quantcast blocklist (32% of June spend on that '
    'line) and set a campaign-level viewability floor (June viewability '
    '51.3% vs the 70% standard).',
    '•  Concentrate Azerion budget on converting ad sets; '
    'Conversion_Commodities spent $4,072 with zero applications.',
    '•  Open a dedicated SEO workstream for the Spanish site: organic '
    'clicks are down 75% since January for identifiable content reasons.',
]:
    r = para(ws, r, NC, b)
r += 1
r = section(ws, r, NC, 'Next Steps')
for b in [
    '•  Complete the Mexico-only search restriction (top efficiency '
    'lever).',
    '•  Link the search (SA360) account in GA4: 27% of June Mexico '
    'sessions are unattributed; July reporting will restate paid search '
    'larger once linked, and we will flag the restatement.',
    '•  Complete the Meta placement audit and pixel-event review, then '
    'reinstate conversion reporting.',
    '•  Hold the budget mix pending the measurement fixes; revisit channel '
    'allocation with clean July attribution.',
]:
    r = para(ws, r, NC, b)
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 50
for col in 'CDEFGH':
    ws.column_dimensions[col].width = 14
ws.freeze_panes = 'A2'
ws.sheet_view.showGridLines = False

# ============================ PERFORMANCE ===================================
ws = wb.create_sheet('Performance')
NC = 10
title_bar(ws, NC, 'Channel Performance Detail — June 2026')
r = 3
r = section(ws, r, NC, 'Channel KPI Table (June 2026)')
r = table(ws, r,
          ['Channel', 'Spend', 'Impressions', 'Clicks', 'CTR', 'CPM', 'CPC',
           'Primary Conversions', 'CPA', 'Viewability'],
          [
              ['Bing Ads', 25658.61, 466582, 21480, 0.046037, '—', 1.19,
               '50 submitted apps', 513.17, '—'],
              ['Meta', 25923.71, 19780693, 407136, 0.020582, 1.31, 0.06,
               'Held (validation)', '—', '—'],
              ['Azerion', 34555.83, 7679074, 9910, 0.001291, 4.50, 3.49,
               '42 submitted apps', 822.76, 0.6847],
              ['Quantcast', 33784.20, 41964872, 11284, 0.000269, 0.81, 2.99,
               '11 view-through', '—', 0.5127],
              ['Total', 119922.35, 69891221, 449810, '—', '—', '—',
               'do not sum', '—', '—'],
          ],
          fmts=[None, CUR, INT, INT, PCT3, CUR, CUR, None, CUR, PCT2],
          total_last=True)
r += 1
r = section(ws, r, NC, 'Spend by Channel — April / May / June')
r = table(ws, r,
          ['Channel', 'April', 'May', 'June', 'MoM (Jun v May)',
           'June share'],
          [
              ['Bing Ads', 15289.21, 15972.00, 25658.61, 0.6065, 0.2140],
              ['Meta', 5227.18, 6626.06, 25923.71, 2.9124, 0.2162],
              ['Quantcast', 15004.71, 26890.00, 33784.20, 0.2564, 0.2817],
              ['Azerion', 10215.00, 27257.72, 34555.83, 0.2677, 0.2882],
              ['Total', 45736.10, 76745.78, 119922.35, 0.5626, 1.0],
          ],
          fmts=[None, CUR, CUR, CUR, PCT2, PCT2],
          total_last=True)
r = para(ws, r, NC,
         'May Quantcast uses the client-reconciled $26,890. Azerion is raw '
         'vendor spend in all months (see Data Notes).')
r += 1
r = section(ws, r, NC, 'Primary Conversions by Channel (per-channel '
                       'definitions; never summed)')
r = table(ws, r,
          ['Channel', 'April', 'May', 'June', 'June CPA', 'Definition'],
          [
              ['Bing (SA360)', 'n/a', 33, 50, 513.17,
               'Submitted applications'],
              ['Meta', 1, 4, 'Held', '—',
               'Pixel events; held pending validation'],
              ['Quantcast', 0, 1, 11, '—', 'View-through (directional)'],
              ['Azerion', 'n/a', 37, 42, 822.76,
               'Submitted applications (vendor)'],
          ],
          fmts=[None, INT, INT, INT, CUR, None])
r += 1
r = section(ws, r, NC, 'Bing Campaigns (June)')
r = table(ws, r,
          ['Campaign', 'Spend', 'Impressions', 'Clicks', 'CTR', 'Avg CPC',
           'Apps', 'CPA'],
          [
              ['AO Generic (policy test)', 13508.22, 164946, 11020,
               0.066810, 1.23, 27, 500.30],
              ['Brand + Generic', 8154.40, 195341, 7559, 0.038696, 1.08, 9,
               906.04],
              ['Platform Intercept', 3995.99, 106295, 2901, 0.027292, 1.38,
               14, 285.43],
              ['Total', 25658.61, 466582, 21480, 0.046037, 1.19, 50,
               513.17],
          ],
          fmts=[None, CUR, INT, INT, PCT3, CUR, INT, CUR],
          total_last=True)
r += 1
r = section(ws, r, NC, 'Azerion Ad Sets (June)')
r = table(ws, r,
          ['Ad Set', 'Spend', 'Impressions', 'Clicks', 'Apps', 'CPA',
           'Viewability'],
          [
              ['Experience', 8996.03, 1999118, 2499, 11, 817.82, 0.6958],
              ['Global Market', 4069.06, 904236, 1134, 10, 406.91, 0.6759],
              ['TradeForex', 4164.07, 925349, 1140, 7, 594.87, 0.6863],
              ['Language Broker', 4927.14, 1094921, 1510, 7, 703.88, 0.6745],
              ['Conversion_Instruments', 4154.38, 923196, 1144, 4, 1038.60,
               0.6831],
              ['Spanish Platform', 4172.71, 927269, 1309, 3, 1390.90,
               0.6723],
              ['Conversion_Commodities', 4072.43, 904985, 1174, 0, '—',
               0.6939],
              ['Total', 34555.83, 7679074, 9910, 42, 822.76, 0.6847],
          ],
          fmts=[None, CUR, INT, INT, INT, CUR, PCT2],
          total_last=True)
r += 1
r = section(ws, r, NC, 'Site Traffic — GA4 Mexico, monthly')
r = table(ws, r,
          ['Metric', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
          [
              ['Sessions', 9380, 22229, 12614, 7577, 5751, 9236],
              ['Unique visitors', 4651, 14172, 6554, 3366, 2592, 5838],
          ],
          fmts=[None, INT, INT, INT, INT, INT, INT])
r = para(ws, r, NC,
         'June sessions +61% vs May; unique visitors +125% vs May and +26% '
         'vs January. February reflects a single paid flight and is not a '
         'valid baseline (see Diagnostics).')
ws.column_dimensions['A'].width = 28
for col in 'BCDEFGHIJ':
    ws.column_dimensions[col].width = 13
ws.column_dimensions['H'].width = 20
ws.freeze_panes = 'A2'

# ============================ DIAGNOSTICS ===================================
ws = wb.create_sheet('Diagnostics')
NC = 7
title_bar(ws, NC, 'Diagnostics — Measurement, Delivery Quality, Organic')
r = 3
r = section(ws, r, NC, 'Platform Clicks vs GA4 Sessions (Mexico, June)')
r = table(ws, r,
          ['Channel', 'Platform clicks', 'GA4 MX sessions', 'Capture',
           'Status'],
          [
              ['Bing Ads', 21480, '663 tagged (+ up to 2,451 unattributed)',
               '~7.6% tagged; ~36% incl. unattributed', 'Tracking gap: '
               'search account not linked in GA4'],
              ['Meta', 407136, 786, '0.19% (May: 0.54%)',
               'Anomaly, worsened; inventory quality'],
              ['Quantcast', 11284, 270, '2.4%', 'Expected range for '
               'display'],
              ['Azerion', 9910, '61 (+294 assumed placement)',
               '~3.6% combined', 'Expected range; geo pending vendor'],
          ],
          fmts=[None, INT, None, None, None])
r = para(ws, r, NC,
         'The largest June GA4 channel is Unassigned (2,853 sessions, 31%), '
         'driven by the unlinked search account. One admin-level fix '
         'recovers it; July reporting will then attribute paid search '
         'roughly 4x larger than June shows.')
r += 1
r = section(ws, r, NC, 'Search Delivery by Market (June)')
r = table(ws, r,
          ['Market', 'Spend', '% of spend', 'Note'],
          [
              ['Mexico (target)', 13021.58, 0.5075, 'June 3 exclusion '
               'already lifting the MX share'],
              ['Out-of-market (all other)', 12637.03, 0.4925,
               'Venezuela, Spain, US lead; fix = Mexico presence-only'],
              ['Total', 25658.61, 1.0, ''],
          ],
          fmts=[None, CUR, PCT2, None],
          total_last=True)
r += 1
r = section(ws, r, NC, 'Delivery Quality — Viewability')
r = table(ws, r,
          ['Line', 'June', 'May', 'Standard', 'Action'],
          [
              ['Quantcast', 0.5127, 0.671, 0.70, '49-site blocklist '
               'delivered (32% of June spend); set viewability floor'],
              ['Azerion', 0.6847, 0.716, 0.70, 'Push over 70%'],
          ],
          fmts=[None, PCT2, PCT2, PCT2, None])
r += 1
r = section(ws, r, NC, 'Organic Search (Search Console, Mexico)')
r = table(ws, r,
          ['Metric', 'January', 'June', 'Change'],
          [
              ['Organic clicks', 4811, 1210, '-75%'],
              ['Demo account page position', 12, 23, 'slid'],
              ['Trading academy page position', 14, 24, 'slid'],
              ['Homepage (brand) position', '—', 2.5, 'improved'],
          ],
          fmts=[None, INT, INT, None])
r = para(ws, r, NC,
         'GA4 shows the same decline independently, so this is ranking '
         'loss, not a tracking artifact. Two visible causes: dated news '
         'and analysis content aging out with nothing replacing it, and '
         'mid-funnel commercial pages sliding in rank while brand '
         'rankings improved. The decline spans every country the Spanish '
         'site serves. Recommendation: a dedicated SEO workstream.')
r += 1
r = section(ws, r, NC, 'Conversion Definitions (per channel)')
r = table(ws, r,
          ['Channel', 'Definition'],
          [
              ['Bing (SA360)', 'SA360 Primary conversions = submitted '
               'applications (live-account confirmation goal). SA360 is '
               'the source of truth; platform UI shows 0.'],
              ['Meta', 'Pixel custom events. Held: 67 of 86 June '
               'conversions come from one retargeting campaign with 126 '
               'landing-page views, not a credible rate.'],
              ['Quantcast', '11 view-through conversions (0 click-through). '
               'Directional only.'],
              ['Azerion', 'Result = submitted application (vendor-defined). '
               'Step 1 starts = 440.'],
          ])
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 40
for col in 'CDEFG':
    ws.column_dimensions[col].width = 34
for row_cells in ws.iter_rows(min_row=2):
    for c in row_cells:
        if c.value and isinstance(c.value, str) and len(c.value) > 40:
            c.alignment = WRAP

# ============================ DATA NOTES ====================================
ws = wb.create_sheet('Data Notes')
NC = 2
title_bar(ws, NC, 'Data Notes — Basis, Corrections, Rules')
r = 3
notes = [
    ('Currency / Timezone', 'USD, America/New_York. Comparison MoM vs May '
     '2026; April shown for trend where available.'),
    ('Conversions', 'Never summed and CPA never blended across channels; '
     'four different event definitions (see Diagnostics).'),
    ('Meta conversions', 'The 86 June platform conversions are held from '
     'reporting: 67 come from one retargeting campaign with 126 '
     'landing-page views, and platform-reported landing-page views exceed '
     'measured site sessions by an order of magnitude. Landing pages are '
     'confirmed tagged; the placement audit and pixel-event review '
     'complete before conversions are reinstated.'),
    ('Meta clicks', 'Link clicks (not all clicks); CTR is link CTR.'),
    ('Azerion spend basis', 'Raw vendor spend in all months (June '
     '$34,555.83; May $27,257.72). The May report stated Azerion at '
     '$29,302 on a fee-inclusive basis; comparisons here use the raw '
     'basis consistently.'),
    ('Quantcast May basis', 'May uses the client-reconciled $26,890 (not '
     'the platform-pulled $25,013.56). June is the platform Budget '
     'Delivered figure, $33,784.20.'),
    ('Search geo', '51% of June search spend served Mexico; 49% '
     '($12,637) served out-of-market users. The June 3 exclusion is '
     'active; the Mexico presence-only restriction completes the fix.'),
    ('Azerion geo', 'Country-level delivery confirmation requested from '
     'the vendor (early July); Mexico-only certification pending reply.'),
    ('SA360-GA4 link', 'The search account is not linked to GA4 property '
     '508849216, leaving 2,451 June Mexico sessions (27%) unattributed. '
     'Once linked, July reporting will restate paid search larger; we '
     'will flag the restatement.'),
    ('Revenue / ROAS', 'Not tracked on any channel; excluded from this '
     'report.'),
    ('Sources', 'SA360, Meta, Quantcast platform APIs; Azerion vendor '
     'workbook; GA4 property 508849216; Search Console. Pulled July 7, '
     '2026.'),
]
for k, v in notes:
    a = ws.cell(r, 1, k)
    a.font = F_LABEL
    a.fill = FILL_GRAY
    a.alignment = WRAP
    b = ws.cell(r, 2, v)
    b.font = F_BODY
    b.alignment = WRAP
    ws.row_dimensions[r].height = 13 * max(1, len(v) // 75 + 1) + 4
    r += 1
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 95
ws.sheet_view.showGridLines = False

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('Saved', os.path.abspath(OUT))
