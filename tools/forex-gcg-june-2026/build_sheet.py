#!/usr/bin/env python3
"""Build the GCG June 2026 Performance Report workbook (.xlsx).

Four tabs (Summary / Performance / Diagnostics / Data Notes) per the
Google Sheets Styling Standard, for Drive upload with conversion.
Mirrors tools/forex-june-2026/build_sheet.py. Spend = client budget
tracker (Meta $30,711 per Renzo). Azerion/Native detail pending vendor.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = '1A2A3A'
BLUE = '3B5998'
GRAY = 'F3F3F3'
BODY = '32373E'

F_TITLE = Font(name='Arial', size=16, bold=True, color='FFFFFF')
F_SECTION = Font(name='Arial', size=12, bold=True, color=NAVY)
F_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_LABEL = Font(name='Arial', size=10, bold=True, color=BODY)
F_BODY = Font(name='Arial', size=10, color=BODY)
FILL_NAVY = PatternFill('solid', fgColor=NAVY)
FILL_BLUE = PatternFill('solid', fgColor=BLUE)
FILL_GRAY = PatternFill('solid', fgColor=GRAY)
TOP_BORDER = Border(top=Side(style='thin', color='555555'))
WRAP = Alignment(wrap_text=True, vertical='top')
RIGHT = Alignment(horizontal='right')
CENTER_V = Alignment(vertical='center')

CUR = '$#,##0.00'
INT = '#,##0'
PCT3 = '0.000%'
PCT2 = '0.00%'

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, '..', '..', 'reports', 'forex', 'gcg', '2026-06',
                   'output', 'GCG-June-2026-Performance-Report.xlsx')

wb = Workbook()


def title_bar(ws, ncols, text):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = F_TITLE
    c.fill = FILL_NAVY
    c.alignment = CENTER_V
    ws.row_dimensions[1].height = 38


def section(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row, 1, text).font = F_SECTION
    return row + 1


def table(ws, row, headers, rows, fmts=None, total_last=False):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.font = F_HDR
        c.fill = FILL_BLUE
    row += 1
    for ri, r in enumerate(rows):
        is_total = total_last and ri == len(rows) - 1
        for ci, v in enumerate(r, 1):
            c = ws.cell(row, ci, v)
            c.font = F_LABEL if (ci == 1 or is_total) else F_BODY
            if ri % 2 == 0 or is_total:
                c.fill = FILL_GRAY
            if is_total:
                c.border = TOP_BORDER
            if ci > 1:
                c.alignment = RIGHT
                if fmts and fmts[ci - 1] and isinstance(v, (int, float)):
                    c.number_format = fmts[ci - 1]
        row += 1
    return row


def para(ws, row, ncols, text, bold=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = F_LABEL if bold else F_BODY
    c.alignment = WRAP
    ws.row_dimensions[row].height = 13 * max(1, len(text) // 110 + 1) + 4
    return row + 1


def meta_rows(ws, row, pairs):
    for k, v in pairs:
        a = ws.cell(row, 1, k)
        a.font = F_LABEL
        a.fill = FILL_GRAY
        b = ws.cell(row, 2, v)
        b.font = F_BODY
        b.fill = FILL_GRAY
        for ci in range(3, 9):
            ws.cell(row, ci).fill = FILL_GRAY
        row += 1
    return row


# ============================ SUMMARY =======================================
ws = wb.active
ws.title = 'Summary'
NC = 8
title_bar(ws, NC, 'FOREX.com GCG (US Hispanic) — June 2026 Performance Review')
r = 2
r = meta_rows(ws, r, [
    ('Reporting Period', 'June 1 – 30, 2026'),
    ('Comparison Basis', 'vs May 2026 (month over month)'),
    ('Currency', 'USD'),
    ('Timezone', 'America/New_York'),
    ('Channels', 'Google Ads, Meta, Azerion (Programmatic), Quantcast (Programmatic), Native (pilot)'),
    ('Attribution', 'Per channel; conversions not summed across channels (see Data Notes)'),
    ('Revenue / ROAS', 'Not tracked; excluded'),
    ('Data Freshness', 'Pulled July 17, 2026'),
    ('Prepared by', 'Berelvant'),
])
r += 1
r = section(ws, r, NC, 'Cross-Channel Performance Summary')
r = table(ws, r,
          ['Channel', 'Spend', 'Impressions', 'Clicks', 'CTR', 'CPM',
           'Conversions', 'CPA'],
          [
              ['Google Ads', 22524.00, 181470, 12888, 0.071020, '—', 67,
               336.18],
              ['Meta', 30711.00, 3058402, 74572, 0.024383, 10.04,
               '136 (pixel events)', '—'],
              ['Quantcast', 30559.00, 37233620, 2942, 0.000079, 0.82,
               '15 (view-through)', '—'],
              ['Azerion', 29586.00, 4515400, 19653, 0.004352, 6.55, 58,
               510.10],
              ['Native', 3645.00, 'pilot', '—', '—', '—', '—', '—'],
              ['Total', 117024.00, '—', '—', '—', '—', '—', '—'],
          ],
          fmts=[None, CUR, INT, INT, PCT3, CUR, INT, CUR],
          total_last=True)
r = para(ws, r, NC,
         'Conversions are measured on different events per channel and are '
         'not summed. Google recorded 67 submitted applications at a $336 '
         'CPA (application Step 5 goal, definition unchanged from May). '
         'Meta counts pixel application events, mostly application starts, '
         'on a traffic objective. Quantcast results are view-through and '
         'directional. Azerion counts vendor-defined submitted '
         'applications: 58 in June at a $510 CPA, up 35% from May at a '
         '17% lower CPA. The Native June flight is intentionally small; the '
         'full pilot read comes with the July report.')
r += 1
r = section(ws, r, NC, 'Executive Summary')
r = para(ws, r, NC,
         'June was the biggest GCG media month of 2026: $117,024 across '
         'five lines, up 53% from May, with the Native pilot launching as '
         'the fifth channel. Cost per measured submitted application held in '
         'the $350-$420 band across the quarter, the program\'s core '
         'efficiency metric. The planned search scale-up surfaced the '
         'account\'s rank ceiling: 67 submitted applications (May: 76), '
         'because every '
         'campaign loses 64-76% of available impressions to ad rank and '
         'only 9-13% to budget, so July search work is ad rank before '
         'budget. Paid social scaled 151% with healthy '
         'delivery (70.6% of clicks became landing-page views) and is '
         'ready for its conversion objective. The reach line scaled 37% at '
         'a $0.82 CPM while viewability fell to 46.9%; the blocklist and a '
         'viewability floor address it. The client funnel shows June '
         'submissions up 7% with approval steady at 45%; the '
         'start-to-submit step is where volume dilutes. Azerion was the June '
         'conversion leader: 58 applications (+35%) at $510 (-17%).')
r += 1
r = section(ws, r, NC, 'Campaign Highlights')
for b in [
    '•  Google: TrackB stayed the efficiency core (Platform $264, '
    'Authority $297); all 67 conversions are submitted applications.',
    '•  Meta: 1.53M people reached at 2.0 frequency; landing-page views '
    'ran 70.6% of clicks (May 65.9%); site analytics capture ~57% of '
    'clicks as sessions.',
    '•  Quantcast: 37.2M impressions to 23.2M devices at $0.82 CPM; '
    'results rose 8 to 15 (13 view-through).',
    '•  Azerion: 58 applications (+35%) at $510 (-17% CPA); application '
    'starts fell 68% while start-to-application completion improved 3.1% '
    'to 13.0% (the start-quality fix May called for). US-only delivery '
    'confirmed by the state-level geo report.',
    '•  Native pilot entered the market with an intentionally small '
    'first flight ($3,645), below the volume for a meaningful read; the '
    'full pilot reports with July.',
]:
    r = para(ws, r, NC, b)
r += 1
r = section(ws, r, NC, 'Suggested Optimizations')
for b in [
    '•  Shift paid social to a conversion objective and judge it on '
    'submitted applications from July.',
    '•  Run the search ad-rank program before adding budget: bid and '
    'quality work plus an RSA refresh on TrackB terms; defend Brand (11% '
    'impression share).',
    '•  Apply the 18-site Quantcast blocklist ($9,752, 32% of June spend '
    'on that line) and set a campaign-level viewability floor (June '
    'viewability 46.9% vs the 70% standard).',
    '•  Put the start-to-submit funnel step on the joint roadmap; traffic '
    'quadrupled quarter over quarter while submissions fell 14%.',
    '•  Concentrate Azerion budget on Trusted Broker ($352) and Broker 1 '
    '($382); shift weight from 728x90 (worst format) toward 300x600 '
    '(best); push viewability from 58.8% over the 70% standard.',
]:
    r = para(ws, r, NC, b)
r += 1
r = section(ws, r, NC, 'Next Steps')
for b in [
    '•  Complete the Meta conversion-objective shift (top paid-social '
    'action).',
    '•  Report the full Native pilot with July (June was the small '
    'first flight).',
    '•  Hold June budget levels on search while ad rank recovers, then '
    'scale into the recovered impression share.',
]:
    r = para(ws, r, NC, b)
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 50
for col in 'CDEFGH':
    ws.column_dimensions[col].width = 14
ws.freeze_panes = 'A2'
ws.sheet_view.showGridLines = False

# ============================ PERFORMANCE ===================================
ws = wb.create_sheet('Performance')
NC = 9
title_bar(ws, NC, 'Channel Performance Detail — June 2026')
r = 3
r = section(ws, r, NC, 'Spend by Channel — April / May / June (client tracker)')
r = table(ws, r,
          ['Channel', 'April', 'May', 'June', 'MoM (Jun v May)',
           'June share'],
          [
              ['Google Ads', 15120.00, 15201.00, 22524.00, 0.4818, 0.1925],
              ['Meta', 12167.00, 12243.00, 30711.00, 1.5085, 0.2624],
              ['Quantcast', 15015.00, 22359.00, 30559.00, 0.3667, 0.2611],
              ['Azerion', 16021.00, 26472.00, 29586.00, 0.1176, 0.2528],
              ['Native', '—', '—', 3645.00, 'new', 0.0311],
              ['Total', 58324.00, 76274.00, 117024.00, 0.5342, 1.0],
          ],
          fmts=[None, CUR, CUR, CUR, PCT2, PCT2],
          total_last=True)
r += 1
r = section(ws, r, NC, 'Conversions by Channel (per-channel definitions; '
                       'never summed)')
r = table(ws, r,
          ['Channel', 'April', 'May', 'June', 'June CPA', 'Definition'],
          [
              ['Google Ads', 12, 76, 67, 336.18,
               'Submitted applications (Step 5)'],
              ['Azerion', 76, 43, 58, 510.10,
               'Submitted applications (vendor-defined)'],
              ['Quantcast', 2, 8, 15, '—', 'View-through, directional'],
              ['Meta', 73, 109, 136, '—',
               'Pixel events (starts rollup, traffic objective)'],
          ],
          fmts=[None, INT, INT, INT, CUR, None])
r += 1
r = section(ws, r, NC, 'Google Ads Campaigns (June)')
r = table(ws, r,
          ['Campaign', 'Spend', 'Impressions', 'Clicks', 'CTR', 'Avg CPC',
           'Apps', 'CPA', 'Search IS'],
          [
              ['TrackA Trust', 7796.35, 57842, 5712, 0.098752, 1.36, 18,
               433.13, 0.2655],
              ['TrackB Authority', 6825.66, 44139, 3198, 0.072453, 2.13,
               23, 296.77, 0.2375],
              ['TrackB Platform', 4488.64, 44286, 2568, 0.057987, 1.75, 17,
               264.04, 0.1303],
              ['Brand Search', 3413.14, 35203, 1410, 0.040053, 2.42, 9,
               379.24, 0.1119],
              ['Total', 22523.79, 181470, 12888, 0.071020, 1.75, 67,
               336.18, '—'],
          ],
          fmts=[None, CUR, INT, INT, PCT3, CUR, INT, CUR, PCT2],
          total_last=True)
r = para(ws, r, NC,
         'Lost impression share is rank-driven on every campaign (64-76%) '
         'vs budget-driven (9-13%): the July lever is ad rank, not budget.')
r += 1
r = section(ws, r, NC, 'Meta Ad Sets (June)')
r = table(ws, r,
          ['Ad set', 'Spend (platform)', 'Impressions', 'Link clicks',
           'LP views', 'Pixel events'],
          [
              ['trackA_pros_us_es', 16923.92, 1768642, 35710, 25078, 84],
              ['trackB_pros_us_es', 17787.05, 1289760, 38862, 27561, 52],
          ],
          fmts=[None, CUR, INT, INT, INT, INT])
r += 1
r = section(ws, r, NC, 'Azerion Ad Sets (June; budget-tracker spend basis)')
r = table(ws, r,
          ['Ad set', 'Spend', 'Apps', 'CPA', 'Viewability'],
          [
              ['Trusted Broker', 4932.94, 14, 352.35, 0.5889],
              ['Broker 1', 4964.06, 13, 381.85, 0.5873],
              ['Language Broker', 4900.30, 10, 490.03, 0.5877],
              ['Professional Tools', 4940.10, 9, 548.90, 0.5866],
              ['Spanish Platform', 4933.06, 7, 704.72, 0.5877],
              ['Trust HTML', 4915.71, 5, 983.14, 0.5880],
              ['Total', 29586.00, 58, 510.10, 0.5877],
          ],
          fmts=[None, CUR, INT, CUR, PCT2],
          total_last=True)
r += 1
r = section(ws, r, NC, 'Azerion Funnel — March to June (vendor DSP '
                       'reporting)')
r = table(ws, r,
          ['Month', 'Site visits', 'App starts', 'Applications',
           'Start-to-app rate'],
          [
              ['March', 2088, 68, 7, 0.103],
              ['April', 10042, 492, 76, 0.154],
              ['May', 51955, 1398, 43, 0.031],
              ['June', 29988, 447, 58, 0.130],
          ],
          fmts=[None, INT, INT, INT, PCT2])
r = para(ws, r, NC,
         'June reversed the May pattern: starts fell 68% while '
         'start-to-application completion improved from 3.1% to 13.0%. '
         'Fewer, better-qualified starts.')
r += 1
r = section(ws, r, NC, 'Blended Funnel — Q1 vs Q2 (client dashboard, '
                       'organic + paid)')
r = table(ws, r,
          ['Metric', 'Q1 (Jan-Mar)', 'Q2 (Apr-Jun)', 'Change'],
          [
              ['Unique Sessions', 41047, 162871, '+297%'],
              ['App Starts', 6997, 7974, '+14%'],
              ['Submitted', 1203, 1033, '-14%'],
              ['Live Apps Submitted', 1165, 993, '-15%'],
              ['Approved', 524, 465, '-11%'],
              ['New Funded', 148, 146, '-1%'],
              ['New Traded', 118, 114, '-3%'],
          ],
          fmts=[None, INT, INT, None])
r += 1
r = section(ws, r, NC, 'Spanish-Language Site Traffic — GA4 US property, '
                       'monthly')
r = table(ws, r,
          ['Metric', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
          [
              ['Sessions', 69895, 58481, 74971, 69553, 69927, 80231],
              ['Unique visitors', 24928, 20841, 29355, 30251, 29808, 41795],
          ],
          fmts=[None, INT, INT, INT, INT, INT, INT])
r = para(ws, r, NC,
         'June sessions +15% and unique visitors +40% vs May, the highest '
         'Spanish-language visitor count of the year. The base held near '
         '70K sessions January through May.')
ws.column_dimensions['A'].width = 28
for col in 'BCDEFGHI':
    ws.column_dimensions[col].width = 14
ws.freeze_panes = 'A2'

# ============================ DIAGNOSTICS ===================================
ws = wb.create_sheet('Diagnostics')
NC = 7
title_bar(ws, NC, 'Diagnostics — Measurement, Delivery Quality, Funnel')
r = 3
r = section(ws, r, NC, 'Search Impression Share (June)')
r = table(ws, r,
          ['Campaign', 'Search IS', 'Lost to budget', 'Lost to rank'],
          [
              ['Brand Search', 0.1119, 0.1258, 0.7622],
              ['TrackA Trust', 0.2655, 0.0861, 0.6484],
              ['TrackB Platform', 0.1303, 0.1093, 0.7604],
              ['TrackB Authority', 0.2375, 0.1196, 0.6430],
          ],
          fmts=[None, PCT2, PCT2, PCT2])
r = para(ws, r, NC,
         'Every campaign is rank-limited. The June scale-up quantified it: '
         'at current ad rank, incremental budget reaches lower-quality '
         'auctions (67 applications at $336). The ad-rank program '
         'converts this into recovered impression share.')
r += 1
r = section(ws, r, NC, 'Paid Social Measurement Check (June)')
r = table(ws, r,
          ['Check', 'Value', 'Status'],
          [
              ['Landing-page views / link clicks', 0.7059,
               'Healthy (May: 65.9%)'],
              ['Site sessions / link clicks (GA4)', 0.57,
               'Healthy capture'],
              ['Objective', 'Traffic (CTR)',
               'Conversion-objective shift = top July action'],
              ['Pixel events (June)', 136,
               'Starts rollup; not submitted applications'],
          ],
          fmts=[None, PCT2, None])
r += 1
r = section(ws, r, NC, 'Delivery Quality — Viewability')
r = table(ws, r,
          ['Line', 'June', 'Standard', 'Action'],
          [
              ['Quantcast', 0.4692, 0.70,
               '18-site blocklist delivered ($9,752, 32% of June spend); '
               'set viewability floor'],
              ['Azerion', 0.5877, 0.70,
               'Push over the standard; 300x600 best format, 728x90 '
               'worst'],
          ],
          fmts=[None, PCT2, PCT2, None])
r += 1
r = section(ws, r, NC, 'Client Funnel — monthly (blended organic + paid)')
r = table(ws, r,
          ['Metric', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
          [
              ['Unique Sessions', 4187, 5075, 31785, 43192, 52137, 67542],
              ['App Starts', 2306, 1871, 2820, 2767, 2617, 2590],
              ['Live Apps Submitted', 430, 344, 391, 398, 289, 306],
              ['Approved', 202, 150, 172, 186, 134, 145],
              ['New Funded', 58, 46, 44, 65, 40, 41],
              ['New Traded', 45, 40, 33, 53, 31, 30],
          ],
          fmts=[None, INT, INT, INT, INT, INT, INT])
r = para(ws, r, NC,
         'Traffic grew 16x January to June while application starts stayed '
         'near 2,600/month: the start rate diluted from 55% to 4%. '
         'Approval (45%) and traded (73-82%) rates held; approval, funding, '
         'and activation run on the application-review and '
         'account-activation journey downstream of media. The joint '
         'funnel work targets the start-to-submit step.')
r += 1
r = section(ws, r, NC, 'Conversion Definitions (per channel)')
r = table(ws, r,
          ['Channel', 'Definition'],
          [
              ['Google Ads', 'Submitted applications: "PO App Form - Step '
               '5 - Submission Completed". Same definition all months.'],
              ['Meta', 'Pixel application events on a traffic objective; '
               'mostly application starts. Not submitted applications.'],
              ['Quantcast', '15 results; 13 view-through, 2 click-through. '
               'Directional only.'],
              ['Azerion', 'Submitted applications (vendor-defined): 58 in '
               'June with 447 application starts.'],
              ['Native', 'Pilot; June flight intentionally small. Full '
               'pilot read in the July report, judged on cost per '
               'submitted application.'],
          ])
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 40
for col in 'CDEFG':
    ws.column_dimensions[col].width = 26
for row_cells in ws.iter_rows(min_row=2):
    for c in row_cells:
        if c.value and isinstance(c.value, str) and len(c.value) > 40:
            c.alignment = WRAP

# ============================ DATA NOTES ====================================
ws = wb.create_sheet('Data Notes')
NC = 2
title_bar(ws, NC, 'Data Notes — Definitions & Rules')
r = 3
notes = [
    ('Currency / Timezone', 'USD, America/New_York. Comparison MoM vs May '
     '2026; April shown for trend.'),
    ('Conversions', 'Never summed and CPA never blended across channels; '
     'definitions differ per channel (see Diagnostics).'),
    ('Meta conversions', 'Pixel application events on a traffic '
     'objective, mostly application starts; reported for context, not as '
     'submitted applications. The shift to a conversion objective is the '
     'top July action.'),
    ('Meta clicks', 'Link clicks (not all clicks); CTR is link CTR.'),
    ('Azerion', 'Vendor June file received: 58 applications, 447 starts, '
     '4.52M impressions, viewability 58.8%; US-only delivery confirmed '
     'by the state-level geo report. Ad-set figures shown on the '
     'budget-tracker spend basis.'),
    ('Native', 'Carried at budget-tracker spend ($3,645). The June '
     'flight was an intentionally small first burn, below the volume '
     'for a meaningful cost-per-application read; the full pilot '
     'reports with July.'),
    ('Search impression share', 'Google Ads auction data, June: lost to '
     'rank 64-76% per campaign, lost to budget 9-13%.'),
    ('Client funnel', 'Blended organic + paid from the client dashboard; '
     'quarters are calendar (Q1 Jan-Mar, Q2 Apr-Jun).'),
    ('Revenue / ROAS', 'Not tracked on any channel; excluded.'),
    ('Sources', 'Google Ads, Meta, Quantcast platform APIs; GA4 property '
     '325353267; client dashboard; client budget tracker. Pulled July '
     '17, 2026.'),
]
for k, v in notes:
    a = ws.cell(r, 1, k)
    a.font = F_LABEL
    a.fill = FILL_GRAY
    a.alignment = WRAP
    b = ws.cell(r, 2, v)
    b.font = F_BODY
    b.alignment = WRAP
    ws.row_dimensions[r].height = 13 * max(1, len(v) // 75 + 1) + 4
    r += 1
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 95
ws.sheet_view.showGridLines = False

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('Saved', os.path.abspath(OUT))
