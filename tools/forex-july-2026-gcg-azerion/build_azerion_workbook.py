"""
Build GCG (US Hispanic) Azerion July 2026 data workbook (Display + Native).
Source: vendor XLSX in reports/forex/gcg/2026-07/data/sources/ (received before
2026-08-19), transcribed as-is. Reporting only.
Applications are vendor-reported (partner methodology), interpreted
independently — never blended with Google or any other system.
Tech fee note: raw + 7.5% is INTERNAL ONLY (billing); client-facing spend is
the client tracker line.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
SUBHEAD_FILL = "D9E1F2"
SECTION_FILL = "2E5496"
ROW_ALT_FILL = "F2F5FB"
GRAY_TEXT = "404040"
BORDER_COLOR = "BFBFBF"
thin = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def title_row(ws, text, row=1, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)


def subtitle_row(ws, text, row=2, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10, color=GRAY_TEXT)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SUBHEAD_FILL)


def section_row(ws, text, row, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SECTION_FILL)


def bullet_row(ws, text, row, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30


def header_row(ws, headers, row):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def data_row(ws, values, row, formats=None, alt=False, total=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if total:
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=SECTION_FILL)
        else:
            c.font = Font(name="Calibri", size=10)
            if alt:
                c.fill = PatternFill("solid", fgColor=ROW_ALT_FILL)
        if formats and formats.get(i):
            c.number_format = formats[i]


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


MONEY = "#,##0.00"
INT = "#,##0"
PCT = "0.00%"
DEC = "0.00"

# Display ad sets: (name, spend, impr, clicks, results, cpa, viewable, viewability)
DISPLAY = [
    ("Spanish Platform", 7310.778, 1218463, 2427, 9, 812.31, 813703, 0.6678),
    ("Language Broker", 6727.632, 1121272, 2164, 16, 420.48, 747490, 0.6666),
    ("Broker 1", 4112.616, 685436, 1064, 13, 316.36, 445712, 0.6503),
    ("Professional Tools", 3553.650, 592275, 1189, 18, 197.43, 368978, 0.6230),
    ("Trust HTML", 3478.548, 579758, 893, 13, 267.58, 366774, 0.6326),
    ("Trusted Broker", 3432.162, 572027, 1155, 11, 312.01, 354907, 0.6204),
]
D_SPEND = sum(x[1] for x in DISPLAY)      # 28,615.386
D_IMPR = sum(x[2] for x in DISPLAY)
D_CLICKS = sum(x[3] for x in DISPLAY)
D_RESULTS = sum(x[4] for x in DISPLAY)    # 80
D_VIEWABLE = sum(x[6] for x in DISPLAY)

NATIVE = [
    ("Trader_laptop_thinking", 1598.184, 199773, 132, 151858, 0.7602),
    ("Multiple_screens_graph", 1597.896, 199737, 125, 144426, 0.7231),
    ("Mobile_desktop_view", 1576.288, 197036, 143, 151199, 0.7674),
    ("Phone_closeup", 1569.984, 196248, 139, 150946, 0.7692),
    ("City_view_MHTN", 1516.944, 189618, 143, 126602, 0.6677),
    ("Third_person_perspective_mobile", 1502.744, 187843, 134, 125510, 0.6682),
]
N_SPEND = sum(x[1] for x in NATIVE)       # 9,362.04
N_IMPR = sum(x[2] for x in NATIVE)
N_CLICKS = sum(x[3] for x in NATIVE)
N_VIEWABLE = sum(x[4] for x in NATIVE)

# ---------------------------------------------------------------------------
# TAB 1: Summary
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Summary")
title_row(ws, "GCG (US Hispanic) — Azerion Display + Native — July 2026")
subtitle_row(ws, "Vendor XLSX (Azerion, account 9969644, US geo filter) · raw vendor spend; client-facing spend = client tracker · apps are vendor-reported, read independently")

r = 4
section_row(ws, "DISPLAY HEADLINE", r); r += 1
for t in [
    f"•  ${D_SPEND:,.2f} raw vendor spend (tracker line $31,477), {D_RESULTS} vendor-reported applications — the highest Azerion application month to date (June 58, +37.9%). Vendor CPA $357.69 raw; on the tracker basis $393.46 vs June's $510 (-22.8%).",
    "•  Efficiency inverted the spend ranking: Professional Tools leads at $197.43 CPA on 18 apps, while the biggest line (Spanish Platform, $7,310.78) produced 9 apps at $812.31. Trust HTML improved from June's $983 laggard to $267.58.",
    f"•  Viewability: {D_VIEWABLE/D_IMPR:.1%} computed from vendor rows (viewable/served) — BELOW the 70% IAB floor. The vendor summary tab claims 71.28%; the discrepancy is flagged in QA and the computed figure is used.",
    "•  Week of Jul 22-28 was the efficiency peak: 20 apps at $144.76 on reduced spend, evidence that tighter delivery converts better.",
]:
    bullet_row(ws, t, r); r += 1
r += 1
section_row(ws, "NATIVE HEADLINE", r); r += 1
for t in [
    f"•  First full Azerion Native month: ${N_SPEND:,.2f} raw, ramped from near-zero in week 1 to $3,402.66 in the final partial week. {N_IMPR:,} impressions, {N_CLICKS} clicks (0.070% CTR), viewability {N_VIEWABLE/N_IMPR:.1%} — above the 70% floor.",
    "•  Upper-funnel line: no conversion tracking on Native in July; judged on delivery and viewability. Vendor's own next step: switch optimization toward CTR.",
    "•  Tracker 'Native' line is $20,298; Azerion Native accounts for roughly $10.1K of it with the internal fee applied. The remainder is expected to be Quantcast Native — BLOCKED this session (Quantcast MCP credential, see KNOWN-BUGS) — reconcile before the Native line is client-ready.",
]:
    bullet_row(ws, t, r); r += 1
set_widths(ws, [112] + [12] * 9)

# ---------------------------------------------------------------------------
# TAB 2: Display Ad Sets
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Display Ad Sets")
title_row(ws, "Azerion Display — Ad Sets (July)")
subtitle_row(ws, "Vendor-reported. Results = applications per Azerion methodology. Sorted by spend.")
headers = ["Ad set", "Spend (raw)", "Impr", "Clicks", "CTR", "Apps", "CPA (raw)", "Viewable imps", "Viewability"]
r = 4
header_row(ws, headers, r); r += 1
f = {2: MONEY, 3: INT, 4: INT, 5: PCT, 6: INT, 7: MONEY, 8: INT, 9: PCT}
for i, x in enumerate(DISPLAY):
    data_row(ws, [x[0], x[1], x[2], x[3], x[3] / x[2], x[4], x[5], x[6], x[7]], r, f, alt=i % 2 == 1)
    r += 1
data_row(ws, ["TOTAL", D_SPEND, D_IMPR, D_CLICKS, D_CLICKS / D_IMPR, D_RESULTS, D_SPEND / D_RESULTS, D_VIEWABLE, D_VIEWABLE / D_IMPR], r, f, total=True)
r += 2
bullet_row(ws, "June carry-over check (concentrate on Trusted Broker + Broker 1): both held efficient ($312 / $316), but July's actual leader is Professional Tools ($197, 18 apps). Spanish Platform is the reallocation candidate: highest spend, worst CPA. No creative-format (728x90 vs 300x600) breakdown in this month's vendor file — June's format recommendation cannot be verified; ask Azerion to restore the format cut.", r)
set_widths(ws, [22, 12, 11, 9, 8, 7, 10, 12, 11])

# ---------------------------------------------------------------------------
# TAB 3: Display Weekly
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Display Weekly")
title_row(ws, "Azerion Display — Weekly Trend (July)")
subtitle_row(ws, "Vendor weekly rows, transcribed. Week of Jul 22-28: spend cut ~64%, best CPA of the month.")
headers = ["Week", "Spend", "Impr", "Clicks", "Apps", "CPA", "Viewability"]
r = 4
header_row(ws, headers, r); r += 1
WEEKS = [
    ("Jul 1-7", 8829.168, 1471528, 2605, 23, 383.88, 0.6350),
    ("Jul 8-14", 7955.988, 1325998, 2179, 15, 530.40, 0.6422),
    ("Jul 15-21", 8033.274, 1338879, 2466, 22, 365.15, 0.6603),
    ("Jul 22-28", 2895.144, 482524, 1452, 20, 144.76, 0.6823),
    ("Jul 29-31", 901.812, 150302, 190, 0, None, 0.6534),
]
f = {2: MONEY, 3: INT, 4: INT, 5: INT, 6: MONEY, 7: PCT}
for i, x in enumerate(WEEKS):
    data_row(ws, list(x), r, f, alt=i % 2 == 1)
    r += 1
set_widths(ws, [12, 12, 11, 9, 7, 10, 11])

# ---------------------------------------------------------------------------
# TAB 4: Native
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Native")
title_row(ws, "Azerion Native — Creatives & Weekly (July)")
subtitle_row(ws, "First full month. Upper-funnel, no conversion tracking; judged on delivery + viewability. Spend nearly flat across 6 creatives by design.")
headers = ["Creative", "Spend (raw)", "Impr", "Clicks", "CTR", "Viewable imps", "Viewability"]
r = 4
header_row(ws, headers, r); r += 1
f = {2: MONEY, 3: INT, 4: INT, 5: PCT, 6: INT, 7: PCT}
for i, x in enumerate(NATIVE):
    data_row(ws, [x[0], x[1], x[2], x[3], x[3] / x[2], x[4], x[5]], r, f, alt=i % 2 == 1)
    r += 1
data_row(ws, ["TOTAL", N_SPEND, N_IMPR, N_CLICKS, N_CLICKS / N_IMPR, N_VIEWABLE, N_VIEWABLE / N_IMPR], r, f, total=True)
r += 2
section_row(ws, "WEEKLY RAMP", r); r += 1
headers2 = ["Week", "Spend", "Impr", "Clicks", "Viewability"]
header_row(ws, headers2, r); r += 1
NWEEKS = [
    ("Jul 1-7", 0.19, 24, 0, 0.0),
    ("Jul 8-14", 273.30, 34163, 112, 0.6874),
    ("Jul 15-21", 963.96, 120495, 145, 0.6649),
    ("Jul 22-28", 4722.12, 590265, 328, 0.7186),
    ("Jul 29-31", 3402.66, 425332, 231, 0.7734),
]
f2 = {2: MONEY, 3: INT, 4: INT, 5: PCT}
for i, x in enumerate(NWEEKS):
    data_row(ws, list(x), r, f2, alt=i % 2 == 1)
    r += 1
r += 1
bullet_row(ws, "Viewability improved as scale grew (68.7% → 77.3% weekly), opposite of the display line. Clicks did not follow scale: CTR fell from 0.33% (wk2) to 0.054% (final week) — the vendor's optimize-to-CTR proposal responds to this.", r)
set_widths(ws, [30, 12, 11, 9, 8, 12, 11])

# ---------------------------------------------------------------------------
# TAB 5: Notes & QA
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Notes & QA")
title_row(ws, "Azerion — Notes, QA & Reconciliation", span=1)
r = 3
for t in [
    "SOURCE",
    "Two vendor XLSX files (display + native), Azerion account 9969644, US geo filter, July 1-31 2026, USD, TZ America/New_York. Clicks are IVT-filtered per vendor notes. Transcribed as-is; raw files untouched in data/sources/.",
    "",
    "RECONCILIATION (internal only — never client-facing)",
    "Display raw $28,615.39 x 1.075 tech fee = $30,761.54 vs tracker Azerion $31,477: delta +$715.46. Same pattern as June (raw + fee + adjustment); tracker stands for all client materials, delta stays internal.",
    "Native raw $9,362.04 x 1.075 = $10,064.19 vs tracker Native $20,298: gap $10,233.81 — expected to be Quantcast Native, unverifiable this session (Quantcast MCP blocked, KNOWN-BUGS 2026-08-19). The Native line is NOT client-ready until reconciled.",
    "Native file internal inconsistency: Summary KPI spend 9,362.232 vs Performance total 9,362.04 (delta $0.19, immaterial; performance rows used).",
    "",
    "VIEWABILITY DISCREPANCY (QA flag)",
    "Display Summary tab claims 71.28% viewability; the Performance rows compute 64.95% (3,097,564 viewable / 4,769,231 served) and no ad set exceeds 66.8%. The computed figure is used everywhere. 64.95% is BELOW the 70% IAB floor — flag to vendor; June carry-over 'push viewability over the standard' is NOT met on display. Native computes 72.68%, above the floor.",
    "",
    "COUNTING RULES",
    "Applications and CPA are Azerion partner-reported and interpreted independently (permanent reporting rule). Never summed with Google submitted apps. Tracker-basis CPA for client materials: $31,477 / 80 = $393.46.",
]:
    if t and t == t.upper() and len(t) < 60:
        section_row(ws, t, r, span=1)
    else:
        bullet_row(ws, t, r, span=1)
        ws.row_dimensions[r].height = 30 if t else 6
    r += 1
ws.column_dimensions["A"].width = 130

out = "reports/forex/gcg/2026-07/data/GCG-Azerion-July-2026-data.xlsx"
wb.save(out)
print(f"saved {out}")

assert abs(D_SPEND - 28615.386) < 0.01
assert D_RESULTS == 80
assert abs(N_SPEND - 9362.04) < 0.01
assert D_IMPR == 4769231 and N_IMPR == 1170255
print("self-check OK: display and native totals match vendor rows")
